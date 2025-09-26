# Unified Excel Reporting System - Combines Original + Enhanced Reporting
# This module merges the comprehensive analysis from both reporting systems

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference, BarChart, LineChart
from datetime import datetime
import os
import logging

logger = logging.getLogger(__name__)

def create_comprehensive_excel_report(container_results, file_share_results, 
                                    storage_data=None, recommendations=None,
                                    cost_analysis=None, reservations_analysis=None, 
                                    savings_plans_analysis=None):
    """
    Create the most comprehensive Excel report combining all analysis types
    
    Args:
        container_results: Blob storage analysis results
        file_share_results: Azure Files analysis results  
        storage_data: Original storage analysis data format
        recommendations: Cost optimization recommendations
        cost_analysis: Monthly spending analysis and trends
        reservations_analysis: Reserved Instance recommendations
        savings_plans_analysis: Savings Plans recommendations
        
    Returns:
        str: Path to generated Excel file
    """
    logger.info("Creating comprehensive Excel report with all analysis types...")
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create all available sheets - comprehensive coverage
    sheets = {}
    
    # 1. Executive Summary (Enhanced) - High-level KPIs for executives
    sheets['Executive Summary'] = wb.create_sheet("Executive Summary")
    
    # 2. Summary (Original) - Detailed statistics overview  
    sheets['Summary'] = wb.create_sheet("Summary")
    
    # 3. Monthly Spending - Cost trends and financial analysis
    if cost_analysis:
        sheets['Monthly Spending'] = wb.create_sheet("Monthly Spending")
    
    # 4. Blob Storage Analysis (Original) - Detailed container/blob analysis
    if container_results or storage_data:
        sheets['Blob Storage Analysis'] = wb.create_sheet("Blob Storage Analysis")
    
    # 5. Azure Files Analysis (Original) - Detailed file shares analysis  
    if file_share_results:
        sheets['Azure Files Analysis'] = wb.create_sheet("Azure Files Analysis")
    
    # 6. Storage Analysis (Enhanced) - Combined storage overview
    sheets['Storage Analysis'] = wb.create_sheet("Storage Analysis")
    
    # 7. Recommendations (Original) - Basic cost optimization
    if recommendations:
        sheets['Recommendations'] = wb.create_sheet("Recommendations")
    
    # 8. Cost Optimization (Enhanced) - Advanced financial recommendations
    sheets['Cost Optimization'] = wb.create_sheet("Cost Optimization")
    
    # 9. Reservations - RI opportunities
    if reservations_analysis:
        sheets['Reservations'] = wb.create_sheet("Reservations")
    
    # 10. Savings Plans - Savings plan opportunities
    if savings_plans_analysis:
        sheets['Savings Plans'] = wb.create_sheet("Savings Plans")
    
    # 11. Detailed Data - Raw data for analysis
    sheets['Detailed Data'] = wb.create_sheet("Detailed Data")
    
    # Populate sheets with data
    try:
        # Executive Summary (Enhanced version)
        try:
            if cost_analysis or reservations_analysis or savings_plans_analysis:
                from .enhanced_reporting import _create_executive_summary_sheet
                _create_executive_summary_sheet(
                    sheets['Executive Summary'], 
                    container_results, file_share_results,
                    cost_analysis, reservations_analysis, savings_plans_analysis
                )
        except Exception as e:
            logger.warning(f"Failed to create Executive Summary sheet: {e}")
        
        # Summary (Original detailed version)
        try:
            if storage_data and recommendations:
                _create_original_summary_sheet(sheets['Summary'], storage_data, recommendations)
        except Exception as e:
            logger.warning(f"Failed to create Summary sheet: {e}")
        
        # Monthly Spending
        try:
            if cost_analysis and 'Monthly Spending' in sheets:
                from .enhanced_reporting import _create_monthly_spending_sheet
                _create_monthly_spending_sheet(sheets['Monthly Spending'], cost_analysis)
        except Exception as e:
            logger.warning(f"Failed to create Monthly Spending sheet: {e}")
        
        # Blob Storage Analysis (Original detailed version)
        try:
            if 'Blob Storage Analysis' in sheets:
                if storage_data:
                    _create_original_blob_analysis_sheet(sheets['Blob Storage Analysis'], storage_data)
                elif container_results:
                    _create_container_results_analysis_sheet(sheets['Blob Storage Analysis'], container_results)
        except Exception as e:
            logger.warning(f"Failed to create Blob Storage Analysis sheet: {e}")
        
        # Azure Files Analysis (Original detailed version)
        try:
            if 'Azure Files Analysis' in sheets and file_share_results:
                _create_original_files_analysis_sheet(sheets['Azure Files Analysis'], file_share_results)
        except Exception as e:
            logger.warning(f"Failed to create Azure Files Analysis sheet: {e}")
        
        # Storage Analysis (Enhanced combined version)
        try:
            from .enhanced_reporting import _create_storage_analysis_sheet
            _create_storage_analysis_sheet(sheets['Storage Analysis'], container_results, file_share_results)
        except Exception as e:
            logger.warning(f"Failed to create Storage Analysis sheet: {e}")
        
        # Recommendations (Original version)
        try:
            if recommendations and 'Recommendations' in sheets:
                _create_original_recommendations_sheet(sheets['Recommendations'], recommendations)
        except Exception as e:
            logger.warning(f"Failed to create Recommendations sheet: {e}")
        
        # Cost Optimization (Enhanced version)
        try:
            from .enhanced_reporting import _create_cost_optimization_sheet
            _create_cost_optimization_sheet(
                sheets['Cost Optimization'], 
                cost_analysis, reservations_analysis, savings_plans_analysis
            )
        except Exception as e:
            logger.warning(f"Failed to create Cost Optimization sheet: {e}")
        
        # Reservations
        try:
            if reservations_analysis and 'Reservations' in sheets:
                from .enhanced_reporting import _create_reservations_sheet
                _create_reservations_sheet(sheets['Reservations'], reservations_analysis)
        except Exception as e:
            logger.warning(f"Failed to create Reservations sheet: {e}")
        
        # Savings Plans  
        try:
            if savings_plans_analysis and 'Savings Plans' in sheets:
                from .enhanced_reporting import _create_savings_plans_sheet
                _create_savings_plans_sheet(sheets['Savings Plans'], savings_plans_analysis)
        except Exception as e:
            logger.warning(f"Failed to create Savings Plans sheet: {e}")
        
        # Detailed Data
        try:
            from .enhanced_reporting import _create_detailed_data_sheet
            _create_detailed_data_sheet(sheets['Detailed Data'], container_results, file_share_results)
        except Exception as e:
            logger.warning(f"Failed to create Detailed Data sheet: {e}")
        
        # Add watermarks and auto-adjust column widths for all sheets
        watermark_text = f"Azure FinOps Analysis Report - Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        for sheet_name, sheet in sheets.items():
            _add_watermark_to_sheet(sheet, watermark_text)
            try:
                _auto_adjust_column_widths(sheet)
            except Exception as e:
                logger.warning(f"Failed to auto-adjust columns for {sheet_name}: {e}")
        
    except Exception as e:
        logger.error(f"Error populating report sheets: {e}")
        logger.info("Continuing with available data...")
    
    # Save the comprehensive report
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"azure_finops_comprehensive_analysis_{timestamp}.xlsx"
    abs_path = os.path.abspath(filename)
    
    wb.save(abs_path)
    logger.info(f"Comprehensive Excel report saved to: {abs_path}")
    print(f"\n📊 Comprehensive Azure FinOps Excel report written to: {abs_path}")
    print(f"📋 Report includes {len(sheets)} analysis sheets:")
    for sheet_name in sheets.keys():
        print(f"   • {sheet_name}")
    
    # Print data summary
    print(f"\n📊 Data Summary:")
    print(f"   • Container Results: {len(container_results)} containers")
    print(f"   • File Share Results: {len(file_share_results)} shares")
    if storage_data:
        print(f"   • Storage Accounts: {len(storage_data.get('storage_accounts', []))}")
    print(f"   • Recommendations: {len(recommendations) if recommendations else 0}")
    
    return abs_path


def _create_original_summary_sheet(sheet, storage_data, recommendations):
    """Create simple summary sheet"""
    # Header
    sheet['A1'] = "Azure Storage Analysis Summary"
    sheet['A1'].font = Font(bold=True, size=16)
    
    if not storage_data:
        sheet['A3'] = "No storage data available"
        return
    
    # Basic summary information
    row = 3
    if isinstance(storage_data, dict):
        sheet[f'A{row}'] = "Subscription ID:"
        sheet[f'B{row}'] = storage_data.get('subscription_id', 'Unknown')
        row += 1
        
        sheet[f'A{row}'] = "Storage Accounts:"
        sheet[f'B{row}'] = len(storage_data.get('storage_accounts', []))
        row += 1
        
        sheet[f'A{row}'] = "Total Cost (90 days):"
        sheet[f'B{row}'] = f"${storage_data.get('total_cost', 0):.2f}"
        row += 1
        
        sheet[f'A{row}'] = "Analysis Date:"
        sheet[f'B{row}'] = storage_data.get('analysis_date', 'Unknown')
        row += 1
    
    # Recommendations summary
    if recommendations:
        row += 2
        sheet[f'A{row}'] = "Recommendations:"
        sheet[f'A{row}'].font = Font(bold=True)
        row += 1
        
        for i, rec in enumerate(recommendations, 1):
            sheet[f'A{row}'] = f"{i}. {rec.get('title', 'Unknown')}"
            sheet[f'B{row}'] = rec.get('priority', 'Unknown')
            row += 1
    
    # Auto-adjust column widths
    _auto_adjust_column_widths(sheet)


def _create_original_blob_analysis_sheet(sheet, storage_data):
    """Create simple blob analysis sheet from storage_data"""
    # Header
    sheet['A1'] = "Blob Storage Analysis (Storage Data)"
    sheet['A1'].font = Font(bold=True, size=16)
    
    if not storage_data or not isinstance(storage_data, dict):
        sheet['A3'] = "No storage data available"
        return
    
    storage_accounts = storage_data.get('storage_accounts', [])
    if not storage_accounts:
        sheet['A3'] = "No storage accounts found"
        return
    
    # Headers
    headers = ['Account Name', 'Location', 'SKU', 'Kind', 'Resource Group']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Data rows
    row = 4
    for account in storage_accounts:
        sheet.cell(row=row, column=1, value=account.get('name', ''))
        sheet.cell(row=row, column=2, value=account.get('location', ''))
        sheet.cell(row=row, column=3, value=account.get('sku_name', ''))
        sheet.cell(row=row, column=4, value=account.get('kind', ''))
        sheet.cell(row=row, column=5, value=account.get('resource_group', ''))
        row += 1
    
    # Auto-adjust column widths
    _auto_adjust_column_widths(sheet)


def _create_container_results_analysis_sheet(sheet, container_results):
    """Create blob analysis sheet from container results format"""
    # Header
    sheet['A1'] = "Blob Storage Analysis (Container Results)"
    sheet['A1'].font = Font(bold=True, size=16)
    
    if not container_results:
        sheet['A3'] = "No container results available"
        return
    
    # Headers for container analysis
    headers = ['Account Name', 'Container Name', 'Blob Count', 'Total Size (GB)', 'Last Modified', 'Access Tier']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Data rows
    row = 4
    for result in container_results:
        sheet.cell(row=row, column=1, value=result.get('account_name', ''))
        sheet.cell(row=row, column=2, value=result.get('container_name', ''))
        sheet.cell(row=row, column=3, value=result.get('blob_count', 0))
        sheet.cell(row=row, column=4, value=round(result.get('total_size_gb', 0), 3))
        sheet.cell(row=row, column=5, value=str(result.get('last_modified', '')))
        sheet.cell(row=row, column=6, value=result.get('access_tier', ''))
        row += 1
    
    # Auto-adjust column widths
    _auto_adjust_column_widths(sheet)


def _create_original_files_analysis_sheet(sheet, file_share_results):
    """Create original-style Azure Files analysis sheet"""
    # Check if we have storage_data with file_shares
    if isinstance(file_share_results, list) and file_share_results:
        # This looks like file_share_results format
        sheet['A1'] = "Azure Files Analysis (File Share Results)"
        sheet['A1'].font = Font(bold=True, size=16)
        
        headers = ['Account Name', 'Share Name', 'File Count', 'Total Size (GB)', 'Last Modified']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        row = 4
        for result in file_share_results:
            sheet.cell(row=row, column=1, value=result.get('account_name', ''))
            sheet.cell(row=row, column=2, value=result.get('share_name', ''))
            sheet.cell(row=row, column=3, value=result.get('file_count', 0))
            sheet.cell(row=row, column=4, value=round(result.get('total_size_gb', 0), 3))
            sheet.cell(row=row, column=5, value=str(result.get('last_modified', '')))
            row += 1
        
        # Auto-adjust column widths
        _auto_adjust_column_widths(sheet)
    else:
        # Try original format
        try:
            from .reporting import _create_file_shares_sheet
            temp_wb = openpyxl.Workbook()
            _create_file_shares_sheet(temp_wb, file_share_results)
            
            source_sheet = temp_wb['Azure Files Analysis'] 
            for row in source_sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        new_cell = sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                        if cell.font:
                            new_cell.font = cell.font
                        if cell.fill:
                            new_cell.fill = cell.fill
        except:
            sheet['A1'] = "Azure Files Analysis - No Data Available"


def _create_original_recommendations_sheet(sheet, recommendations):
    """Create simple recommendations sheet"""
    # Header
    sheet['A1'] = "Cost Optimization Recommendations"
    sheet['A1'].font = Font(bold=True, size=16)
    
    if not recommendations:
        sheet['A3'] = "No recommendations available"
        return
    
    # Headers
    headers = ['Priority', 'Type', 'Title', 'Description', 'Potential Savings']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Data rows
    row = 4
    for rec in recommendations:
        sheet.cell(row=row, column=1, value=rec.get('priority', ''))
        sheet.cell(row=row, column=2, value=rec.get('type', ''))
        sheet.cell(row=row, column=3, value=rec.get('title', ''))
        sheet.cell(row=row, column=4, value=rec.get('description', ''))
        sheet.cell(row=row, column=5, value=rec.get('potential_savings', ''))
        row += 1


def _auto_adjust_column_widths(sheet):
    """Auto-adjust column widths based on content"""
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            if cell.value:
                # Convert to string and get length
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        
        # Set column width with some padding (minimum 12, maximum 50)
        adjusted_width = min(max(max_length + 2, 12), 50)
        sheet.column_dimensions[column_letter].width = adjusted_width


def _add_watermark_to_sheet(sheet, watermark_text):
    """Add watermark to sheet"""
    try:
        from .reporting import add_watermark_to_sheet
        add_watermark_to_sheet(sheet, watermark_text)
    except:
        # Fallback - just add text in bottom right
        sheet['Z50'] = watermark_text
        sheet['Z50'].font = Font(size=8, color="CCCCCC")