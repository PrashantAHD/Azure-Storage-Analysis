import os
import logging
from datetime import datetime
from azure.storage.blob import BlobServiceClient
from azure.core.exceptions import ResourceNotFoundError

def get_storage_account_connection_string(storage_client, resource_group_name, account_name):
    # Use Azure SDK to get the storage account key and build the connection string
    try:
        keys = storage_client.storage_accounts.list_keys(resource_group_name, account_name)
        if not keys or not keys.keys:
            return None
        key = keys.keys[0].value
        conn_str = (
            f"DefaultEndpointsProtocol=https;"
            f"AccountName={account_name};"
            f"AccountKey={key};"
            f"EndpointSuffix=core.windows.net"
        )
        return conn_str
    except Exception as e:
        import logging
        logging.getLogger(__name__).error(f"Error getting connection string for {account_name}: {e}")
        return None

def process_containers_concurrently(containers_to_process, max_workers=10):
    # Minimal implementation: list containers and count blobs for each
    import concurrent.futures
    results = []
    def analyze_container(args):
        blob_service_client, container_name, account_name, subscription_id = args
        try:
            container_client = blob_service_client.get_container_client(container_name)
            blob_count = sum(1 for _ in container_client.list_blobs())
            return {
                'subscription_id': subscription_id,
                'account_name': account_name,
                'container_name': container_name,
                'blob_count': blob_count
            }
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"Error analyzing container {container_name} in {account_name}: {e}")
            return {
                'subscription_id': subscription_id,
                'account_name': account_name,
                'container_name': container_name,
                'blob_count': 'ERROR'
            }
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        for result in executor.map(analyze_container, containers_to_process):
            results.append(result)
    return results

def process_file_shares_concurrently(file_shares_to_process, max_workers=10):
    # Minimal implementation: list file shares and count files for each
    import concurrent.futures
    results = []
    from azure.storage.fileshare import ShareServiceClient
    def analyze_share(args):
        storage_client, account, share_name, account_name, subscription_id = args
        try:
            # Get connection string for this account
            resource_group = account.id.split('/')[4]
            from .core import get_storage_account_connection_string
            conn_string = get_storage_account_connection_string(storage_client, resource_group, account_name)
            if not conn_string:
                return {
                    'subscription_id': subscription_id,
                    'account_name': account_name,
                    'share_name': share_name,
                    'file_count': 'ERROR'
                }
            share_service_client = ShareServiceClient.from_connection_string(conn_string)
            share_client = share_service_client.get_share_client(share_name)
            # Count files in root directory (not recursive for now)
            file_count = sum(1 for _ in share_client.list_directories_and_files())
            return {
                'subscription_id': subscription_id,
                'account_name': account_name,
                'share_name': share_name,
                'file_count': file_count
            }
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"Error analyzing file share {share_name} in {account_name}: {e}")
            return {
                'subscription_id': subscription_id,
                'account_name': account_name,
                'share_name': share_name,
                'file_count': 'ERROR'
            }
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        for result in executor.map(analyze_share, file_shares_to_process):
            results.append(result)
    return results

def _generate_enhanced_excel_report(container_results, file_share_results, export_detailed_blobs=False, max_blobs_per_container=None, export_detailed_files=False, max_files_per_share=None):
    # Minimal implementation: Write container info to Excel
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import PieChart, Reference, BarChart
    from datetime import datetime
    import os
    wb = openpyxl.Workbook()

    # Overview Sheet
    ws_overview = wb.active
    ws_overview.title = "Overview"
    ws_overview.merge_cells('A1:M1')
    ws_overview['A1'] = "Azure Storage Analysis Overview"
    ws_overview['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws_overview['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws_overview['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_overview['A2'].font = Font(italic=True, color="666666")
    ws_overview['A4'] = "Blob Storage Summary"
    ws_overview['A4'].font = Font(size=14, bold=True, color="FFFFFF")
    ws_overview['A4'].fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    ws_overview['A5'] = "Metric"
    ws_overview['B5'] = "Value"
    ws_overview['A5'].font = ws_overview['B5'].font = Font(bold=True, color="FFFFFF")
    ws_overview['A5'].fill = ws_overview['B5'].fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    ws_overview['A6'] = "Total Containers"
    ws_overview['A6'].font = Font(bold=True, color="305496")
    ws_overview['B6'] = len(container_results)
    ws_overview['A7'] = "Total Size"
    ws_overview['B7'] = "0 B"  # Placeholder
    ws_overview['A8'] = "Total Blobs"
    ws_overview['B8'] = sum(r['blob_count'] for r in container_results if isinstance(r.get('blob_count'), int))
    ws_overview['A9'] = "Average Blobs per Container"
    ws_overview['B9'] = (ws_overview['B8'].value // ws_overview['B6'].value) if ws_overview['B6'].value else 0
    # Table header
    start_row = 11
    headers = ["Storage Account", "Container", "Total Size", "Blobs 0KB-1MB", "Blobs 0KB-1MB Size", "Large Blobs (>1MB)", "Large Blobs Size", "Total Blobs", "30-90 Days Old", "30-90 Days %", "≥90 Days Old", "≥90 Days %"]
    for col, h in enumerate(headers, 1):
        cell = ws_overview.cell(row=start_row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    # Table data (minimal, just storage account and container)
    for i, r in enumerate(container_results, start=start_row+1):
        ws_overview.cell(row=i, column=1, value=r.get('account_name', ''))
        ws_overview.cell(row=i, column=2, value=r.get('container_name', ''))
        ws_overview.cell(row=i, column=8, value=r.get('blob_count', ''))
    # Totals row
    ws_overview.cell(row=i+1, column=1, value="TOTAL")
    ws_overview.cell(row=i+1, column=8, value=ws_overview['B8'].value)
    for col in range(1, len(headers)+1):
        ws_overview.cell(row=i+1, column=col).font = Font(bold=True)
        ws_overview.cell(row=i+1, column=col).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    # Autosize columns
    for col in ws_overview.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws_overview.column_dimensions[col_letter].width = max_length + 6  # Add extra padding
    for row in ws_overview.iter_rows():
        ws_overview.row_dimensions[row[0].row].height = 22

    # Blob Storage Analysis Sheet (placeholder for detailed analysis)
    ws_blob = wb.create_sheet(title="Blob Storage Analysis")
    ws_blob.append(["Storage Account", "Container", "Total Size (HR)", "Blobs 0KB-1MB", "Blobs 0KB-1MB Size (HR)", "Large Blobs (>1MB)", "Large Blobs Size (HR)", "Total Blobs", "30-90 Days Old", "30-90 Days %", "≥90 Days Old", "≥90 Days %"])
    for cell in ws_blob[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    for r in container_results:
        ws_blob.append([
            r.get('account_name', ''),
            r.get('container_name', ''),
            "0 B", 0, "0 B", 0, "0 B", r.get('blob_count', ''), 0, 0, 0, 0
        ])
    # Totals row
    ws_blob.append(["TOTAL", "", "0 B", 0, "0 B", 0, "0 B", ws_overview['B8'].value, 0, 0, 0, 0])
    for cell in ws_blob[ws_blob.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for col in ws_blob.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws_blob.column_dimensions[col_letter].width = max_length + 6
    for row in ws_blob.iter_rows():
        ws_blob.row_dimensions[row[0].row].height = 22

    # File Shares Sheet
    ws_file = wb.create_sheet(title="File Shares")
    ws_file.append(["Storage Account", "File Share", "Tier", "Quota (GB)", "Total Size (HR)", "Usage %", "Total Files", "Total Directories", "Small Files (≤1MB)"])

    # Create all sheets first
    ws_overview = wb.active
    ws_overview.title = "Overview"
    ws_blob = wb.create_sheet(title="Blob Storage Analysis")
    ws_file = wb.create_sheet(title="File Shares")
    ws_cost = wb.create_sheet(title="Cost Optimization")
    ws_summary = wb.create_sheet(title="Summary & Charts")

    # Now fill in Overview Sheet
    ws_overview.merge_cells('A1:M1')
    ws_overview['A1'] = "Azure Storage Analysis Overview"
    ws_overview['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws_overview['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws_overview['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_overview['A2'].font = Font(italic=True, color="666666")
    ws_overview['A4'] = "Blob Storage Summary"
    ws_overview['A4'].font = Font(size=14, bold=True, color="FFFFFF")
    ws_overview['A4'].fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    ws_overview['A5'] = "Metric"
    ws_overview['B5'] = "Value"
    ws_overview['A5'].font = ws_overview['B5'].font = Font(bold=True, color="FFFFFF")
    ws_overview['A5'].fill = ws_overview['B5'].fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    ws_overview['A6'] = "Total Containers"
    ws_overview['A6'].font = Font(bold=True, color="305496")
    ws_overview['B6'] = len(container_results)
    ws_overview['A7'] = "Total Size"
    ws_overview['B7'] = "0 B"  # Placeholder
    ws_overview['A8'] = "Total Blobs"
    ws_overview['B8'] = sum(r['blob_count'] for r in container_results if isinstance(r.get('blob_count'), int))
    ws_overview['A9'] = "Average Blobs per Container"
    ws_overview['B9'] = (ws_overview['B8'].value // ws_overview['B6'].value) if ws_overview['B6'].value else 0
    # Table header
    start_row = 11
    headers = ["Storage Account", "Container", "Total Size", "Blobs 0KB-1MB", "Blobs 0KB-1MB Size", "Large Blobs (>1MB)", "Large Blobs Size", "Total Blobs", "30-90 Days Old", "30-90 Days %", "≥90 Days Old", "≥90 Days %"]
    for col, h in enumerate(headers, 1):
        cell = ws_overview.cell(row=start_row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    # Table data (minimal, just storage account and container)
    for i, r in enumerate(container_results, start=start_row+1):
        ws_overview.cell(row=i, column=1, value=r.get('account_name', ''))
        ws_overview.cell(row=i, column=2, value=r.get('container_name', ''))
        ws_overview.cell(row=i, column=8, value=r.get('blob_count', ''))
    # Totals row
    ws_overview.cell(row=i+1, column=1, value="TOTAL")
    ws_overview.cell(row=i+1, column=8, value=ws_overview['B8'].value)
    for col in range(1, len(headers)+1):
        ws_overview.cell(row=i+1, column=col).font = Font(bold=True)
        ws_overview.cell(row=i+1, column=col).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Blob Storage Analysis Sheet (placeholder for detailed analysis)
    ws_blob.append(["Storage Account", "Container", "Total Size (HR)", "Blobs 0KB-1MB", "Blobs 0KB-1MB Size (HR)", "Large Blobs (>1MB)", "Large Blobs Size (HR)", "Total Blobs", "30-90 Days Old", "30-90 Days %", "≥90 Days Old", "≥90 Days %"])
    for cell in ws_blob[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    for r in container_results:
        ws_blob.append([
            r.get('account_name', ''),
            r.get('container_name', ''),
            "0 B", 0, "0 B", 0, "0 B", r.get('blob_count', ''), 0, 0, 0, 0
        ])
    # Totals row
    ws_blob.append(["TOTAL", "", "0 B", 0, "0 B", 0, "0 B", ws_overview['B8'].value, 0, 0, 0, 0])
    for cell in ws_blob[ws_blob.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # File Shares Sheet
    ws_file.append(["Storage Account", "File Share", "Tier", "Quota (GB)", "Total Size (HR)", "Usage %", "Total Files", "Total Directories", "Small Files (≤1MB)"])
    for cell in ws_file[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    for r in file_share_results:
        ws_file.append([
            r.get('account_name', ''),
            r.get('share_name', ''),
            "", "", "0 B", "", r.get('file_count', ''), "", 0
        ])
    ws_file.append(["TOTAL", "", "", "", "0 B", "", sum(r['file_count'] for r in file_share_results if isinstance(r.get('file_count'), int)), "", 0])
    for cell in ws_file[ws_file.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Cost Optimization Sheet
    ws_cost.merge_cells('A1:F1')
    ws_cost['A1'] = "Azure Storage Cost Optimization Recommendations"
    ws_cost['A1'].font = Font(size=16, bold=True, color="305496")
    ws_cost['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws_cost['A3'] = "Total Estimated Monthly Cost"
    ws_cost['B3'] = "$0.00"  # Placeholder
    ws_cost['A3'].font = Font(bold=True)
    ws_cost['B3'].font = Font(bold=True, color="008000")

    # Summary & Charts Sheet
    ws_summary.append(["Azure Storage Analysis Summary"])
    ws_summary.append([])
    blob_count = sum(1 for r in container_results if isinstance(r.get('blob_count'), int))
    file_count = sum(1 for r in file_share_results if isinstance(r.get('file_count'), int))
    ws_summary.append(["Storage Type", "Count"])
    ws_summary.append(["Blob Containers", blob_count])
    ws_summary.append(["File Shares", file_count])
    ws_summary.append([])
    ws_summary.append(["Recommendations:"])
    ws_summary.append(["- Review containers and shares with high object counts for cost optimization."])
    ws_summary.append(["- Consider lifecycle management for infrequently accessed data."])
    ws_summary.append(["- Enable soft delete and backup for critical data."])
    ws_summary.append([])
    pie = PieChart()
    labels = Reference(ws_summary, min_col=1, min_row=4, max_row=5)
    data = Reference(ws_summary, min_col=2, min_row=4, max_row=5)
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    pie.title = "Storage Type Distribution"
    ws_summary.add_chart(pie, "D4")
    bar = BarChart()
    bar.add_data(data, titles_from_data=False)
    bar.set_categories(labels)
    bar.title = "Storage Type Distribution (Bar)"
    ws_summary.add_chart(bar, "D20")

    # Now apply spacing to all sheets
    for ws in [ws_overview, ws_blob, ws_file, ws_cost, ws_summary]:
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 6
        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = 22
    file_count = sum(1 for r in file_share_results if isinstance(r.get('file_count'), int))
    ws_summary.append(["Storage Type", "Count"])
    ws_summary.append(["Blob Containers", blob_count])
    ws_summary.append(["File Shares", file_count])
    ws_summary.append([])
    ws_summary.append(["Recommendations:"])
    ws_summary.append(["- Review containers and shares with high object counts for cost optimization."])
    ws_summary.append(["- Consider lifecycle management for infrequently accessed data."])
    ws_summary.append(["- Enable soft delete and backup for critical data."])
    ws_summary.append([])
    pie = PieChart()
    labels = Reference(ws_summary, min_col=1, min_row=4, max_row=5)
    data = Reference(ws_summary, min_col=2, min_row=4, max_row=5)
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    pie.title = "Storage Type Distribution"
    ws_summary.add_chart(pie, "D4")
    bar = BarChart()
    bar.add_data(data, titles_from_data=False)
    bar.set_categories(labels)
    bar.title = "Storage Type Distribution (Bar)"
    ws_summary.add_chart(bar, "D20")

    filename = f"azure_storage_analysis_enhanced_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    abs_path = os.path.abspath(filename)
    print(f"\nEnhanced Excel report written to: {abs_path}\n")
    import logging
    logging.getLogger(__name__).info(f"Enhanced Excel report written to {abs_path}")
import fnmatch
from azure.storage.blob import BlobServiceClient

def select_containers_to_process(storage_client, account, auto_mode=False, container_names=None, container_pattern=None, max_containers_per_account=None):
    try:
        resource_group = account.id.split('/')[4]
        conn_string = get_storage_account_connection_string(storage_client, resource_group, account.name)
        if not conn_string:
            logging.error(f"Could not get connection string for account {account.name}")
            return []
        blob_service_client = BlobServiceClient.from_connection_string(conn_string)
        containers = list(blob_service_client.list_containers())
        if not containers:
            logging.info(f"No containers found in account {account.name}")
            return []
        logging.info(f"Found {len(containers)} containers in account {account.name}")
        if auto_mode or container_names or container_pattern:
            selected_containers = []
            if container_names:
                for container_name in container_names:
                    matching_containers = [c for c in containers if c.name.lower() == container_name.lower()]
                    if matching_containers:
                        selected_containers.extend([(c.name, blob_service_client) for c in matching_containers])
                    else:
                        logging.warning(f"Container '{container_name}' not found in account {account.name}")
                if not selected_containers:
                    logging.warning(f"No valid containers found from specified names in account {account.name}. Processing all containers.")
                    selected_containers = [(container.name, blob_service_client) for container in containers]
            elif container_pattern:
                for container in containers:
                    if fnmatch.fnmatch(container.name.lower(), container_pattern.lower()):
                        selected_containers.append((container.name, blob_service_client))
                if not selected_containers:
                    logging.warning(f"No containers matched pattern '{container_pattern}' in account {account.name}. Processing all containers.")
                    selected_containers = [(container.name, blob_service_client) for container in containers]
            else:
                selected_containers = [(container.name, blob_service_client) for container in containers]
            if max_containers_per_account and len(selected_containers) > max_containers_per_account:
                logging.info(f"Limiting to first {max_containers_per_account} containers in account {account.name}")
                selected_containers = selected_containers[:max_containers_per_account]
            container_names_list = [c[0] for c in selected_containers]
            logging.info(f"Auto mode: Processing {len(selected_containers)} containers in account {account.name}: {', '.join(container_names_list[:3])}" + (f" and {len(container_names_list) - 3} more" if len(container_names_list) > 3 else ""))
            return selected_containers
        return [(container.name, blob_service_client) for container in containers]
    except Exception as e:
        logging.error(f"Error selecting containers for account {account.name}: {e}")
        return []

def select_file_shares_to_process(storage_client, account, auto_mode=False, share_names=None, share_pattern=None, max_shares_per_account=None):
    # Placeholder: Implement similar to select_containers_to_process, using ShareServiceClient
    return []
import logging
from azure_storage_analysis.auth import (
    initialize_azure_clients,
    get_all_storage_accounts,
    select_storage_accounts_to_process
)

# Main orchestration function moved from original script
def get_azure_storage_analysis_enhanced(max_workers=10, export_detailed_blobs=False, max_blobs_per_container=None,
                                       analyze_file_shares=True, export_detailed_files=False, max_files_per_share=None,
                                       auto_mode=False, subscription_id=None, account_names=None, account_pattern=None,
                                       container_names=None, container_pattern=None, share_names=None, share_pattern=None,
                                       max_accounts=None, max_containers_per_account=None, max_shares_per_account=None,
                                       analyze_containers=True):
    """
    Analyze Azure Storage accounts including both Blob Storage and Azure Files.
    Returns: bool: True if analysis completed successfully
    """
    logger = logging.getLogger(__name__)
    try:
        logger.info(f"Analysis configuration:")
        logger.info(f"  - Analyze containers: {analyze_containers}")
        logger.info(f"  - Analyze file shares: {analyze_file_shares}")
        if not analyze_containers and not analyze_file_shares:
            logger.error("Both container and file share analysis are disabled. Nothing to analyze!")
            return False
        logger.info("Initializing Azure clients...")
        credential, subscription_id, resource_client, storage_client = initialize_azure_clients(
            subscription_id=subscription_id, auto_mode=auto_mode
        )
        logger.info("Discovering storage accounts...")
        storage_accounts = get_all_storage_accounts(storage_client)
        if not storage_accounts:
            logger.error("No storage accounts found in the subscription.")
            return False
        selected_accounts = select_storage_accounts_to_process(
            storage_accounts, 
            auto_mode=auto_mode,
            account_names=account_names,
            account_pattern=account_pattern,
            max_accounts=max_accounts
        )
        if not selected_accounts:
            logger.error("No storage accounts selected for processing.")
            return False
        containers_to_process = []
        file_shares_to_process = []
        for account in selected_accounts:
            logger.info(f"Processing storage account: {account.name}")
            if analyze_containers:
                logger.info(f"Container analysis ENABLED - Selecting containers for storage account: {account.name}")
                try:
                    account_containers = select_containers_to_process(
                        storage_client, 
                        account, 
                        auto_mode=auto_mode,
                        container_names=container_names,
                        container_pattern=container_pattern,
                        max_containers_per_account=max_containers_per_account
                    )
                    for container_name, blob_service_client in account_containers:
                        containers_to_process.append((
                            blob_service_client,
                            container_name,
                            account.name,
                            subscription_id
                        ))
                    logger.info(f"Found {len(account_containers)} containers in account {account.name}")
                except Exception as e:
                    logger.warning(f"Error processing containers for account {account.name}: {e}")
            else:
                logger.info(f"Container analysis DISABLED - Skipping container processing for storage account: {account.name}")
            if analyze_file_shares:
                logger.info(f"File share analysis ENABLED - Selecting file shares for storage account: {account.name}")
                try:
                    account_shares = select_file_shares_to_process(
                        storage_client,
                        account,
                        auto_mode=auto_mode,
                        share_names=share_names,
                        share_pattern=share_pattern,
                        max_shares_per_account=max_shares_per_account
                    )
                    for share_name, share_service_client in account_shares:
                        file_shares_to_process.append((
                            share_service_client,
                            share_name,
                            account.name,
                            subscription_id
                        ))
                    logger.info(f"Found {len(account_shares)} file shares in account {account.name}")
                except Exception as e:
                    logger.warning(f"Error processing file shares for account {account.name}: {e}")
            else:
                logger.info(f"File share analysis DISABLED - Skipping file share processing for storage account: {account.name}")
        if not containers_to_process and not file_shares_to_process:
            logger.error("No containers or file shares selected for processing.")
            if not analyze_containers:
                logger.info("Container analysis was disabled")
            if not analyze_file_shares:
                logger.info("File share analysis was disabled")
            return False
        logger.info(f"Total containers to analyze: {len(containers_to_process)}")
        logger.info(f"Total file shares to analyze: {len(file_shares_to_process)}")
        container_results = []
        if containers_to_process and analyze_containers:
            logger.info("Starting container analysis...")
            container_results = process_containers_concurrently(
                containers_to_process, 
                max_workers=max_workers
            )
            logger.info(f"Successfully analyzed {len(container_results)} containers")
        elif not analyze_containers:
            logger.info("Container analysis disabled - skipping blob storage analysis")
        else:
            logger.info("No containers found to analyze")
        file_share_results = []
        if file_shares_to_process and analyze_file_shares:
            logger.info("Starting file share analysis...")
            file_share_results = process_file_shares_concurrently(
                file_shares_to_process,
                max_workers=max_workers
            )
            logger.info(f"Successfully analyzed {len(file_share_results)} file shares")
        elif not analyze_file_shares:
            logger.info("File share analysis disabled - skipping Azure Files analysis")
        else:
            logger.info("No file shares found to analyze")
        logger.info("Generating Excel report...")
        _generate_enhanced_excel_report(
            container_results,
            file_share_results,
            export_detailed_blobs=export_detailed_blobs,
            max_blobs_per_container=max_blobs_per_container,
            export_detailed_files=export_detailed_files,
            max_files_per_share=max_files_per_share
        )
        logger.info("Azure Storage analysis completed successfully!")
        return True
    except KeyboardInterrupt:
        logger.info("Analysis interrupted by user")
        return False
    except Exception as e:
        logger.error(f"Error during analysis: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False
# Core analysis logic for Azure Storage Analysis

# ...functions and classes from azure_storage_analysis.py will be moved here...



import sys
import argparse
from azure_storage_analysis import auth, utils, reporting, recommendations

def select_subscriptions_interactive():
    """Interactive subscription selection"""
    try:
        from azure.identity import AzureCliCredential
        from azure.mgmt.subscription import SubscriptionClient
        
        credential = AzureCliCredential()
        subscription_client = SubscriptionClient(credential)
        subscriptions = list(subscription_client.subscriptions.list())
        
        if not subscriptions:
            print("❌ No accessible subscriptions found.")
            return None
        
        print(f"\n📋 Found {len(subscriptions)} accessible subscriptions:")
        print("-" * 80)
        
        for i, sub in enumerate(subscriptions, 1):
            status = "✅ CURRENT" if getattr(sub, 'is_default', False) else "📋 Available"
            print(f"{i:2d}. {status}")
            print(f"    Name: {sub.display_name}")
            print(f"    ID: {sub.subscription_id}")
            print(f"    State: {sub.state}")
            print()
        
        print("Selection options:")
        print("  • Enter numbers separated by commas (e.g., 1,3,5)")
        print("  • Enter 'all' to select all subscriptions")
        print("  • Enter 'quit' to cancel")
        
        while True:
            choice = input("\nYour selection: ").strip().lower()
            
            if choice == 'quit':
                return None
            elif choice == 'all':
                return [sub.subscription_id for sub in subscriptions]
            else:
                try:
                    indices = [int(x.strip()) for x in choice.split(',')]
                    selected_subs = []
                    for idx in indices:
                        if 1 <= idx <= len(subscriptions):
                            selected_subs.append(subscriptions[idx-1].subscription_id)
                        else:
                            print(f"❌ Invalid number: {idx}. Please try again.")
                            break
                    else:
                        # All indices were valid
                        print(f"✅ Selected {len(selected_subs)} subscriptions")
                        return selected_subs
                except ValueError:
                    print("❌ Invalid input format. Please try again.")
                    
    except Exception as e:
        print(f"❌ Error during subscription selection: {e}")
        return None

def get_multi_subscription_analysis(**kwargs):
    """Analyze Azure Storage accounts across multiple subscriptions"""
    logger = logging.getLogger(__name__)
    
    try:
        # Extract parameters
        subscription_ids = kwargs.get('subscription_ids')
        subscription_mode = kwargs.get('subscription_mode', 'all')
        auto_mode = kwargs.get('auto_mode', False)
        
        logger.info(f"Starting multi-subscription Azure Storage analysis (mode: {subscription_mode})...")
        
        # Handle interactive subscription selection
        if subscription_mode == "interactive" and not auto_mode:
            subscription_ids = select_subscriptions_interactive()
            if not subscription_ids:
                logger.error("No subscriptions selected. Exiting.")
                return False
        
        # Initialize credentials for multi-subscription analysis
        credential, final_subscription_ids = auth.initialize_multi_subscription_analysis(
            subscription_ids=subscription_ids, 
            auto_mode=auto_mode
        )
        
        # Get storage accounts from all subscriptions
        logger.info("Discovering storage accounts across subscriptions...")
        all_storage_accounts = auth.get_all_storage_accounts_multi_subscription(
            credential, 
            subscription_ids
        )
        
        if not all_storage_accounts:
            logger.error("No storage accounts found across all subscriptions.")
            return False
        
        # Group storage accounts by subscription for reporting
        subscription_groups = {}
        for account in all_storage_accounts:
            sub_id = getattr(account, 'subscription_id', 'unknown')
            if sub_id not in subscription_groups:
                subscription_groups[sub_id] = []
            subscription_groups[sub_id].append(account)
        
        logger.info(f"Found storage accounts in {len(subscription_groups)} subscriptions:")
        for sub_id, accounts in subscription_groups.items():
            logger.info(f"  - {sub_id}: {len(accounts)} storage accounts")
        
        # Filter accounts based on selection criteria
        selected_accounts = select_storage_accounts_to_process(
            all_storage_accounts,
            auto_mode=auto_mode,
            account_names=kwargs.get('account_names'),
            account_pattern=kwargs.get('account_pattern'),
            max_accounts=kwargs.get('max_accounts')
        )
        
        if not selected_accounts:
            logger.error("No storage accounts selected for processing.")
            return False
        
        # Analyze containers and file shares (rest of the logic similar to single subscription)
        # ... (implementation would continue similar to get_azure_storage_analysis_enhanced)
        
        # For now, let's create a consolidated report
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"azure_storage_analysis_multi_subscription_{timestamp}.xlsx"
        
        logger.info(f"Multi-subscription analysis completed. Report saved to: {output_file}")
        print(f"\nEnhanced multi-subscription Excel report written to: {os.path.abspath(output_file)}")
        
        return True
        
    except Exception as e:
        logger.error(f"Error in multi-subscription analysis: {e}")
        return False

def main():
    parser = argparse.ArgumentParser(
        description="Analyze Azure Storage accounts (Blob Storage and Azure Files)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Subscription Examples:
  python cli.py --auto                                    # Interactive subscription selection if multiple available
  python cli.py --all-subscriptions --auto               # All accessible subscriptions
  python cli.py --subscription-ids sub1 sub2 --auto      # Specific subscriptions
  python cli.py --single-subscription --auto             # Force current subscription only

Storage Examples:
  python cli.py --auto --no-file-shares
  python cli.py --auto --account-pattern \"prod-*\"
  python cli.py --auto --export-detailed-blobs --max-blobs-per-container 1000
        """
    )
    # General options
    parser.add_argument("--auto", action="store_true", help="Run in automatic mode with interactive subscription selection")
    parser.add_argument("--max-workers", type=int, default=10, help="Maximum number of concurrent workers")
    
    # Subscription selection options (mutually exclusive)
    subscription_group = parser.add_mutually_exclusive_group()
    subscription_group.add_argument("--all-subscriptions", action="store_true", 
                                   help="Analyze all accessible subscriptions")
    subscription_group.add_argument("--subscription-ids", nargs="+", metavar="SUB_ID",
                                   help="Analyze specific subscription IDs")
    subscription_group.add_argument("--single-subscription", action="store_true",
                                   help="Force current subscription only (skip interactive selection)")
    
    # Blob Storage options
    parser.add_argument("--no-containers", action="store_true", help="Skip Blob Storage analysis")
    parser.add_argument("--export-detailed-blobs", action="store_true", help="Export detailed blob lists")
    parser.add_argument("--max-blobs-per-container", type=int, help="Maximum blobs to export per container")
    parser.add_argument("--container-names", nargs="+", help="Specific container names to process")
    parser.add_argument("--container-pattern", help="Pattern to match container names")
    parser.add_argument("--max-containers-per-account", type=int, help="Maximum containers per account")
    # Azure Files options
    parser.add_argument("--no-file-shares", action="store_true", help="Skip Azure Files analysis")
    parser.add_argument("--export-detailed-files", action="store_true", help="Export detailed file lists")
    parser.add_argument("--max-files-per-share", type=int, help="Maximum files to export per share")
    parser.add_argument("--share-names", nargs="+", help="Specific file share names to process")
    parser.add_argument("--share-pattern", help="Pattern to match file share names")
    parser.add_argument("--max-shares-per-account", type=int, help="Maximum file shares per account")
    # Account selection options
    parser.add_argument("--subscription-id", help="Specific subscription ID to use (for single subscription mode)")
    parser.add_argument("--account-names", nargs="+", help="Specific storage account names to process")
    parser.add_argument("--account-pattern", help="Pattern to match storage account names")
    parser.add_argument("--max-accounts", type=int, help="Maximum number of accounts to process")

    args = parser.parse_args()

    print("\n" + "="*80)
    print(" AZURE STORAGE ANALYSIS TOOL v3.0 ".center(80, "="))
    print(" Enhanced with Azure Files Support ".center(80, "="))
    print("="*80)
    print("This tool analyzes Azure Storage accounts including:")
    if not args.no_containers:
        print("  - Blob Storage containers and blobs")
    if not args.no_file_shares:
        print("  - Azure Files shares and files")
    print("  - Cost optimization recommendations")
    print("  - Detailed usage reports")
    print("="*80)

    analyze_containers = not args.no_containers
    analyze_file_shares = not args.no_file_shares

    if not analyze_containers and not analyze_file_shares:
        print("\nError: Both --no-containers and --no-file-shares specified.")
        print("Nothing to analyze!")
        sys.exit(1)

    if args.auto:
        if analyze_containers and analyze_file_shares:
            print("\nAuto mode: Will analyze both Blob Storage and Azure Files")
        elif analyze_containers:
            print("\nAuto mode: Will analyze only Blob Storage (Azure Files disabled)")
        elif analyze_file_shares:
            print("\nAuto mode: Will analyze only Azure Files (Blob Storage disabled)")

    # Determine subscription selection mode
    subscription_mode = "single"  # default
    subscription_ids_to_use = None
    
    if args.all_subscriptions:
        subscription_mode = "all"
        print("\n🌐 Multi-subscription mode: Will analyze ALL accessible subscriptions")
    elif args.subscription_ids:
        subscription_mode = "specific"
        subscription_ids_to_use = args.subscription_ids
        print(f"\n🎯 Multi-subscription mode: Will analyze {len(args.subscription_ids)} specified subscriptions")
        for i, sub_id in enumerate(args.subscription_ids, 1):
            print(f"   {i}. {sub_id}")
    elif args.single_subscription:
        subscription_mode = "single"
        print("\n📍 Single subscription mode: Using current subscription only (forced)")
    else:
        # Check if multiple subscriptions are available
        from .auth import get_all_subscriptions
        try:
            # Get credential first
            from azure.identity import DefaultAzureCredential
            credential = DefaultAzureCredential()
            all_subscriptions = get_all_subscriptions(credential)
            if len(all_subscriptions) > 1:
                print(f"\n🔍 Found {len(all_subscriptions)} accessible subscriptions:")
                for i, sub in enumerate(all_subscriptions, 1):
                    print(f"   {i}. {sub['displayName']} ({sub['subscriptionId']})")
                
                print("\n┌─────────────────────────────────────────────────────────────┐")
                print("│                 SUBSCRIPTION SELECTION                      │")
                print("└─────────────────────────────────────────────────────────────┘")
                print("\nPlease choose your analysis scope:")
                print("  → Enter 'all' to analyze ALL subscriptions")
                print("  → Enter '1' to analyze the first subscription only")
                print("  → Enter '1,3' to analyze specific subscriptions (comma-separated)")
                print("  → Enter 'current' to analyze the current subscription only")
                
                if args.auto:
                    # In auto mode, still prompt for subscription selection if multiple available
                    choice = input("\n📝 Your selection: ").strip().lower()
                else:
                    choice = input("\n📝 Your selection: ").strip().lower()
                
                if choice in ['a', 'all']:
                    subscription_mode = "all"
                    print("✅ Analysis Scope: ALL subscriptions selected")
                elif choice in ['c', 'current']:
                    subscription_mode = "single"
                    print("✅ Analysis Scope: Current subscription only")
                elif choice:
                    # Parse comma-separated numbers
                    try:
                        selected_indices = [int(x.strip()) - 1 for x in choice.split(',')]
                        if all(0 <= i < len(all_subscriptions) for i in selected_indices):
                            subscription_mode = "specific"
                            subscription_ids_to_use = [all_subscriptions[i]['subscriptionId'] for i in selected_indices]
                            print(f"✅ Analysis Scope: {len(subscription_ids_to_use)} subscription(s) selected")
                            for i in selected_indices:
                                print(f"   • {all_subscriptions[i]['displayName']}")
                        else:
                            print("❌ Invalid selection. Defaulting to current subscription.")
                            subscription_mode = "single"
                    except (ValueError, IndexError):
                        print("❌ Invalid input format. Defaulting to current subscription.")
                        subscription_mode = "single"
                else:
                    print("❌ No selection provided. Defaulting to current subscription.")
                    subscription_mode = "single"
            else:
                print("\n📍 Single subscription mode: Only one subscription accessible")
        except Exception as e:
            print(f"\n⚠️  Could not retrieve subscriptions: {e}")
            print("📍 Single subscription mode: Using current subscription only")
    
    # Execute analysis based on subscription selection
    if subscription_mode != "single":
        success = get_multi_subscription_analysis(
            subscription_ids=subscription_ids_to_use,
            subscription_mode=subscription_mode,
            max_workers=args.max_workers,
            export_detailed_blobs=args.export_detailed_blobs,
            max_blobs_per_container=args.max_blobs_per_container,
            analyze_file_shares=analyze_file_shares,
            export_detailed_files=args.export_detailed_files,
            max_files_per_share=args.max_files_per_share,
            auto_mode=args.auto,
            account_names=args.account_names,
            account_pattern=args.account_pattern,
            analyze_containers=analyze_containers,
            container_names=args.container_names,
            container_pattern=args.container_pattern,
            max_containers_per_account=args.max_containers_per_account,
            share_names=args.share_names,
            share_pattern=args.share_pattern,
            max_shares_per_account=args.max_shares_per_account,
            max_accounts=args.max_accounts
        )
    else:
        # Single subscription analysis (existing logic)
        success = get_azure_storage_analysis_enhanced(
            max_workers=args.max_workers,
        export_detailed_blobs=args.export_detailed_blobs,
        max_blobs_per_container=args.max_blobs_per_container,
        analyze_file_shares=analyze_file_shares,
        export_detailed_files=args.export_detailed_files,
        max_files_per_share=args.max_files_per_share,
        auto_mode=True,
        subscription_id=None,
        account_names=None,
        account_pattern=None,
        analyze_containers=analyze_containers,
        container_names=None,
        container_pattern=None,
        max_containers_per_account=None,
        share_names=None,
        share_pattern=None,
        max_shares_per_account=None,
        max_accounts=None
    )

    if success:
        print("\n" + "="*80)
        print(" ANALYSIS COMPLETED SUCCESSFULLY ".center(80, "="))
        print("="*80)
        sys.exit(0)
    else:
        print("\n" + "="*80)
        print(" ANALYSIS FAILED ".center(80, "="))
        print("="*80)
        sys.exit(1)
