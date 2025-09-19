# Reporting and Excel/CSV export logic for Azure Storage Analysis

import os
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import csv

def create_enhanced_excel_report(storage_data, recommendations, output_file):
    """Create enhanced Excel report with multiple sheets"""
    logger = logging.getLogger(__name__)
    
    try:
        workbook = Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        # Create summary sheet
        _create_summary_sheet(workbook, storage_data, recommendations)
        
        # Create blob storage analysis sheet  
        _create_blob_analysis_sheet(workbook, storage_data)
        
        # Create recommendations sheet
        _create_recommendations_sheet(workbook, recommendations)
        
        # Create file shares sheet if data exists
        file_shares_data = [account for account in storage_data if account.get('file_shares')]
        if file_shares_data:
            _create_file_shares_sheet(workbook, file_shares_data)
        
        # Add watermark to all sheets
        watermark_text = f"Azure Storage Analysis Report - Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        for sheet in workbook.worksheets:
            add_watermark_to_sheet(sheet, watermark_text)
        
        workbook.save(output_file)
        logger.info(f"Enhanced Excel report saved to: {output_file}")
        return True
        
    except Exception as e:
        logger.error(f"Error creating Excel report: {e}")
        return False

def _create_summary_sheet(workbook, storage_data, recommendations):
    """Create summary overview sheet"""
    sheet = workbook.create_sheet("Summary")
    
    # Header styling
    header_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Title
    sheet['A1'] = "Azure Storage Analysis Summary"
    sheet['A1'].font = Font(bold=True, size=16)
    sheet.merge_cells('A1:D1')
    
    # Summary statistics
    total_accounts = len(storage_data)
    total_containers = sum(len(account.get('containers', [])) for account in storage_data)
    total_blobs = sum(
        sum(container.get('blob_count', 0) for container in account.get('containers', []))
        for account in storage_data
    )
    
    summary_data = [
        ['Metric', 'Value'],
        ['Total Storage Accounts', total_accounts],
        ['Total Containers', total_containers], 
        ['Total Blobs', f"{total_blobs:,}"],
        ['Analysis Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['Total Recommendations', len(recommendations)]
    ]
    
    # Write summary data
    for i, row in enumerate(summary_data, start=3):
        for j, value in enumerate(row, start=1):
            cell = sheet.cell(row=i, column=j, value=value)
            if i == 3:  # Header row
                cell.font = header_font
                cell.fill = header_fill
    
    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

def _create_blob_analysis_sheet(workbook, storage_data):
    """Create detailed blob storage analysis sheet"""
    sheet = workbook.create_sheet("Blob Storage Analysis")
    
    # Headers
    headers = [
        'Storage Account', 'Container', 'Total Size (HR)', 'Blobs 0KB-1MB', 
        'Blobs 0KB-1MB Size (HR)', 'Large Blobs (>1MB)', 'Large Blobs Size (HR)', 
        'Total Blobs', '30-90 Days Old', '30-90 Days %', '≥90 Days Old', '≥90 Days %'
    ]
    
    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Write data
    row_num = 2
    for account in storage_data:
        account_name = account.get('account_name', 'Unknown')
        containers = account.get('containers', [])
        
        for container in containers:
            data_row = [
                account_name,
                container.get('name', 'Unknown'),
                container.get('total_size_hr', '0 B'),
                container.get('small_blobs_count', 0),
                container.get('small_blobs_size_hr', '0 B'),
                container.get('large_blobs_count', 0),
                container.get('large_blobs_size_hr', '0 B'),
                container.get('blob_count', 0),
                container.get('blobs_30_90_days', 0),
                f"{container.get('blobs_30_90_days_pct', 0):.1f}%",
                container.get('blobs_90_plus_days', 0),
                f"{container.get('blobs_90_plus_days_pct', 0):.1f}%"
            ]
            
            for col, value in enumerate(data_row, start=1):
                sheet.cell(row=row_num, column=col, value=value)
            
            row_num += 1
    
    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

def _create_recommendations_sheet(workbook, recommendations):
    """Create cost optimization recommendations sheet"""
    sheet = workbook.create_sheet("Recommendations")
    
    # Headers
    headers = ['Priority', 'Type', 'Account', 'Description', 'Potential Savings', 'Action']
    
    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Write recommendations
    for row, rec in enumerate(recommendations, start=2):
        data_row = [
            rec.get('priority', 'Medium'),
            rec.get('type', 'General'),
            rec.get('account', 'All'),
            rec.get('description', ''),
            rec.get('potential_savings', 'TBD'),
            rec.get('action', '')
        ]
        
        for col, value in enumerate(data_row, start=1):
            cell = sheet.cell(row=row, column=col, value=value)
            
            # Color-code by priority
            if rec.get('priority') == 'High':
                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            elif rec.get('priority') == 'Medium':
                cell.fill = PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid")
    
    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min((max_length + 2) * 1.2, 50)  # Cap at 50 for readability
        sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

def _create_file_shares_sheet(workbook, file_shares_data):
    """Create Azure Files analysis sheet"""
    sheet = workbook.create_sheet("Azure Files Analysis")
    
    # Headers
    headers = ['Storage Account', 'File Share', 'Total Size', 'File Count', 'Last Modified']
    
    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Write file shares data
    row_num = 2
    for account in file_shares_data:
        account_name = account.get('account_name', 'Unknown')
        file_shares = account.get('file_shares', [])
        
        for share in file_shares:
            data_row = [
                account_name,
                share.get('name', 'Unknown'),
                share.get('size_hr', '0 B'),
                share.get('file_count', 0),
                share.get('last_modified', 'Unknown')
            ]
            
            for col, value in enumerate(data_row, start=1):
                sheet.cell(row=row_num, column=col, value=value)
            
            row_num += 1

def add_watermark_to_sheet(sheet, watermark_text):
    """Add watermark to sheet"""
    row = sheet.max_row + 2
    cell = sheet.cell(row=row, column=1)
    cell.value = watermark_text
    cell.font = Font(italic=True, color="888888", size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center")

def save_excel_with_watermark(workbook, output_file, watermark_text):
    """Save Excel file with watermark on all sheets"""
    for sheet in workbook.worksheets:
        add_watermark_to_sheet(sheet, watermark_text)
    workbook.save(output_file)

def save_csv_with_watermark(rows, output_file, watermark_text):
    """Save CSV file with watermark"""
    with open(output_file, mode='w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow([watermark_text])
        writer.writerow([])
        for row in rows:
            writer.writerow(row)
