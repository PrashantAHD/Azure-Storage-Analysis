# Enhanced Excel Reporting with Cost Management, Reservations, and Savings Plans

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference, BarChart, LineChart
from datetime import datetime
import os
import logging

logger = logging.getLogger(__name__)

def create_enhanced_excel_report_with_cost_analysis(container_results, file_share_results, 
                                                   cost_analysis=None, reservations_analysis=None, 
                                                   savings_plans_analysis=None):
    """
    Create comprehensive Excel report with cost management analysis
    
    Args:
        container_results: Blob storage analysis results
        file_share_results: Azure Files analysis results
        cost_analysis: Monthly spending analysis and trends
        reservations_analysis: Reserved Instance recommendations
        savings_plans_analysis: Savings Plans recommendations
        
    Returns:
        str: Path to generated Excel file
    """
    logger.info("Creating enhanced Excel report with cost analysis...")
    
    wb = openpyxl.Workbook()
    
    # Remove default sheet and create our sheets
    wb.remove(wb.active)
    
    # Create all sheets
    sheets = {
        'Executive Summary': wb.create_sheet("Executive Summary"),
        'Monthly Spending': wb.create_sheet("Monthly Spending"),
        'Storage Analysis': wb.create_sheet("Storage Analysis"),
        'Reservations': wb.create_sheet("Reservations"),
        'Savings Plans': wb.create_sheet("Savings Plans"),
        'Cost Optimization': wb.create_sheet("Cost Optimization"),
        'Detailed Data': wb.create_sheet("Detailed Data")
    }
    
    # Create Executive Summary
    _create_executive_summary_sheet(sheets['Executive Summary'], container_results, file_share_results, 
                                   cost_analysis, reservations_analysis, savings_plans_analysis)
    
    # Create Monthly Spending Analysis
    if cost_analysis:
        _create_monthly_spending_sheet(sheets['Monthly Spending'], cost_analysis)
    
    # Create Storage Analysis
    _create_storage_analysis_sheet(sheets['Storage Analysis'], container_results, file_share_results)
    
    # Create Reservations Analysis
    if reservations_analysis:
        _create_reservations_sheet(sheets['Reservations'], reservations_analysis)
    
    # Create Savings Plans Analysis
    if savings_plans_analysis:
        _create_savings_plans_sheet(sheets['Savings Plans'], savings_plans_analysis)
    
    # Create Cost Optimization Summary
    _create_cost_optimization_sheet(sheets['Cost Optimization'], cost_analysis, 
                                  reservations_analysis, savings_plans_analysis)
    
    # Create Detailed Data Sheet
    _create_detailed_data_sheet(sheets['Detailed Data'], container_results, file_share_results)
    
    # Format all sheets
    for sheet in sheets.values():
        _format_sheet(sheet)
    
    # Save file
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"azure_finops_analysis_enhanced_{timestamp}.xlsx"
    wb.save(filename)
    
    abs_path = os.path.abspath(filename)
    logger.info(f"Enhanced Excel report saved to: {abs_path}")
    print(f"\n📊 Enhanced FinOps Excel report written to: {abs_path}")
    
    return abs_path

def _create_executive_summary_sheet(ws, container_results, file_share_results, 
                                  cost_analysis, reservations_analysis, savings_plans_analysis):
    """Create executive summary with key metrics and recommendations"""
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "Azure FinOps Analysis - Executive Summary"
    ws['A1'].font = Font(size=20, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %H:%M')}"
    ws['A2'].font = Font(italic=True, size=12)
    
    current_row = 4
    
    # Key Metrics Section
    ws[f'A{current_row}'] = "📊 KEY METRICS"
    ws[f'A{current_row}'].font = Font(size=14, bold=True, color="2E75B6")
    current_row += 2
    
    # Storage metrics
    total_containers = len(container_results)
    total_shares = len(file_share_results)
    total_blobs = sum(r.get('blob_count', 0) for r in container_results if isinstance(r.get('blob_count'), int))
    
    metrics = [
        ("Storage Containers", total_containers),
        ("File Shares", total_shares),
        ("Total Blobs", f"{total_blobs:,}"),
    ]
    
    # Add cost metrics if available
    if cost_analysis:
        total_spending = cost_analysis.get('total_spending', {})
        if total_spending:
            latest_month = max(total_spending.keys()) if total_spending else None
            if latest_month:
                current_cost = total_spending[latest_month].get('total_cost', 0)
                metrics.append(("Current Monthly Cost", f"${current_cost:,.2f}"))
    
    # Add savings potential
    total_potential_savings = 0
    if reservations_analysis:
        total_potential_savings += reservations_analysis.get('total_potential_savings', 0)
    if savings_plans_analysis:
        sp_summary = savings_plans_analysis.get('summary', {})
        total_potential_savings += sp_summary.get('total_annual_savings', 0)
    
    if total_potential_savings > 0:
        metrics.append(("Annual Savings Potential", f"${total_potential_savings:,.2f}"))
    
    for i, (metric, value) in enumerate(metrics):
        row = current_row + i
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(color="0066CC")
    
    current_row += len(metrics) + 2
    
    # Monthly Spending Trend Section
    if cost_analysis and cost_analysis.get('month_over_month_analysis'):
        ws[f'A{current_row}'] = "📈 MONTHLY SPENDING TREND"
        ws[f'A{current_row}'].font = Font(size=14, bold=True, color="2E75B6")
        current_row += 2
        
        mom_analysis = cost_analysis.get('month_over_month_analysis', [])
        if mom_analysis:
            latest_change = mom_analysis[-1]
            
            trend_text = f"From {latest_change['from_month']} to {latest_change['to_month']}:"
            change_amount = latest_change['change_amount']
            change_percent = latest_change['change_percent']
            
            if change_amount > 0:
                trend_details = f"📈 INCREASED by ${change_amount:,.2f} ({change_percent:.1f}%)"
                color = "FF6B6B"  # Red for increase
            elif change_amount < 0:
                trend_details = f"📉 DECREASED by ${abs(change_amount):,.2f} ({abs(change_percent):.1f}%)"
                color = "51CF66"  # Green for decrease
            else:
                trend_details = "➡️ STABLE (no significant change)"
                color = "868E96"  # Gray for stable
            
            ws[f'A{current_row}'] = trend_text
            ws[f'A{current_row+1}'] = trend_details
            ws[f'A{current_row+1}'].font = Font(bold=True, color=color)
            
            current_row += 3
    
    # Top Recommendations Section
    ws[f'A{current_row}'] = "🎯 TOP RECOMMENDATIONS"
    ws[f'A{current_row}'].font = Font(size=14, bold=True, color="2E75B6")
    current_row += 2
    
    recommendations = []
    
    # Add high-priority reservation recommendations
    if reservations_analysis:
        blob_recs = reservations_analysis.get('blob_storage_reservations', [])
        files_recs = reservations_analysis.get('files_reservations', [])
        high_priority_recs = [r for r in blob_recs + files_recs if r.get('recommendation_priority') == 'High']
        
        for rec in high_priority_recs[:2]:  # Top 2
            savings = rec.get('annual_savings', 0)
            service = rec.get('service', 'Storage')
            recommendations.append(f"💰 Purchase {service} reservation (${savings:,.0f}/year savings)")
    
    # Add high-priority savings plan recommendations
    if savings_plans_analysis:
        compute_plans = savings_plans_analysis.get('compute_savings_plans', [])
        azure_plans = savings_plans_analysis.get('azure_savings_plans', [])
        high_priority_plans = [p for p in compute_plans + azure_plans if p.get('recommendation_priority') == 'High']
        
        for plan in high_priority_plans[:1]:  # Top 1
            savings = plan.get('estimated_annual_savings', 0)
            plan_type = plan.get('plan_type', 'Savings Plan')
            recommendations.append(f"📋 Implement {plan_type} (${savings:,.0f}/year savings)")
    
    # Add general recommendations
    if total_containers > 50:
        recommendations.append("🏗️ Implement lifecycle management policies for automated tier transitions")
    
    if total_potential_savings > 10000:
        recommendations.append("⚡ High savings potential detected - prioritize cost optimization initiatives")
    
    for i, rec in enumerate(recommendations[:5]):  # Top 5 recommendations
        ws[f'A{current_row + i}'] = f"{i+1}. {rec}"
        ws[f'A{current_row + i}'].font = Font(size=11)
    
    current_row += len(recommendations) + 2
    
    # Quick Actions Section
    ws[f'A{current_row}'] = "⚡ QUICK ACTIONS"
    ws[f'A{current_row}'].font = Font(size=14, bold=True, color="2E75B6")
    current_row += 2
    
    quick_actions = [
        "Review 'Reservations' tab for immediate purchase opportunities",
        "Check 'Savings Plans' tab for flexible commitment options", 
        "Analyze 'Monthly Spending' tab for cost trend insights",
        "Implement recommendations from 'Cost Optimization' tab"
    ]
    
    for i, action in enumerate(quick_actions):
        ws[f'A{current_row + i}'] = f"• {action}"
        ws[f'A{current_row + i}'].font = Font(size=11)

def _create_monthly_spending_sheet(ws, cost_analysis):
    """Create monthly spending analysis with charts"""
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = "Monthly Storage Spending Analysis"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center")
    
    current_row = 3
    
    # Monthly spending data
    total_spending = cost_analysis.get('total_spending', {})
    if total_spending:
        # Headers
        headers = ["Month", "Total Cost", "Change from Previous", "Change %", "Trend"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header)
            ws.cell(row=current_row, column=col).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=current_row, column=col).fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        
        current_row += 1
        
        # Data
        months = sorted(total_spending.keys())
        mom_changes = {change['to_month']: change for change in cost_analysis.get('month_over_month_analysis', [])}
        
        chart_data_start = current_row
        for month_key in months:
            month_data = total_spending[month_key]
            month_name = month_data.get('month_name', month_key)
            cost = month_data.get('total_cost', 0)
            
            # Get change data
            change_info = None
            for change in cost_analysis.get('month_over_month_analysis', []):
                if change['to_month'] == month_name:
                    change_info = change
                    break
            
            if change_info:
                change_amount = change_info.get('change_amount', 0)
                change_percent = change_info.get('change_percent', 0)
                trend = "📈" if change_amount > 0 else "📉" if change_amount < 0 else "➡️"
                change_text = f"${change_amount:+,.2f}"
            else:
                change_amount = 0
                change_percent = 0
                trend = "➡️"
                change_text = "N/A"
            
            ws.cell(row=current_row, column=1, value=month_name)
            ws.cell(row=current_row, column=2, value=f"${cost:,.2f}")
            ws.cell(row=current_row, column=3, value=change_text)
            ws.cell(row=current_row, column=4, value=f"{change_percent:+.1f}%" if change_percent != 0 else "N/A")
            ws.cell(row=current_row, column=5, value=trend)
            
            # Color code the change
            if change_amount > 0:
                ws.cell(row=current_row, column=3).font = Font(color="E74C3C")  # Red
                ws.cell(row=current_row, column=4).font = Font(color="E74C3C")
            elif change_amount < 0:
                ws.cell(row=current_row, column=3).font = Font(color="27AE60")  # Green
                ws.cell(row=current_row, column=4).font = Font(color="27AE60")
            
            current_row += 1
        
        chart_data_end = current_row - 1
        
        # Create spending trend chart
        if len(months) >= 2:
            chart = LineChart()
            chart.title = "Monthly Storage Spending Trend"
            chart.style = 10
            chart.y_axis.title = 'Cost ($)'
            chart.x_axis.title = 'Month'
            
            # Data for chart (column B = costs)
            data = Reference(ws, min_col=2, min_row=chart_data_start, max_row=chart_data_end, max_col=2)
            categories = Reference(ws, min_col=1, min_row=chart_data_start, max_row=chart_data_end)
            
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(categories)
            
            # Add chart to sheet
            ws.add_chart(chart, "H3")
    
    # Service breakdown section
    current_row += 3
    ws[f'A{current_row}'] = "Service Breakdown Analysis"
    ws[f'A{current_row}'].font = Font(size=14, bold=True, color="E74C3C")
    current_row += 2
    
    service_breakdown = cost_analysis.get('storage_service_breakdown', {})
    if service_breakdown:
        # Headers for service breakdown
        headers = ["Service", "Total Cost", "Avg Monthly", "Trend", "Subscriptions"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header)
            ws.cell(row=current_row, column=col).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=current_row, column=col).fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        
        current_row += 1
        
        # Service data
        pie_chart_start = current_row
        for service, data in service_breakdown.items():
            total_cost = data.get('total_cost', 0)
            avg_cost = data.get('average_monthly_cost', 0)
            trend_direction = data.get('trend_direction', 'stable')
            subscription_count = data.get('subscription_count', 0)
            
            trend_icon = "📈" if trend_direction == 'increasing' else "📉" if trend_direction == 'decreasing' else "➡️"
            
            ws.cell(row=current_row, column=1, value=service)
            ws.cell(row=current_row, column=2, value=f"${total_cost:,.2f}")
            ws.cell(row=current_row, column=3, value=f"${avg_cost:,.2f}")
            ws.cell(row=current_row, column=4, value=trend_icon)
            ws.cell(row=current_row, column=5, value=subscription_count)
            
            current_row += 1
        
        pie_chart_end = current_row - 1
        
        # Create service breakdown pie chart
        if len(service_breakdown) >= 2:
            pie_chart = PieChart()
            pie_chart.title = "Storage Cost by Service"
            
            # Data for pie chart
            data = Reference(ws, min_col=2, min_row=pie_chart_start, max_row=pie_chart_end)
            categories = Reference(ws, min_col=1, min_row=pie_chart_start, max_row=pie_chart_end)
            
            pie_chart.add_data(data, titles_from_data=False)
            pie_chart.set_categories(categories)
            
            # Add chart
            ws.add_chart(pie_chart, "H20")

def _create_storage_analysis_sheet(ws, container_results, file_share_results):
    """Create storage analysis sheet with container and file share details"""
    
    # Title
    ws.merge_cells('A1:G1')
    ws['A1'] = "Azure Storage Resource Analysis"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center")
    
    current_row = 3
    
    # Blob Storage Section
    if container_results:
        ws[f'A{current_row}'] = "Blob Storage Containers"
        ws[f'A{current_row}'].font = Font(size=14, bold=True, color="3498DB")
        current_row += 2
        
        # Headers
        headers = ["Subscription", "Storage Account", "Container", "Blob Count", "Status"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header)
            ws.cell(row=current_row, column=col).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=current_row, column=col).fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
        
        current_row += 1
        
        # Container data
        for container in container_results:
            ws.cell(row=current_row, column=1, value=container.get('subscription_id', 'N/A'))
            ws.cell(row=current_row, column=2, value=container.get('account_name', 'N/A'))
            ws.cell(row=current_row, column=3, value=container.get('container_name', 'N/A'))
            
            blob_count = container.get('blob_count', 'ERROR')
            ws.cell(row=current_row, column=4, value=blob_count if blob_count != 'ERROR' else 0)
            
            status = "✅ Active" if blob_count != 'ERROR' and blob_count > 0 else "⚠️ Empty/Error"
            ws.cell(row=current_row, column=5, value=status)
            
            current_row += 1
        
        current_row += 2
    
    # Azure Files Section
    if file_share_results:
        ws[f'A{current_row}'] = "Azure File Shares"
        ws[f'A{current_row}'].font = Font(size=14, bold=True, color="3498DB")
        current_row += 2
        
        # Headers
        headers = ["Subscription", "Storage Account", "File Share", "File Count", "Status"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header)
            ws.cell(row=current_row, column=col).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=current_row, column=col).fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
        
        current_row += 1
        
        # File share data
        for share in file_share_results:
            ws.cell(row=current_row, column=1, value=share.get('subscription_id', 'N/A'))
            ws.cell(row=current_row, column=2, value=share.get('account_name', 'N/A'))
            ws.cell(row=current_row, column=3, value=share.get('share_name', 'N/A'))
            
            file_count = share.get('file_count', 'ERROR')
            ws.cell(row=current_row, column=4, value=file_count if file_count != 'ERROR' else 0)
            
            status = "✅ Active" if file_count != 'ERROR' and file_count > 0 else "⚠️ Empty/Error"
            ws.cell(row=current_row, column=5, value=status)
            
            current_row += 1

def _create_reservations_sheet(ws, reservations_analysis):
    """Create reservations analysis sheet"""
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "Azure Reserved Instances Analysis"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="8E44AD", end_color="8E44AD", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center")
    
    current_row = 3
    
    # Summary section
    summary = reservations_analysis.get('summary', {})
    if summary:
        ws[f'A{current_row}'] = "Reservation Summary"
        ws[f'A{current_row}'].font = Font(size=14, bold=True, color="8E44AD")
        current_row += 2
        
        summary_metrics = [
            ("Total Recommendations", summary.get('total_recommendations', 0)),
            ("Total Annual Savings", f"${summary.get('total_annual_savings', 0):,.2f}"),
            ("Total Upfront Investment", f"${summary.get('total_upfront_cost', 0):,.2f}"),
            ("Net Savings (Year 1)", f"${summary.get('net_savings_year_1', 0):,.2f}"),
            ("ROI Percentage", f"{summary.get('roi_percentage', 0):.1f}%"),
            ("High Priority Count", summary.get('high_priority_count', 0))
        ]
        
        for i, (metric, value) in enumerate(summary_metrics):
            ws.cell(row=current_row + i, column=1, value=metric)
            ws.cell(row=current_row + i, column=2, value=value)
            ws.cell(row=current_row + i, column=1).font = Font(bold=True)
            ws.cell(row=current_row + i, column=2).font = Font(color="8E44AD")
        
        current_row += len(summary_metrics) + 3
    
    # Detailed recommendations
    ws[f'A{current_row}'] = "Detailed Reservation Recommendations"
    ws[f'A{current_row}'].font = Font(size=14, bold=True, color="8E44AD")
    current_row += 2
    
    # Headers
    headers = ["Priority", "Service", "Type", "Term", "Capacity", "Monthly Savings", "Annual Savings", "Upfront Cost", "Confidence"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=current_row, column=col, value=header)
        ws.cell(row=current_row, column=col).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=current_row, column=col).fill = PatternFill(start_color="6C3483", end_color="6C3483", fill_type="solid")
    
    current_row += 1
    
    # Combine all reservation recommendations
    all_reservations = (reservations_analysis.get('blob_storage_reservations', []) + 
                       reservations_analysis.get('files_reservations', []))
    
    # Sort by priority and savings
    sorted_reservations = sorted(all_reservations, 
                               key=lambda x: (
                                   {'High': 3, 'Medium': 2, 'Low': 1}.get(x.get('recommendation_priority', 'Low'), 1),
                                   x.get('annual_savings', 0)
                               ), 
                               reverse=True)
    
    for reservation in sorted_reservations:
        priority = reservation.get('recommendation_priority', 'Low')
        service = reservation.get('service', 'Unknown')
        res_type = reservation.get('reservation_type', 'Unknown')
        term = reservation.get('term', 'Unknown')
        capacity = f"{reservation.get('capacity_tb', 0)} TB"
        monthly_savings = f"${reservation.get('monthly_savings', 0):,.2f}"
        annual_savings = f"${reservation.get('annual_savings', 0):,.2f}"
        upfront = f"${reservation.get('upfront_cost', 0):,.2f}"
        confidence = reservation.get('confidence', 'Unknown')
        
        ws.cell(row=current_row, column=1, value=priority)
        ws.cell(row=current_row, column=2, value=service)
        ws.cell(row=current_row, column=3, value=res_type)
        ws.cell(row=current_row, column=4, value=term)
        ws.cell(row=current_row, column=5, value=capacity)
        ws.cell(row=current_row, column=6, value=monthly_savings)
        ws.cell(row=current_row, column=7, value=annual_savings)
        ws.cell(row=current_row, column=8, value=upfront)
        ws.cell(row=current_row, column=9, value=confidence)
        
        # Color code by priority
        if priority == 'High':
            for col in range(1, 10):
                ws.cell(row=current_row, column=col).fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
        elif priority == 'Medium':
            for col in range(1, 10):
                ws.cell(row=current_row, column=col).fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")
        
        current_row += 1

def _create_savings_plans_sheet(ws, savings_plans_analysis):
    """Create savings plans analysis sheet"""
    
    # Title
    ws.merge_cells('A1:I1')
    ws['A1'] = "Azure Savings Plans Analysis"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center")
    
    current_row = 3
    
    # Summary section
    summary = savings_plans_analysis.get('summary', {})
    if summary:
        ws[f'A{current_row}'] = "Savings Plans Summary"
        ws[f'A{current_row}'].font = Font(size=14, bold=True, color="F39C12")
        current_row += 2
        
        summary_metrics = [
            ("Total Plans Recommended", summary.get('total_plans', 0)),
            ("Total Annual Savings", f"${summary.get('total_annual_savings', 0):,.2f}"),
            ("Total Annual Commitment", f"${summary.get('total_annual_commitment', 0):,.2f}"),
            ("Average Savings %", f"{summary.get('average_savings_percentage', 0):.1f}%"),
            ("Compute Plans", summary.get('compute_plans_count', 0)),
            ("Azure Plans", summary.get('azure_plans_count', 0))
        ]
        
        for i, (metric, value) in enumerate(summary_metrics):
            ws.cell(row=current_row + i, column=1, value=metric)
            ws.cell(row=current_row + i, column=2, value=value)
            ws.cell(row=current_row + i, column=1).font = Font(bold=True)
            ws.cell(row=current_row + i, column=2).font = Font(color="F39C12")
        
        current_row += len(summary_metrics) + 3
    
    # Detailed recommendations
    ws[f'A{current_row}'] = "Detailed Savings Plans Recommendations"
    ws[f'A{current_row}'].font = Font(size=14, bold=True, color="F39C12")
    current_row += 2
    
    # Headers
    headers = ["Priority", "Plan Type", "Term", "Monthly Commitment", "Annual Savings", "Savings %", "Flexibility", "Confidence", "Services Covered"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=current_row, column=col, value=header)
        ws.cell(row=current_row, column=col).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=current_row, column=col).fill = PatternFill(start_color="D68910", end_color="D68910", fill_type="solid")
    
    current_row += 1
    
    # Combine all savings plan recommendations
    all_plans = (savings_plans_analysis.get('compute_savings_plans', []) + 
                savings_plans_analysis.get('azure_savings_plans', []))
    
    # Sort by priority and savings
    sorted_plans = sorted(all_plans, 
                         key=lambda x: (
                             {'High': 3, 'Medium': 2, 'Low': 1}.get(x.get('recommendation_priority', 'Low'), 1),
                             x.get('estimated_annual_savings', 0)
                         ), 
                         reverse=True)
    
    for plan in sorted_plans:
        priority = plan.get('recommendation_priority', 'Low')
        plan_type = plan.get('plan_type', 'Unknown')
        term = plan.get('term', 'Unknown')
        monthly_commitment = f"${plan.get('monthly_commitment', 0):,.2f}"
        annual_savings = f"${plan.get('estimated_annual_savings', 0):,.2f}"
        savings_pct = f"{plan.get('savings_percentage', 0):.1f}%"
        flexibility = plan.get('flexibility', 'Unknown')
        confidence = plan.get('confidence', 'Unknown')
        services = ', '.join(plan.get('covered_services', []))[:30] + "..." if len(', '.join(plan.get('covered_services', []))) > 30 else ', '.join(plan.get('covered_services', []))
        
        ws.cell(row=current_row, column=1, value=priority)
        ws.cell(row=current_row, column=2, value=plan_type)
        ws.cell(row=current_row, column=3, value=term)
        ws.cell(row=current_row, column=4, value=monthly_commitment)
        ws.cell(row=current_row, column=5, value=annual_savings)
        ws.cell(row=current_row, column=6, value=savings_pct)
        ws.cell(row=current_row, column=7, value=flexibility)
        ws.cell(row=current_row, column=8, value=confidence)
        ws.cell(row=current_row, column=9, value=services)
        
        # Color code by priority
        if priority == 'High':
            for col in range(1, 10):
                ws.cell(row=current_row, column=col).fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
        elif priority == 'Medium':
            for col in range(1, 10):
                ws.cell(row=current_row, column=col).fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")
        
        current_row += 1

def _create_cost_optimization_sheet(ws, cost_analysis, reservations_analysis, savings_plans_analysis):
    """Create consolidated cost optimization recommendations sheet"""
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = "Cost Optimization Action Plan"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center")
    
    current_row = 3
    
    # Total savings potential
    total_savings = 0
    if reservations_analysis:
        total_savings += reservations_analysis.get('total_potential_savings', 0)
    if savings_plans_analysis:
        sp_summary = savings_plans_analysis.get('summary', {})
        total_savings += sp_summary.get('total_annual_savings', 0)
    
    ws[f'A{current_row}'] = f"💰 TOTAL ANNUAL SAVINGS POTENTIAL: ${total_savings:,.2f}"
    ws[f'A{current_row}'].font = Font(size=14, bold=True, color="27AE60")
    current_row += 3
    
    # Priority recommendations
    ws[f'A{current_row}'] = "🎯 HIGH PRIORITY ACTIONS"
    ws[f'A{current_row}'].font = Font(size=14, bold=True, color="E74C3C")
    current_row += 2
    
    high_priority_actions = []
    
    # Add high-priority reservations
    if reservations_analysis:
        all_reservations = (reservations_analysis.get('blob_storage_reservations', []) + 
                           reservations_analysis.get('files_reservations', []))
        high_priority_reservations = [r for r in all_reservations if r.get('recommendation_priority') == 'High']
        
        for reservation in high_priority_reservations[:3]:  # Top 3
            action = f"Purchase {reservation.get('service', 'Storage')} reservation - {reservation.get('term', 'Unknown')} term"
            savings = f"${reservation.get('annual_savings', 0):,.0f}/year"
            investment = f"${reservation.get('upfront_cost', 0):,.0f} upfront"
            high_priority_actions.append((action, savings, investment))
    
    # Add high-priority savings plans
    if savings_plans_analysis:
        all_plans = (savings_plans_analysis.get('compute_savings_plans', []) + 
                    savings_plans_analysis.get('azure_savings_plans', []))
        high_priority_plans = [p for p in all_plans if p.get('recommendation_priority') == 'High']
        
        for plan in high_priority_plans[:2]:  # Top 2
            action = f"Implement {plan.get('plan_type', 'Savings Plan')} - {plan.get('term', 'Unknown')} term"
            savings = f"${plan.get('estimated_annual_savings', 0):,.0f}/year"
            investment = f"${plan.get('monthly_commitment', 0):,.0f}/month commitment"
            high_priority_actions.append((action, savings, investment))
    
    # Display high priority actions
    if high_priority_actions:
        headers = ["Action", "Annual Savings", "Investment Required"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header)
            ws.cell(row=current_row, column=col).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=current_row, column=col).fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        
        current_row += 1
        
        for action, savings, investment in high_priority_actions:
            ws.cell(row=current_row, column=1, value=action)
            ws.cell(row=current_row, column=2, value=savings)
            ws.cell(row=current_row, column=3, value=investment)
            
            # Highlight high savings
            if "savings" in savings and float(savings.replace('$', '').replace(',', '').replace('/year', '')) > 5000:
                for col in range(1, 4):
                    ws.cell(row=current_row, column=col).fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
            
            current_row += 1
        
        current_row += 2
    
    # Implementation timeline
    ws[f'A{current_row}'] = "📅 IMPLEMENTATION TIMELINE"
    ws[f'A{current_row}'].font = Font(size=14, bold=True, color="3498DB")
    current_row += 2
    
    timeline_items = [
        ("Week 1-2", "Review and validate reservation recommendations", "Finance team approval for upfront investments"),
        ("Week 3-4", "Implement high-priority reservations", "Purchase storage reservations through Azure portal"),
        ("Month 2", "Implement Savings Plans", "Set up hourly commitments and monitoring"),
        ("Month 3", "Monitor and optimize", "Track utilization and adjust commitments as needed"),
        ("Ongoing", "Regular reviews", "Monthly cost optimization reviews and adjustments")
    ]
    
    headers = ["Timeframe", "Action", "Details"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=current_row, column=col, value=header)
        ws.cell(row=current_row, column=col).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=current_row, column=col).fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
    
    current_row += 1
    
    for timeframe, action, details in timeline_items:
        ws.cell(row=current_row, column=1, value=timeframe)
        ws.cell(row=current_row, column=2, value=action)
        ws.cell(row=current_row, column=3, value=details)
        
        current_row += 1

def _create_detailed_data_sheet(ws, container_results, file_share_results):
    """Create detailed raw data sheet"""
    
    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = "Detailed Raw Data"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="95A5A6", end_color="95A5A6", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center")
    
    current_row = 3
    
    # Container details
    if container_results:
        ws[f'A{current_row}'] = "Container Analysis Results"
        ws[f'A{current_row}'].font = Font(size=12, bold=True)
        current_row += 2
        
        headers = ["Subscription ID", "Account Name", "Container Name", "Blob Count", "Analysis Status"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header)
            ws.cell(row=current_row, column=col).font = Font(bold=True)
        
        current_row += 1
        
        for container in container_results:
            ws.cell(row=current_row, column=1, value=container.get('subscription_id', 'N/A'))
            ws.cell(row=current_row, column=2, value=container.get('account_name', 'N/A'))
            ws.cell(row=current_row, column=3, value=container.get('container_name', 'N/A'))
            ws.cell(row=current_row, column=4, value=container.get('blob_count', 0))
            ws.cell(row=current_row, column=5, value="Success" if container.get('blob_count') != 'ERROR' else "Error")
            current_row += 1
        
        current_row += 2
    
    # File share details
    if file_share_results:
        ws[f'A{current_row}'] = "File Share Analysis Results"
        ws[f'A{current_row}'].font = Font(size=12, bold=True)
        current_row += 2
        
        headers = ["Subscription ID", "Account Name", "Share Name", "File Count", "Analysis Status"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header)
            ws.cell(row=current_row, column=col).font = Font(bold=True)
        
        current_row += 1
        
        for share in file_share_results:
            ws.cell(row=current_row, column=1, value=share.get('subscription_id', 'N/A'))
            ws.cell(row=current_row, column=2, value=share.get('account_name', 'N/A'))
            ws.cell(row=current_row, column=3, value=share.get('share_name', 'N/A'))
            ws.cell(row=current_row, column=4, value=share.get('file_count', 0))
            ws.cell(row=current_row, column=5, value="Success" if share.get('file_count') != 'ERROR' else "Error")
            current_row += 1

def _format_sheet(ws):
    """Apply consistent formatting to a worksheet"""
    
    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Set row height
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 20