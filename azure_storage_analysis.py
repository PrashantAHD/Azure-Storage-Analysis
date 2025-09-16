#!/usr/bin/env python3
"""
Azure Storage Analysis Tool

This script analyzes Azure Storage accounts and containers with concurrent processing
and generates comprehensive Excel reports including cost optimization recommendations.

"""

import os
import sys
import re
import gzip
import json
import fnmatch
import configparser
import subprocess
import logging
import concurrent.futures
import argparse
from datetime import datetime, timedelta, timezone
from collections import defaultdict

# Azure SDK imports

from azure.identity import AzureCliCredential, DefaultAzureCredential, InteractiveBrowserCredential
from azure.mgmt.resource import ResourceManagementClient
from azure.mgmt.subscription import SubscriptionClient
from azure.mgmt.storage import StorageManagementClient
from azure.storage.blob import BlobServiceClient, ContainerClient
from azure.core.exceptions import ResourceNotFoundError, ClientAuthenticationError
from azure.storage.fileshare import ShareServiceClient


from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference, Series, PieChart

# Modern presentation-ready style palette
class PresentationStyles:
    # Modern blue gradient for main headers
    HEADER_FILL = PatternFill(start_color="4F81BD", end_color="1F497D", fill_type="solid")
    HEADER_FONT = Font(name="Segoe UI", bold=True, color="FFFFFF", size=12)
    # Accent for section titles
    SECTION_FILL = PatternFill(start_color="F79646", end_color="FFC000", fill_type="solid")
    SECTION_FONT = Font(name="Segoe UI", bold=True, color="FFFFFF", size=13)
    # Title style
    TITLE_FONT = Font(name="Segoe UI", bold=True, color="305496", size=16)
    # Subtle gray for summary rows
    SUMMARY_FILL = PatternFill(start_color="E2EFDA", end_color="D9E1F2", fill_type="solid")
    SUMMARY_FONT = Font(name="Segoe UI", bold=True, color="305496", size=12)
    # Border for all cells
    BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    # Alignment
    CENTER = Alignment(horizontal="center", vertical="center")
    RIGHT = Alignment(horizontal="right", vertical="center")
    LEFT = Alignment(horizontal="left", vertical="center")

# Configure the logger
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# ============================================================================
# AZURE AUTHENTICATION AND MANAGEMENT FUNCTIONS
# ============================================================================

def get_available_azure_subscriptions():
    """
    Get a list of available Azure subscriptions from Azure CLI.
    
    Returns:
        list: List of available subscription dictionaries with id and name
    """
    subscriptions = []
    
    try:
        # Try to use Azure CLI to list subscriptions
        result = subprocess.run(['az', 'account', 'list', '--output', 'json'], 
                                capture_output=True, text=True)
        
        if result.returncode == 0:
            subscription_data = json.loads(result.stdout)
            for sub in subscription_data:
                subscriptions.append({
                    'id': sub['id'],
                    'name': sub['name'],
                    'state': sub.get('state', 'Unknown'),
                    'is_default': sub.get('isDefault', False)
                })
        else:
            logger.warning(f"Error running Azure CLI: {result.stderr}")
    except FileNotFoundError:
        logger.warning("Azure CLI not found. Make sure it's installed and in your PATH.")
    except json.JSONDecodeError:
        logger.warning("Error parsing Azure CLI output")
    
    # If we couldn't get subscriptions from CLI, try using the SDK
    if not subscriptions:
        try:
            credential = DefaultAzureCredential()
            subscription_client = SubscriptionClient(credential)
            
            for sub in subscription_client.subscriptions.list():
                subscriptions.append({
                    'id': sub.subscription_id,
                    'name': sub.display_name,
                    'state': sub.state,
                    'is_default': False  # Can't determine this from the SDK
                })
        except Exception as e:
            logger.error(f"Error listing subscriptions with SDK: {e}")
    
    return subscriptions

def select_azure_subscription():
    """
    Display available Azure subscriptions and let the user select one.
    
    Returns:
        str: Selected subscription ID or None if canceled
    """
    # Get available subscriptions
    subscriptions = get_available_azure_subscriptions()
    
    if not subscriptions:
        logger.warning("No Azure subscriptions found. Please configure Azure CLI first.")
        return None
    
    # Get the current subscription from Azure CLI
    try:
        result = subprocess.run(['az', 'account', 'show', '--output', 'json'], 
                                capture_output=True, text=True)
        
        if result.returncode == 0:
            current_sub = json.loads(result.stdout)
            current_sub_id = current_sub.get('id')
        else:
            current_sub_id = None
    except Exception:
        current_sub_id = None
    
    # Display subscriptions
    print("\n" + "="*80)
    print(" SELECT AZURE SUBSCRIPTION ".center(80, "="))
    print("="*80)
    print(f"Available Azure subscriptions ({len(subscriptions)}):")
    
    for i, sub in enumerate(subscriptions, 1):
        # Mark the current subscription
        current_marker = " (current)" if sub['id'] == current_sub_id else ""
        # Format the display line
        print(f"{i}. {sub['name']}{current_marker} - ID: {sub['id']}, State: {sub['state']}")
    
    # Add option to keep current subscription
    if current_sub_id:
        print(f"\nC. Continue with current subscription")
    
    # Add option to cancel
    print("X. Cancel and exit")
    
    # Get user selection
    while True:
        choice = input("\nEnter your choice (number, 'c', or 'x'): ").strip().lower()
        
        if choice == 'x':
            return None  # User canceled
        elif choice == 'c' and current_sub_id:
            return current_sub_id  # Keep current subscription
        else:
            try:
                idx = int(choice)
                if 1 <= idx <= len(subscriptions):
                    return subscriptions[idx-1]['id']
                else:
                    print(f"Please enter a number between 1 and {len(subscriptions)}")
            except ValueError:
                print("Please enter a valid choice")

def check_and_login_to_azure(auto_mode=False):
    """
    Check if the user is logged in to Azure and prompt for login if not.
    
    Args:
        auto_mode (bool, optional): If True, skip interactive prompts. Defaults to False.
    
    Returns:
        bool: True if login was successful
    """
    # Check current login status
    try:
        result = subprocess.run(['az', 'account', 'show'], 
                              capture_output=True, text=True)
        
        if result.returncode == 0:
            logger.info("Already logged in to Azure")
            return True
    except Exception:
        pass
    
    # In auto mode, cannot proceed without existing login
    if auto_mode:
        logger.error("Auto mode requires existing Azure CLI login. Please run 'az login' first.")
        return False
    
    # Not logged in, prompt user to login
    print("\n" + "="*80)
    print(" AZURE LOGIN REQUIRED ".center(80, "="))
    print("="*80)
    print("\nYou need to log in to Azure to continue.")
    print("\nPlease choose a login method:")
    print("1. Interactive browser login")
    print("2. Use Azure CLI login")
    print("X. Cancel and exit")
    
    while True:
        choice = input("\nEnter your choice (1, 2, or 'x'): ").strip().lower()
        
        if choice == 'x':
            return False  # User canceled
        elif choice == '1':
            try:
                # Use interactive browser login
                print("\nLaunching browser for interactive login...")
                credential = InteractiveBrowserCredential()
                # Force token acquisition to verify login
                token = credential.get_token("https://management.azure.com/.default")
                if token:
                    logger.info("Successfully logged in via browser")
                    return True
            except Exception as e:
                logger.error(f"Error during interactive login: {e}")
        elif choice == '2':
            try:
                # Use Azure CLI login
                print("\nLaunching Azure CLI login...")
                result = subprocess.run(['az', 'login'], 
                                      capture_output=False, text=True)
                
                if result.returncode == 0:
                    logger.info("Successfully logged in via Azure CLI")
                    return True
                else:
                    logger.error("Azure CLI login failed")
            except Exception as e:
                logger.error(f"Error during Azure CLI login: {e}")
        else:
            print("Please enter 1, 2, or 'x'")
    
    return False

def initialize_azure_clients(subscription_id=None, auto_mode=False):
    """
    Initialize Azure clients with optional subscription selection.
    
    Args:
        subscription_id (str, optional): Azure subscription ID. Defaults to None.
        auto_mode (bool, optional): If True, skip interactive prompts. Defaults to False.
    
    Returns:
        tuple: (credential, subscription_id, resource_client, storage_client)
    """
    # Ensure the user is logged in
    if not check_and_login_to_azure(auto_mode):
        logger.error("Azure login required to continue. Exiting.")
        sys.exit(1)
    
    # Handle subscription selection
    if not subscription_id:
        if auto_mode:
            # In auto mode, use the current/default subscription
            try:
                result = subprocess.run(['az', 'account', 'show', '--output', 'json'], 
                                      capture_output=True, text=True)
                if result.returncode == 0:
                    current_sub = json.loads(result.stdout)
                    subscription_id = current_sub.get('id')
                    logger.info(f"Auto mode: Using current subscription {subscription_id}")
                else:
                    logger.error("Could not determine current subscription in auto mode")
                    sys.exit(1)
            except Exception as e:
                logger.error(f"Error getting current subscription in auto mode: {e}")
                sys.exit(1)
        else:
            subscription_id = select_azure_subscription()
            if not subscription_id:
                logger.error("Subscription selection canceled. Exiting.")
                sys.exit(1)
    
    # Set subscription as active in Azure CLI
    try:
        subprocess.run(['az', 'account', 'set', '--subscription', subscription_id], 
                      capture_output=True, check=True)
        logger.info(f"Set active subscription to {subscription_id}")
    except Exception as e:
        logger.warning(f"Error setting active subscription in Azure CLI: {e}")
    
    # Initialize credential and clients
    try:
        # Try to use AzureCliCredential first (works well with CLI login)
        try:
            credential = AzureCliCredential()
            # Test the credential
            subscription_client = SubscriptionClient(credential)
            test_sub = next(subscription_client.subscriptions.list())
            logger.info("Using AzureCliCredential")
        except Exception:
            # Fall back to DefaultAzureCredential
            credential = DefaultAzureCredential()
            logger.info("Using DefaultAzureCredential")
        
        # Initialize resource client
        resource_client = ResourceManagementClient(credential, subscription_id)
        
        # Initialize storage client
        storage_client = StorageManagementClient(credential, subscription_id)
        
        return credential, subscription_id, resource_client, storage_client
    
    except Exception as e:
        logger.error(f"Error initializing Azure clients: {e}")
        sys.exit(1)

def get_all_storage_accounts(storage_client):
    """
    Get all storage accounts in the subscription.
    
    Args:
        storage_client (StorageManagementClient): Azure Storage Management client
    
    Returns:
        list: List of storage accounts
    """
    try:
        storage_accounts = list(storage_client.storage_accounts.list())
        logger.info(f"Found {len(storage_accounts)} storage accounts")
        return storage_accounts
    except Exception as e:
        logger.error(f"Error listing storage accounts: {e}")
        return []

def select_storage_accounts_to_process(storage_accounts, auto_mode=False, account_names=None, account_pattern=None, max_accounts=None):
    """
    Allow user to select the storage accounts to process.
    
    Args:
        storage_accounts (list): List of all available storage accounts
        auto_mode (bool, optional): If True, process all accounts automatically. Defaults to False.
        account_names (list, optional): Specific account names to process. Defaults to None.
        account_pattern (str, optional): Pattern to match account names. Defaults to None.
        max_accounts (int, optional): Maximum number of accounts to process. Defaults to None.
    
    Returns:
        list: Selected storage accounts to process
    """
    total_accounts = len(storage_accounts)
    logger.info(f"Total available storage accounts: {total_accounts}")
    
    # Auto mode or specific filters provided
    if auto_mode or account_names or account_pattern:
        if account_names:
            # Filter by specific account names
            selected_accounts = []
            for account_name in account_names:
                matching_accounts = [a for a in storage_accounts if a.name.lower() == account_name.lower()]
                if matching_accounts:
                    selected_accounts.extend(matching_accounts)
                else:
                    logger.warning(f"Storage account '{account_name}' not found in subscription")
            
            if not selected_accounts:
                logger.warning("No valid storage accounts found from specified names. Processing all accounts.")
                selected_accounts = storage_accounts
        
        elif account_pattern:
            # Filter by pattern
            selected_accounts = []
            for account in storage_accounts:
                if fnmatch.fnmatch(account.name.lower(), account_pattern.lower()):
                    selected_accounts.append(account)
            
            if not selected_accounts:
                logger.warning(f"No storage accounts matched pattern '{account_pattern}'. Processing all accounts.")
                selected_accounts = storage_accounts
        
        else:
            # Process all accounts
            selected_accounts = storage_accounts
        
        # Apply max_accounts limit
        if max_accounts and len(selected_accounts) > max_accounts:
            logger.info(f"Limiting to first {max_accounts} storage accounts")
            selected_accounts = selected_accounts[:max_accounts]
        
        account_names_list = [a.name for a in selected_accounts]
        logger.info(f"Auto mode: Processing {len(selected_accounts)} storage accounts: {', '.join(account_names_list[:3])}" + 
                   (f" and {len(account_names_list) - 3} more" if len(account_names_list) > 3 else ""))
        return selected_accounts
    
    # Interactive mode
    # Prompt for selection method
    print("\nHow would you like to select storage accounts to process?")
    print("1. Process all storage accounts")
    print("2. Process a specific number of storage accounts")
    print("3. Select specific storage accounts by number")
    print("4. Enter specific storage account names")
    print("5. Enter storage account name patterns (using wildcards)")
    
    selection_method = 0
    while selection_method not in [1, 2, 3, 4, 5]:
        try:
            selection_method = int(input("Enter your choice (1-5): ").strip())
            if selection_method not in [1, 2, 3, 4, 5]:
                logger.warning("Invalid choice. Please enter a number between 1 and 5.")
        except ValueError:
            logger.warning("Invalid input. Please enter a number.")
    
    # Process all storage accounts
    if selection_method == 1:
        logger.info(f"Processing all {total_accounts} storage accounts.")
        return storage_accounts
    
    # Process a specific number of storage accounts
    elif selection_method == 2:
        number_to_process = 0
        while number_to_process < 1 or number_to_process > total_accounts:
            try:
                number_to_process = int(input(f"Enter number of storage accounts to process (1-{total_accounts}): ").strip())
                if number_to_process < 1 or number_to_process > total_accounts:
                    logger.warning(f"Invalid number. Please enter a value between 1 and {total_accounts}.")
            except ValueError:
                logger.warning("Invalid input. Please enter a number.")
        
        logger.info(f"Processing the first {number_to_process} storage accounts.")
        return storage_accounts[:number_to_process]
    
    # Select specific storage accounts by number
    elif selection_method == 3:
        print("\nAvailable storage accounts:")
        for idx, account in enumerate(storage_accounts, start=1):
            print(f"{idx}. {account.name} (Location: {account.location}, Kind: {account.kind})")
        
        selected_indices = []
        valid_selection = False
        
        while not valid_selection:
            try:
                selection_input = input("Enter storage account numbers to process (comma-separated, e.g., 1,3,5): ").strip()
                indices = [int(idx.strip()) for idx in selection_input.split(',')]
                
                # Validate indices
                invalid_indices = [idx for idx in indices if idx < 1 or idx > total_accounts]
                if invalid_indices:
                    logger.warning(f"Invalid account numbers: {invalid_indices}. Please enter values between 1 and {total_accounts}.")
                else:
                    selected_indices = indices
                    valid_selection = True
            except ValueError:
                logger.warning("Invalid input. Please enter numbers separated by commas.")
        
        selected_accounts = [storage_accounts[idx-1] for idx in selected_indices]
        logger.info(f"Processing {len(selected_accounts)} selected storage accounts.")
        return selected_accounts
    
    # Enter specific storage account names
    elif selection_method == 4:
        print("\nEnter specific storage account names to process (one per line).")
        print("Enter an empty line when finished.")
        
        account_names = []
        while True:
            account_name = input("Storage account name (or empty to finish): ").strip()
            if not account_name:
                break
                
            # Check if account exists
            matching_accounts = [a for a in storage_accounts if a.name.lower() == account_name.lower()]
            if matching_accounts:
                account_names.extend(matching_accounts)
            else:
                logger.warning(f"Storage account '{account_name}' not found in your subscription. Please check the name.")
        
        if not account_names:
            logger.warning("No valid storage accounts selected. Processing all storage accounts.")
            return storage_accounts
        
        logger.info(f"Processing {len(account_names)} specified storage accounts.")
        return account_names
    
    # Enter storage account name patterns (using wildcards)
    elif selection_method == 5:
        print("\nEnter storage account name patterns to match (e.g., 'prod-*' or '*-backup').")
        print("Use * as a wildcard to match any characters.")
        print("Enter an empty line when finished.")
        
        patterns = []
        while True:
            pattern = input("Storage account pattern (or empty to finish): ").strip()
            if not pattern:
                break
            patterns.append(pattern)
        
        if not patterns:
            logger.warning("No patterns provided. Processing all storage accounts.")
            return storage_accounts
        
        # Match storage accounts against patterns
        matched_accounts = set()
        for pattern in patterns:
            for account in storage_accounts:
                if fnmatch.fnmatch(account.name.lower(), pattern.lower()):
                    matched_accounts.add(account)
        
        matched_accounts = list(matched_accounts)
        
        if not matched_accounts:
            logger.warning("No storage accounts matched the provided patterns. Processing all storage accounts.")
            return storage_accounts
        
        account_names = [a.name for a in matched_accounts]
        logger.info(f"Found {len(matched_accounts)} storage accounts matching the patterns: {', '.join(account_names)}")
        return matched_accounts

def get_storage_account_connection_string(storage_client, resource_group_name, account_name):
    """
    Get the connection string for a storage account.
    
    Args:
        storage_client (StorageManagementClient): Azure Storage Management client
        resource_group_name (str): Resource group name
        account_name (str): Storage account name
    
    Returns:
        str: Storage account connection string
    """
    try:
        # Get storage account keys
        keys = storage_client.storage_accounts.list_keys(resource_group_name, account_name)
        key = keys.keys[0].value
        
        # Create connection string
        connection_string = f"DefaultEndpointsProtocol=https;AccountName={account_name};AccountKey={key};EndpointSuffix=core.windows.net"
        return connection_string
    except Exception as e:
        logger.error(f"Error getting connection string for {account_name}: {e}")
        return None
      
# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def human_readable_size(size_bytes):
    """
    Convert bytes to human-readable format.
    
    Args:
        size_bytes (int): Size in bytes
    
    Returns:
        tuple: (value, unit) for human-readable size
    """
    for unit in ['B', 'KB', 'MB', 'GB', 'TB', 'PB']:
        if size_bytes < 1024.0:
            return (round(size_bytes, 2), unit)
        size_bytes /= 1024.0
    return (round(size_bytes, 2), 'PB')

# ============================================================================
# CONTAINER ANALYSIS FUNCTIONS
# ============================================================================

def analyze_container(blob_service_client, container_name, account_name, subscription_id=None):
    """
    Analyze a single Azure Storage container and collect statistics.
    
    Args:
        blob_service_client (BlobServiceClient): Azure Blob Service client
        container_name (str): Name of the container to analyze
        account_name (str): Storage account name
        subscription_id (str, optional): Azure subscription ID. Defaults to None.
    
    Returns:
        dict: Container analysis results or None
    """
    try:
        container_client = blob_service_client.get_container_client(container_name)
        
        # Initialize variables for size and object count
        total_size = 0
        total_blob_count = 0
        
        # Object size categories
        small_blob_count = 0
        small_blob_size = 0
        large_blob_count = 0
        large_blob_size = 0
        
        between_30_and_90_days = 0
        between_90_and_180_days = 0
        not_accessed_90_days = 0
        not_accessed_180_days = 0
        
        # Access pattern tracking
        hot_blobs = 0  # Modified in last 30 days
        warm_blobs = 0  # Modified between 30-90 days
        cold_blobs = 0  # Modified between 90-180 days
        archive_blobs = 0  # Not modified in 180+ days
        total_hot_size = 0
        total_warm_size = 0
        total_cold_size = 0
        total_archive_size = 0

        # Define size thresholds (in bytes)
        MAX_SMALL_BLOB_SIZE = 1024 * 1024  # 1 MB

        # Define time thresholds with UTC timezone
        now = datetime.now(timezone.utc)
        threshold_30_days = now - timedelta(days=30)
        threshold_90_days = now - timedelta(days=90)
        threshold_180_days = now - timedelta(days=180)

        # List all blobs in the container
        blobs = container_client.list_blobs()
        
        for blob in blobs:
            blob_size = blob.size
            last_modified = blob.last_modified

            total_size += blob_size
            total_blob_count += 1
            
            # Categorize blobs by size
            if blob_size <= MAX_SMALL_BLOB_SIZE:
                # Objects up to 1 MB
                small_blob_count += 1
                small_blob_size += blob_size
            else:
                # Objects larger than 1 MB
                large_blob_count += 1
                large_blob_size += blob_size
            
            # Check blob's last modified timestamp
            if last_modified >= threshold_30_days:
                hot_blobs += 1
                total_hot_size += blob_size
            elif threshold_90_days <= last_modified < threshold_30_days:
                between_30_and_90_days += 1
                warm_blobs += 1
                total_warm_size += blob_size
            elif threshold_180_days <= last_modified < threshold_90_days:
                between_90_and_180_days += 1
                cold_blobs += 1
                total_cold_size += blob_size
                not_accessed_90_days += 1
            else:  # last_modified < threshold_180_days
                not_accessed_90_days += 1
                not_accessed_180_days += 1
                archive_blobs += 1
                total_archive_size += blob_size

        # Calculate percentages for age metrics
        pct_30_90_days = (between_30_and_90_days / total_blob_count * 100) if total_blob_count > 0 else 0
        pct_over_90_days = (not_accessed_90_days / total_blob_count * 100) if total_blob_count > 0 else 0
        pct_over_180_days = (not_accessed_180_days / total_blob_count * 100) if total_blob_count > 0 else 0
        
        # Calculate access pattern percentages
        pct_hot_blobs = (hot_blobs / total_blob_count * 100) if total_blob_count > 0 else 0
        pct_warm_blobs = (warm_blobs / total_blob_count * 100) if total_blob_count > 0 else 0
        pct_cold_blobs = (cold_blobs / total_blob_count * 100) if total_blob_count > 0 else 0
        pct_archive_blobs = (archive_blobs / total_blob_count * 100) if total_blob_count > 0 else 0
        
        pct_hot_size = (total_hot_size / total_size * 100) if total_size > 0 else 0
        pct_warm_size = (total_warm_size / total_size * 100) if total_size > 0 else 0
        pct_cold_size = (total_cold_size / total_size * 100) if total_size > 0 else 0
        pct_archive_size = (total_archive_size / total_size * 100) if total_size > 0 else 0

        # Log detailed container information
        hr_total_size = human_readable_size(total_size)
        hr_small_size = human_readable_size(small_blob_size)
        hr_large_size = human_readable_size(large_blob_size)
        
        logger.info(f"Container: {container_name} in account {account_name}")
        logger.info(f"Total Size: {hr_total_size[0]} {hr_total_size[1]}")
        logger.info(f"Blobs Between 0KB and 1MB: {small_blob_count}")
        logger.info(f"Blobs Between 0KB and 1MB Total Size: {hr_small_size[0]} {hr_small_size[1]}")
        logger.info(f"Large Blobs (>1MB) Total Size: {hr_large_size[0]} {hr_large_size[1]}")
        logger.info(f"Number of Large Blobs (>1MB): {large_blob_count}")
        logger.info(f"Total Number of Blobs: {total_blob_count}")
        logger.info(f"Blobs between 30 and 90 days old: {between_30_and_90_days} ({pct_30_90_days:.1f}%)")
        logger.info(f"Blobs between 90 and 180 days old: {between_90_and_180_days} ({(between_90_and_180_days/total_blob_count*100 if total_blob_count > 0 else 0):.1f}%)")
        logger.info(f"Blobs not accessed >= 90 days: {not_accessed_90_days} ({pct_over_90_days:.1f}%)")
        logger.info(f"Blobs not accessed >= 180 days: {not_accessed_180_days} ({pct_over_180_days:.1f}%)")
        logger.info("-" * 40)

        return {
            'container_name': container_name,
            'account_name': account_name,
            'subscription_id': subscription_id,
            'total_size': total_size,
            'small_blob_count': small_blob_count,
            'small_blob_size': small_blob_size,
            'large_blob_count': large_blob_count,
            'large_blob_size': large_blob_size,
            'total_blob_count': total_blob_count,
            'between_30_and_90_days': between_30_and_90_days,
            'between_90_and_180_days': between_90_and_180_days,
            'not_accessed_90_days': not_accessed_90_days,
            'not_accessed_180_days': not_accessed_180_days,
            'pct_30_90_days': pct_30_90_days,
            'pct_over_90_days': pct_over_90_days,
            'pct_over_180_days': pct_over_180_days,
            'hot_blobs': hot_blobs,
            'warm_blobs': warm_blobs,
            'cold_blobs': cold_blobs,
            'archive_blobs': archive_blobs,
            'total_hot_size': total_hot_size,
            'total_warm_size': total_warm_size,
            'total_cold_size': total_cold_size,
            'total_archive_size': total_archive_size,
            'pct_hot_blobs': pct_hot_blobs,
            'pct_warm_blobs': pct_warm_blobs,
            'pct_cold_blobs': pct_cold_blobs,
            'pct_archive_blobs': pct_archive_blobs,
            'pct_hot_size': pct_hot_size,
            'pct_warm_size': pct_warm_size,
            'pct_cold_size': pct_cold_size,
            'pct_archive_size': pct_archive_size,
            'client': blob_service_client.get_container_client(container_name)  # Add client for detailed export
        }

    except Exception as e:
        logger.error(f"Error analyzing container {container_name} in account {account_name}: {e}")
        return None

def select_containers_to_process(storage_client, account, auto_mode=False, container_names=None, container_pattern=None, max_containers_per_account=None):
    """
    Allow user to select which containers to process in a storage account.
    
    Args:
        storage_client (StorageManagementClient): Azure Storage Management client
        account (StorageAccount): Storage account object
        auto_mode (bool, optional): If True, process all containers automatically. Defaults to False.
        container_names (list, optional): Specific container names to process. Defaults to None.
        container_pattern (str, optional): Pattern to match container names. Defaults to None.
        max_containers_per_account (int, optional): Maximum containers to process per account. Defaults to None.
    
    Returns:
        list: List of (container_name, blob_service_client) tuples
    """
    try:
        # Get resource group from account ID
        resource_group = account.id.split('/')[4]
        
        # Get connection string
        conn_string = get_storage_account_connection_string(storage_client, resource_group, account.name)
        if not conn_string:
            logger.error(f"Could not get connection string for account {account.name}")
            return []
        
        # Create blob service client
        blob_service_client = BlobServiceClient.from_connection_string(conn_string)
        
        # List all containers
        containers = list(blob_service_client.list_containers())
        
        if not containers:
            logger.info(f"No containers found in account {account.name}")
            return []
        
        logger.info(f"Found {len(containers)} containers in account {account.name}")
        
        # Auto mode or specific filters provided
        if auto_mode or container_names or container_pattern:
            selected_containers = []
            
            if container_names:
                # Filter by specific container names
                for container_name in container_names:
                    matching_containers = [c for c in containers if c.name.lower() == container_name.lower()]
                    if matching_containers:
                        selected_containers.extend([(c.name, blob_service_client) for c in matching_containers])
                    else:
                        logger.warning(f"Container '{container_name}' not found in account {account.name}")
                
                if not selected_containers:
                    logger.warning(f"No valid containers found from specified names in account {account.name}. Processing all containers.")
                    selected_containers = [(container.name, blob_service_client) for container in containers]
            
            elif container_pattern:
                # Filter by pattern
                for container in containers:
                    if fnmatch.fnmatch(container.name.lower(), container_pattern.lower()):
                        selected_containers.append((container.name, blob_service_client))
                
                if not selected_containers:
                    logger.warning(f"No containers matched pattern '{container_pattern}' in account {account.name}. Processing all containers.")
                    selected_containers = [(container.name, blob_service_client) for container in containers]
            
            else:
                # Process all containers
                selected_containers = [(container.name, blob_service_client) for container in containers]
            
            # Apply max_containers_per_account limit
            if max_containers_per_account and len(selected_containers) > max_containers_per_account:
                logger.info(f"Limiting to first {max_containers_per_account} containers in account {account.name}")
                selected_containers = selected_containers[:max_containers_per_account]
            
            container_names_list = [c[0] for c in selected_containers]
            logger.info(f"Auto mode: Processing {len(selected_containers)} containers in account {account.name}: {', '.join(container_names_list[:3])}" + 
                       (f" and {len(container_names_list) - 3} more" if len(container_names_list) > 3 else ""))
            return selected_containers
        
        # Interactive mode
        # Prompt for selection method
        print(f"\nHow would you like to select containers in account {account.name}?")
        print("1. Process all containers")
        print("2. Select specific containers by number")
        print("3. Enter container name patterns (using wildcards)")
        
        selection_method = 0
        while selection_method not in [1, 2, 3]:
            try:
                selection_method = int(input("Enter your choice (1-3): ").strip())
                if selection_method not in [1, 2, 3]:
                    logger.warning("Invalid choice. Please enter a number between 1 and 3.")
            except ValueError:
                logger.warning("Invalid input. Please enter a number.")
        
        # Process all containers
        if selection_method == 1:
            logger.info(f"Processing all {len(containers)} containers in account {account.name}")
            return [(container.name, blob_service_client) for container in containers]
        
        # Select specific containers by number
        elif selection_method == 2:
            print("\nAvailable containers:")
            for idx, container in enumerate(containers, start=1):
                # Try to get container properties
                try:
                    container_client = blob_service_client.get_container_client(container.name)
                    blobs = list(container_client.list_blobs(max_results=1))
                    blob_count = "1+" if blobs else "0"
                except Exception:
                    blob_count = "?"
                
                print(f"{idx}. {container.name} (Blobs: {blob_count})")
            
            selected_indices = []
            valid_selection = False
            
            while not valid_selection:
                try:
                    selection_input = input("Enter container numbers to process (comma-separated, e.g., 1,3,5): ").strip()
                    indices = [int(idx.strip()) for idx in selection_input.split(',')]
                    
                    # Validate indices
                    invalid_indices = [idx for idx in indices if idx < 1 or idx > len(containers)]
                    if invalid_indices:
                        logger.warning(f"Invalid container numbers: {invalid_indices}. Please enter values between 1 and {len(containers)}.")
                    else:
                        selected_indices = indices
                        valid_selection = True
                except ValueError:
                    logger.warning("Invalid input. Please enter numbers separated by commas.")
            
            selected_containers = [(containers[idx-1].name, blob_service_client) for idx in selected_indices]
            logger.info(f"Processing {len(selected_containers)} selected containers")
            return selected_containers
        
        # Enter container name patterns (using wildcards)
        elif selection_method == 3:
            print("\nEnter container name patterns to match (e.g., 'prod-*' or '*-backup').")
            print("Use * as a wildcard to match any characters.")
            print("Enter an empty line when finished.")
            
            patterns = []
            while True:
                pattern = input("Container pattern (or empty to finish): ").strip()
                if not pattern:
                    break
                patterns.append(pattern)
            
            if not patterns:
                logger.warning("No patterns provided. Processing all containers.")
                return [(container.name, blob_service_client) for container in containers]
            
            # Match containers against patterns
            matched_containers = []
            for pattern in patterns:
                for container in containers:
                    if fnmatch.fnmatch(container.name.lower(), pattern.lower()):
                        matched_containers.append((container.name, blob_service_client))
            
            # Remove duplicates
            unique_containers = list({c[0]: c for c in matched_containers}.values())
            
            if not unique_containers:
                logger.warning("No containers matched the provided patterns. Processing all containers.")
                return [(container.name, blob_service_client) for container in containers]
            
            container_names = [c[0] for c in unique_containers]
            logger.info(f"Found {len(unique_containers)} containers matching the patterns: {', '.join(container_names[:5])}" + 
                       (f" and {len(container_names) - 5} more" if len(container_names) > 5 else ""))
            return unique_containers
    
    except Exception as e:
        logger.error(f"Error selecting containers for account {account.name}: {e}")
        return []

def process_container_with_retry(blob_service_client, container_name, account_name, subscription_id=None):
    """
    Process a container with automatic credential refresh if needed.
    
    Args:
        blob_service_client (BlobServiceClient): Azure Blob Service client
        container_name (str): Container name
        account_name (str): Storage account name
        subscription_id (str, optional): Azure subscription ID. Defaults to None.
    
    Returns:
        dict: Container analysis result or None
    """
    try:
        return analyze_container(blob_service_client, container_name, account_name, subscription_id)
    except ClientAuthenticationError:
        logger.warning(f"Authentication error while processing container {container_name}. Refreshing credentials...")
        
        # Re-authenticate and try again
        try:
            # Re-initialize the clients
            credential, sub_id, _, storage_client = initialize_azure_clients(subscription_id)
            
            # Get the resource group name for the storage account
            storage_accounts = list(storage_client.storage_accounts.list())
            account = next((acc for acc in storage_accounts if acc.name == account_name), None)
            
            if not account:
                logger.error(f"Storage account {account_name} not found after refreshing credentials")
                return None
            
            # Extract resource group from ID
            resource_group = account.id.split('/')[4]
            
            # Get connection string
            conn_string = get_storage_account_connection_string(storage_client, resource_group, account_name)
            if not conn_string:
                logger.error(f"Could not get connection string for {account_name}")
                return None
            
            # Create new blob service client
            new_blob_service_client = BlobServiceClient.from_connection_string(conn_string)
            
            # Retry with new credentials
            logger.info(f"Retrying container {container_name} with refreshed credentials")
            return analyze_container(new_blob_service_client, container_name, account_name, subscription_id)
        except Exception as refresh_error:
            logger.error(f"Credential refresh failed: {refresh_error}")
            return None
    except Exception as e:
        logger.error(f"Error processing container {container_name}: {e}")
        return None

def process_containers_concurrently(containers_to_process, max_workers=10):
    """
    Process containers concurrently using ThreadPoolExecutor.
    
    Args:
        containers_to_process (list): List of (blob_service_client, container_name, account_name, subscription_id) tuples
        max_workers (int, optional): Maximum number of concurrent workers. Defaults to 10.
    
    Returns:
        list: Processed container results
    """
    container_results = []
    effective_workers = min(max_workers, len(containers_to_process))
    
    logger.info(f"Starting analysis with {effective_workers} concurrent workers")
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=effective_workers) as executor:
        futures = {
            executor.submit(
                process_container_with_retry, 
                blob_service_client,
                container_name, 
                account_name, 
                subscription_id
            ): (container_name, account_name) 
            for blob_service_client, container_name, account_name, subscription_id in containers_to_process
        }
        
        completed = 0
        total = len(containers_to_process)
        
        for future in concurrent.futures.as_completed(futures):
            container_name, account_name = futures[future]
            completed += 1
            logger.info(f"Progress: {completed}/{total} containers ({(completed/total)*100:.1f}%)")
            
            result = future.result()
            if result:
                container_results.append(result)

    return container_results

# ============================================================================
# COST OPTIMIZATION FUNCTIONS
# ============================================================================

def recommend_azure_tier(size_bytes, days_since_modified):
    """
    Recommend an appropriate Azure Storage access tier based on blob size and age.
    
    Args:
        size_bytes (int): Size of the blob in bytes
        days_since_modified (int): Days since the blob was last modified
    
    Returns:
        str: Recommended access tier
    """
    # Define thresholds
    MIN_COOL_SIZE = 256 * 1024  # 256 KB minimum billable size for Cool tier
    MIN_COLD_SIZE = 256 * 1024  # 256 KB minimum billable size for Cold tier
    
    # If blob is smaller than 256KB, Hot tier is often more cost-effective
    if size_bytes < MIN_COOL_SIZE:
        return "Hot"
    
    # Age-based recommendations
    if days_since_modified > 180:
        # Blobs older than 6 months - rarely accessed
        return "Archive" if size_bytes >= 1024 * 1024 else "Cold"
    elif days_since_modified > 90:
        # Blobs older than 90 days - very infrequently accessed
        return "Cold"
    elif days_since_modified > 30:
        # Blobs older than 30 days - infrequently accessed
        return "Cool"
    else:
        # Recently accessed blobs
        return "Hot"

def generate_cost_optimization_report(container_results):
    """
    Generate a detailed cost optimization report based on container analysis.
    
    Args:
        container_results (list): List of container analysis results
    
    Returns:
        dict: Cost optimization recommendations
    """
    # Initialize recommendations structure
    recommendations = {
        'lifecycle_policies': [],
        'access_tier_migrations': [],
        'small_blobs': [],
        'old_blobs': [],
        'estimated_savings': 0.0
    }
    
    # Pricing data in USD per GB-month (approximate)
    pricing = {
        'Hot': 0.0184,     # General purpose v2 hot tier price
        'Cool': 0.01,      # General purpose v2 cool tier price
        'Cold': 0.005,     # General purpose v2 cold tier price
        'Archive': 0.00099  # Archive tier price
    }
    
    # Analyze each container for optimization opportunities
    for result in container_results:
        container_name = result['container_name']
        account_name = result['account_name']
        total_size_gb = result['total_size'] / (1024**3)  # Convert to GB
        current_monthly_cost = total_size_gb * pricing['Hot']  # Assume all is in Hot tier
        
        # Track potential savings for this container
        container_potential_savings = 0.0
        
        # 1. Check for lifecycle policy opportunities (cold blobs)
        if result['pct_over_90_days'] > 20:  # If more than 20% of blobs are old
            cold_size_gb = result['total_cold_size'] / (1024**3) if 'total_cold_size' in result else 0
            
            # If we don't have cold size from analysis, estimate it
            if cold_size_gb == 0:
                cold_size_gb = total_size_gb * (result['pct_over_90_days'] / 100)
            
            # Calculate potential savings with different tiers based on age
            # For data older than 180 days, recommend Archive tier
            # For data between 90-180 days, recommend Cold tier
            very_old_percent = result.get('pct_over_180_days', result['pct_over_90_days'] / 2)  # Estimate if not available
            very_old_size_gb = cold_size_gb * (very_old_percent / 100) if very_old_percent > 0 else 0
            moderately_old_size_gb = cold_size_gb - very_old_size_gb
            
            hot_cost = cold_size_gb * pricing['Hot']
            archive_cost = very_old_size_gb * pricing['Archive']
            cold_cost = moderately_old_size_gb * pricing['Cold']
            new_cost = archive_cost + cold_cost
            savings = hot_cost - new_cost
            
            if savings > 1.0:  # Only recommend if savings are significant (>$1/month)
                # Create appropriate recommendation based on data distribution
                if very_old_size_gb > moderately_old_size_gb:
                    primary_tier = "Archive"
                    secondary_tier = "Cold"
                else:
                    primary_tier = "Cold"
                    secondary_tier = "Archive"
                
                recommendations['lifecycle_policies'].append({
                    'account': account_name,
                    'container': container_name,
                    'recommendation': f"Implement lifecycle policy to transition blobs older than 90 days to {primary_tier} tier and blobs older than 180 days to {secondary_tier} tier",
                    'affected_blobs': result['not_accessed_90_days'],
                    'affected_size_gb': round(cold_size_gb, 2),
                    'monthly_savings': round(savings, 2)
                })
                container_potential_savings += savings
        
        # 2. Check for small blob consolidation opportunities
        if result['small_blob_count'] > 10000:  # If container has many small blobs
            small_blob_size_gb = result['small_blob_size'] / (1024**3)
            # Savings are harder to estimate here but include reduced transaction costs and better performance
            # Azure charges per transaction, so fewer blobs means fewer GET/LIST operations
            estimated_request_savings = min(result['small_blob_count'] * 0.0000004, 20)  # Cap at $20
            
            if estimated_request_savings > 1.0:
                recommendations['small_blobs'].append({
                    'account': account_name,
                    'container': container_name,
                    'recommendation': f"Consider consolidating {result['small_blob_count']} small blobs into larger blobs",
                    'affected_blobs': result['small_blob_count'],
                    'affected_size_gb': round(small_blob_size_gb, 2),
                    'monthly_savings': round(estimated_request_savings, 2)
                })
                container_potential_savings += estimated_request_savings
        
        # 3. Check for access tier migration opportunities (warm blobs - 30-90 days)
        if 'total_warm_size' in result and result['total_warm_size'] > 0:
            warm_size_gb = result['total_warm_size'] / (1024**3)
            
            # Calculate potential savings with Cool tier
            hot_cost = warm_size_gb * pricing['Hot']
            cool_cost = warm_size_gb * pricing['Cool']
            savings = hot_cost - cool_cost
            
            if savings > 1.0:  # Only recommend if savings are significant
                recommendations['access_tier_migrations'].append({
                    'account': account_name,
                    'container': container_name,
                    'recommendation': f"Migrate infrequently accessed blobs (30-90 days old) to Cool tier",
                    'affected_blobs': result['warm_blobs'],
                    'affected_size_gb': round(warm_size_gb, 2),
                    'monthly_savings': round(savings, 2)
                })
                container_potential_savings += savings
                
        # 4. Check for Cold tier migration opportunities (for 90-180 days data if available)
        if 'total_cold_size' in result and result['total_cold_size'] > 0:
            # Filter out very old data that would go to Archive
            very_old_percent = result.get('pct_over_180_days', 0)
            cold_but_not_archive_size_gb = result['total_cold_size'] * (100 - very_old_percent) / 100
            cold_but_not_archive_size_gb = cold_but_not_archive_size_gb / (1024**3)  # Convert to GB
            
            if cold_but_not_archive_size_gb > 0:
                # Calculate potential savings with Cold tier vs Hot
                hot_cost = cold_but_not_archive_size_gb * pricing['Hot']
                cold_cost = cold_but_not_archive_size_gb * pricing['Cold']
                savings = hot_cost - cold_cost
                
                if savings > 1.0:  # Only recommend if savings are significant
                    affected_blobs = int(result['cold_blobs'] * (100 - very_old_percent) / 100)
                    recommendations['access_tier_migrations'].append({
                        'account': account_name,
                        'container': container_name,
                        'recommendation': f"Migrate rarely accessed blobs (90-180 days old) to Cold tier",
                        'affected_blobs': affected_blobs,
                        'affected_size_gb': round(cold_but_not_archive_size_gb, 2),
                        'monthly_savings': round(savings, 2)
                    })
                    container_potential_savings += savings
        
        # Add container's potential savings to total
        recommendations['estimated_savings'] += container_potential_savings
    
    # Sort recommendations by potential savings
    for category in ['lifecycle_policies', 'access_tier_migrations', 'small_blobs']:
        recommendations[category] = sorted(
            recommendations[category], 
            key=lambda x: x.get('monthly_savings', 0), 
            reverse=True
        )
    
    return recommendations

# ============================================================================
# EXCEL REPORTING FUNCTIONS
# ============================================================================

def _generate_excel_report(container_results, export_detailed_blobs=False, max_blobs_per_container=None):
    """
    Generate an Excel report from container analysis results.
    
    Args:
        container_results (list): List of container analysis results
        export_detailed_blobs (bool, optional): Whether to export detailed blob lists. Defaults to False.
        max_blobs_per_container (int, optional): Maximum number of blobs to export per container. Defaults to None.
    """
    # Initialize workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Azure Storage Analysis"

    # Define styles
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    right_alignment = Alignment(horizontal="right", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    
    warning_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

    # Add headers
    headers = [
        "Storage Account",
        "Container", 
        "Total Size (HR)", 
        "Blobs 0KB-1MB", 
        "Blobs 0KB-1MB Size (HR)",
        "Large Blobs (>1MB)", 
        "Large Blobs Size (HR)", 
        "Total Blobs", 
        "30-90 Days Old", 
        "30-90 Days %", 
        "≥90 Days Old",
        "≥90 Days %"
    ]
    
    sheet.append(headers)
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    # Aggregate totals
    total_size = 0
    total_small_blob_count = 0
    total_small_blob_size = 0
    total_large_blob_count = 0
    total_large_blob_size = 0
    total_blob_count = 0
    total_30_90_days = 0
    total_over_90_days = 0
    total_hot_blobs = 0
    total_warm_blobs = 0
    total_cold_blobs = 0

    # Add data rows
    for idx, result in enumerate(container_results, start=2):
        total_size_hr = human_readable_size(result['total_size'])
        small_blob_size_hr = human_readable_size(result['small_blob_size'])
        large_blob_size_hr = human_readable_size(result['large_blob_size'])
        
        total_size += result['total_size']
        total_small_blob_count += result['small_blob_count']
        total_small_blob_size += result['small_blob_size']
        total_large_blob_count += result['large_blob_count']
        total_large_blob_size += result['large_blob_size']
        total_blob_count += result['total_blob_count']
        total_30_90_days += result['between_30_and_90_days']
        total_over_90_days += result['not_accessed_90_days']
        
        if 'hot_blobs' in result:
            total_hot_blobs += result['hot_blobs']
        if 'warm_blobs' in result:
            total_warm_blobs += result['warm_blobs']
        if 'cold_blobs' in result:
            total_cold_blobs += result['cold_blobs']

        row = [
            result['account_name'],
            result['container_name'],
            f"{total_size_hr[0]} {total_size_hr[1]}",
            result['small_blob_count'],
            f"{small_blob_size_hr[0]} {small_blob_size_hr[1]}",
            result['large_blob_count'],
            f"{large_blob_size_hr[0]} {large_blob_size_hr[1]}",
            result['total_blob_count'],
            result['between_30_and_90_days'],
            result['pct_30_90_days'],  # Store as number, not string with %
            result['not_accessed_90_days'],
            result['pct_over_90_days']  # Store as number, not string with %
        ]
        
        sheet.append(row)
        
        # Apply warning highlighting to rows with high percentages of old objects
        if result['pct_over_90_days'] > 50:
            for col in range(10, 13):  # Columns with age percentages
                cell = sheet.cell(row=idx, column=col)
                cell.fill = warning_fill
                
        # Format percentage cells properly and align all data cells
        for col_num in range(1, len(headers) + 1):
            cell = sheet.cell(row=idx, column=col_num)
            
            # Left align the storage account and container columns
            if col_num in [1, 2]:
                cell.alignment = left_alignment
            # Right align all numeric and data columns 
            else:
                cell.alignment = right_alignment
            
            # Set percentage format for percentage columns
            if col_num == 10 or col_num == 12:
                cell.number_format = '0.0"%"'

    # Add summary row
    total_pct_30_90 = (total_30_90_days / total_blob_count * 100) if total_blob_count > 0 else 0
    total_pct_over_90 = (total_over_90_days / total_blob_count * 100) if total_blob_count > 0 else 0
    total_size_hr = human_readable_size(total_size)
    total_small_size_hr = human_readable_size(total_small_blob_size)
    total_large_size_hr = human_readable_size(total_large_blob_size)
    
    summary_row = [
        "TOTAL",
        "",  # No container for total row
        f"{total_size_hr[0]} {total_size_hr[1]}",
        total_small_blob_count,
        f"{total_small_size_hr[0]} {total_small_size_hr[1]}",
        total_large_blob_count,
        f"{total_large_size_hr[0]} {total_large_size_hr[1]}",
        total_blob_count,
        total_30_90_days,
        total_pct_30_90,  # Store as number, not string with %
        total_over_90_days,
        total_pct_over_90  # Store as number, not string with %
    ]
    
    # Add empty row before summary
    sheet.append([])
    
    # Add summary row with formatting
    summary_row_idx = len(container_results) + 3
    sheet.append(summary_row)
    
    summary_font = Font(bold=True)
    summary_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=summary_row_idx, column=col)
        cell.font = summary_font
        cell.fill = summary_fill
        
        # Left align the "TOTAL" label
        if col == 1:
            cell.alignment = left_alignment
        # Right align all numeric values
        else:
            cell.alignment = right_alignment
        
        # Format percentage cells in summary row
        if col == 10 or col == 12:
            cell.number_format = '0.0"%"'

    # Adjust column widths
    for column in sheet.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = max_length + 2

    # Create a summary by storage account
    if len(set(result.get('account_name', '') for result in container_results)) > 1:
        _generate_account_summary(workbook, container_results)
    
    # Generate cost optimization recommendations
    cost_recommendations = generate_cost_optimization_report(container_results)
    
    # Add cost optimization report
    export_cost_optimization_report_to_excel(workbook, cost_recommendations)

    # Export detailed blob list if requested
    if export_detailed_blobs:
        logger.info("Exporting detailed blob list - this may take some time for large containers...")
        try:
            # Only include container results that have client information
            container_results_with_clients = [r for r in container_results if 'client' in r]
            
            if container_results_with_clients:
                logger.info(f"Beginning detailed export for {len(container_results_with_clients)} containers")
                # Export detailed blobs to the workbook
                total_blobs = export_detailed_blob_list_to_excel(
                    container_results_with_clients, workbook, max_blobs_per_container)
                logger.info(f"Exported details for {total_blobs} blobs")
            else:
                logger.warning("No container clients available for detailed blob export")
        except Exception as e:
            logger.error(f"Error during detailed blob export: {e}")
            import traceback
            logger.error(traceback.format_exc())

    # Save the Excel file
    output_file = f"azure_storage_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    workbook.save(output_file)
    logger.info(f"Analysis complete. Results saved to {output_file}")

    # Log summary statistics
    logger.info(f"Summary Statistics:")
    logger.info(f"Total Containers Analyzed: {len(container_results)}")
    logger.info(f"Total Size Across All Containers: {total_size_hr[0]} {total_size_hr[1]}")
    logger.info(f"Total Blobs Across All Containers: {total_blob_count}")
    logger.info(f"Blobs Between 30-90 Days Old: {total_30_90_days} ({total_pct_30_90:.1f}%)")
    logger.info(f"Blobs Not Modified ≥90 Days: {total_over_90_days} ({total_pct_over_90:.1f}%)")
    
    if container_results and 'hot_blobs' in container_results[0]:
        logger.info(f"Hot Blobs (modified in last 30 days): {total_hot_blobs}")
        logger.info(f"Warm Blobs (modified 30-90 days ago): {total_warm_blobs}")
        logger.info(f"Cold Blobs (not modified in 90+ days): {total_cold_blobs}")
    
    if cost_recommendations:
        logger.info(f"Estimated Monthly Savings: ${cost_recommendations['estimated_savings']:.2f}")

def export_cost_optimization_report_to_excel(workbook, recommendations):
    """
    Export the cost optimization report to an Excel worksheet.
    
    Args:
        workbook (openpyxl.Workbook): Excel workbook
        recommendations (dict): Cost optimization recommendations
    """
    # Create a new sheet for the cost optimization report
    sheet = workbook.create_sheet(title="Cost Optimization")
    
    # Modern presentation styles
    # Title
    sheet.append(["Azure Storage Cost Optimization Recommendations"])
    sheet.merge_cells(f'A1:F1')
    cell = sheet.cell(row=1, column=1)
    cell.font = PresentationStyles.TITLE_FONT
    cell.alignment = PresentationStyles.CENTER
    cell.fill = PresentationStyles.HEADER_FILL
    cell.border = PresentationStyles.BORDER

    # Add summary with total savings
    sheet.append([])
    sheet.append(["Total Estimated Monthly Savings:", f"${recommendations['estimated_savings']:.2f}"])
    cell = sheet.cell(row=3, column=1)
    cell.font = PresentationStyles.SUMMARY_FONT
    cell.fill = PresentationStyles.SUMMARY_FILL
    cell.border = PresentationStyles.BORDER

    savings_cell = sheet.cell(row=3, column=2)
    savings_cell.font = Font(bold=True, color="006100")
    savings_cell.fill = PresentationStyles.SUMMARY_FILL
    savings_cell.border = PresentationStyles.BORDER

    # Add spacing
    sheet.append([])

    # Add sections for each recommendation type
    current_row = 5
    
    # 1. Lifecycle Policies section
    if recommendations['small_blobs']:
        sheet.append(["Small Blobs Consolidation Recommendations"])
        section_cell = sheet.cell(row=current_row, column=1)
        section_cell.font = PresentationStyles.CATEGORY_FONT
        section_cell.fill = PresentationStyles.CATEGORY_FILL
        current_row += 1

        # Add headers
        headers = ["Storage Account", "Container", "Recommendation", "Affected Blobs", "Affected Size (GB)", "Monthly Savings"]
        sheet.append(headers)
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=current_row, column=col_num)
            cell.font = PresentationStyles.HEADER_FONT
            cell.fill = PresentationStyles.HEADER_FILL
            cell.alignment = PresentationStyles.CENTER
        current_row += 1

        # Add recommendations
        for rec in recommendations['small_blobs']:
            row = [
                rec['account'],
                rec['container'],
                rec['recommendation'],
                rec['affected_blobs'],
                rec['affected_size_gb'],
                f"${rec['monthly_savings']:.2f}"
            ]
            sheet.append(row)

            # Style cells
            for col_num in range(1, len(headers) + 1):
                cell = sheet.cell(row=current_row, column=col_num)

                # Align based on content type
                if col_num in [1, 2, 3]:  # Text columns
                    cell.alignment = PresentationStyles.LEFT
                else:  # Numeric columns
                    cell.alignment = PresentationStyles.RIGHT

            current_row += 1

        # Add spacing
        sheet.append([])
        current_row += 1
        # Add spacing
        sheet.append([])
        current_row += 1
    
    # 2. Access Tier Migrations section
    if recommendations['access_tier_migrations']:
        sheet.append(["Access Tier Migration Recommendations"])
        section_cell = sheet.cell(row=current_row, column=1)
        section_cell.font = PresentationStyles.CATEGORY_FONT
        section_cell.fill = PresentationStyles.CATEGORY_FILL
        current_row += 1

        # Add headers
        headers = ["Storage Account", "Container", "Recommendation", "Affected Blobs", "Affected Size (GB)", "Monthly Savings"]
        sheet.append(headers)
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=current_row, column=col_num)
            cell.font = PresentationStyles.HEADER_FONT
            cell.fill = PresentationStyles.HEADER_FILL
            cell.alignment = PresentationStyles.CENTER
        current_row += 1

        # Add recommendations
        for rec in recommendations['access_tier_migrations']:
            row = [
                rec['account'],
                rec['container'],
                rec['recommendation'],
                rec['affected_blobs'],
                rec['affected_size_gb'],
                f"${rec['monthly_savings']:.2f}"
            ]
            sheet.append(row)

            # Style cells
            for col_num in range(1, len(headers) + 1):
                cell = sheet.cell(row=current_row, column=col_num)

                # Align based on content type
                if col_num in [1, 2, 3]:  # Text columns
                    cell.alignment = PresentationStyles.LEFT
                else:  # Numeric columns
                    cell.alignment = PresentationStyles.RIGHT

            current_row += 1

        # Add spacing
        sheet.append([])
        current_row += 1
    
    # 3. Small Blobs Consolidation section
    if recommendations['small_blobs']:
        sheet.append(["Small Blobs Consolidation Recommendations"])
        section_cell = sheet.cell(row=current_row, column=1)
        section_cell.font = PresentationStyles.CATEGORY_FONT
        section_cell.fill = PresentationStyles.CATEGORY_FILL
        current_row += 1
        
        # Add headers
        headers = ["Storage Account", "Container", "Recommendation", "Affected Blobs", "Affected Size (GB)", "Monthly Savings"]
        sheet.append(headers)
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=current_row, column=col_num)
            cell.font = PresentationStyles.HEADER_FONT
            cell.fill = PresentationStyles.HEADER_FILL
            cell.alignment = PresentationStyles.CENTER
        current_row += 1
        
        # Add recommendations
        for rec in recommendations['small_blobs']:
            row = [
                rec['account'],
                rec['container'],
                rec['recommendation'],
                rec['affected_blobs'],
                rec['affected_size_gb'],
                f"${rec['monthly_savings']:.2f}"
            ]
            sheet.append(row)
            
            # Style cells
            for col_num in range(1, len(headers) + 1):
                cell = sheet.cell(row=current_row, column=col_num)
                
                # Align based on content type
                if col_num in [1, 2, 3]:  # Text columns
                    cell.alignment = PresentationStyles.LEFT
                else:  # Numeric columns
                    cell.alignment = PresentationStyles.RIGHT
            
            current_row += 1
    
    # Adjust column widths
    sheet.column_dimensions['A'].width = 25  # Storage Account
    sheet.column_dimensions['B'].width = 25  # Container
    sheet.column_dimensions['C'].width = 50  # Recommendation
    sheet.column_dimensions['D'].width = 18  # Affected Blobs
    sheet.column_dimensions['E'].width = 20  # Affected Size
    sheet.column_dimensions['F'].width = 18  # Monthly Savings
    
    # Add a chart to visualize potential savings
    add_cost_savings_chart(sheet, recommendations)

def add_cost_savings_chart(sheet, recommendations):
    """
    Add a bar chart showing potential cost savings by category.
    
    Args:
        sheet (openpyxl.Worksheet): Excel worksheet
        recommendations (dict): Cost optimization recommendations
    """
    try:
        # Calculate savings by category
        lifecycle_savings = sum(rec.get('monthly_savings', 0) for rec in recommendations.get('lifecycle_policies', []))
        access_tier_savings = sum(rec.get('monthly_savings', 0) for rec in recommendations.get('access_tier_migrations', []))
        small_blobs_savings = sum(rec.get('monthly_savings', 0) for rec in recommendations.get('small_blobs', []))
        
        # Skip chart creation if there are no savings
        if lifecycle_savings == 0 and access_tier_savings == 0 and small_blobs_savings == 0:
            logger.info("No savings data available - skipping cost savings chart")
            return
        
        # Add data for chart
        chart_row = sheet.max_row + 3
        sheet.cell(row=chart_row, column=1).value = "Savings Category"
        sheet.cell(row=chart_row, column=2).value = "Monthly Savings ($)"
        
        sheet.cell(row=chart_row+1, column=1).value = "Lifecycle Policies"
        sheet.cell(row=chart_row+1, column=2).value = lifecycle_savings
        
        sheet.cell(row=chart_row+2, column=1).value = "Access Tier Migrations"
        sheet.cell(row=chart_row+2, column=2).value = access_tier_savings
        
        sheet.cell(row=chart_row+3, column=1).value = "Small Blobs Consolidation"
        sheet.cell(row=chart_row+3, column=2).value = small_blobs_savings
        
        # Create bar chart
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Potential Monthly Savings by Category"
        chart.y_axis.title = "Savings Category"
        chart.x_axis.title = "Monthly Savings ($)"
        
        # Add data to chart
        data = Reference(sheet, min_col=2, min_row=chart_row, max_row=chart_row+3, max_col=2)
        cats = Reference(sheet, min_col=1, min_row=chart_row+1, max_row=chart_row+3)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # Add chart to worksheet
        sheet.add_chart(chart, "H5")
        logger.info("Cost savings chart created successfully")
    except Exception as e:
        logger.warning(f"Error creating cost savings chart: {e}")
        # Continue without chart - better to have a working Excel file with no chart
        # than a corrupted file

def _generate_account_summary(workbook, container_results):
    """
    Generate a summary sheet grouped by storage account.
    
    Args:
        workbook (openpyxl.Workbook): Excel workbook
        container_results (list): List of container analysis results
    """
    # Create a new sheet for account summary
    account_sheet = workbook.create_sheet(title="Account Summary")
    
    # Define styles
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    right_alignment = Alignment(horizontal="right", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    
    account_font = Font(bold=True)
    account_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    # Add headers
    headers = [
        "Storage Account",
        "Subscription ID", 
        "Container Count", 
        "Total Size (HR)", 
        "Total Blobs", 
        "Blobs 30-90 Days Old", 
        "≥90 Days Old Blobs"
    ]
    account_sheet.append(headers)
    for col_num, header in enumerate(headers, start=1):
        cell = account_sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Group containers by account
    account_data = {}
    for result in container_results:
        account_name = result.get('account_name', "Unknown")
        subscription_id = result.get('subscription_id', "Unknown")
        
        if account_name not in account_data:
            account_data[account_name] = {
                'subscription_id': subscription_id,
                'container_count': 0,
                'total_size': 0,
                'total_blobs': 0,
                'blobs_30_90_days': 0,
                'blobs_over_90_days': 0
            }
        
        account_data[account_name]['container_count'] += 1
        account_data[account_name]['total_size'] += result['total_size']
        account_data[account_name]['total_blobs'] += result['total_blob_count']
        account_data[account_name]['blobs_30_90_days'] += result['between_30_and_90_days']
        account_data[account_name]['blobs_over_90_days'] += result['not_accessed_90_days']
    
    # Sort accounts by total size
    sorted_accounts = sorted(account_data.items(), 
                            key=lambda x: x[1]['total_size'], 
                            reverse=True)
    
    # Add account summary rows
    row_idx = 2
    for account_name, data in sorted_accounts:
        total_size_hr = human_readable_size(data['total_size'])
        
        row = [
            account_name,
            data['subscription_id'],
            data['container_count'],
            f"{total_size_hr[0]} {total_size_hr[1]}",
            data['total_blobs'],
            data['blobs_30_90_days'],
            data['blobs_over_90_days']
        ]
        
        account_sheet.append(row)
        
        # Format cells
        for col_num in range(1, len(headers) + 1):
            cell = account_sheet.cell(row=row_idx, column=col_num)
            
            # Left align the account name and subscription ID columns
            if col_num in [1, 2]:
                cell.alignment = left_alignment
                if col_num == 1:
                    cell.font = account_font
            # Right align all numeric columns 
            else:
                cell.alignment = right_alignment
        
        row_idx += 1
    
    # Adjust column widths
    for column in account_sheet.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        account_sheet.column_dimensions[column[0].column_letter].width = max_length + 2
        
    # Add a chart
    pie = PieChart()
    pie.title = "Storage Distribution by Account"
    
    # Create data for the chart (account names and total sizes)
    data_refs = Reference(account_sheet, min_col=4, min_row=2, max_row=row_idx-1)
    labels_refs = Reference(account_sheet, min_col=1, min_row=2, max_row=row_idx-1)
    
    pie.add_data(data_refs, titles_from_data=False)
    pie.set_categories(labels_refs)
    
    # Add the chart to the sheet
    account_sheet.add_chart(pie, "I2")
    
    logger.info(f"Account summary generated with {len(sorted_accounts)} accounts")
    
def export_detailed_blob_list_to_excel(container_results, workbook, max_blobs_per_container=None):
    """
    Export detailed blob list with metadata to a new sheet in the existing Excel workbook.
    
    Args:
        container_results (list): List of container analysis results with client access
        workbook (openpyxl.Workbook): Existing Excel workbook
        max_blobs_per_container (int, optional): Maximum number of blobs to export per container.
            None means no limit. Defaults to None.
    
    Returns:
        int: Total number of blobs exported
    """
    # Configuration constants
    MAX_EXCEL_ROWS = 1000000  # Leave buffer below Excel's 1,048,576 limit
    
    # Modern presentation header styles
    thin_border = PresentationStyles.BORDER
    
    # Create the initial sheet
    main_sheet = workbook.create_sheet(title="Detailed Blob List")
    
    # Define headers
    headers = [
        "Storage Account",
        "Container", 
        "Blob Name",
        "Size (Bytes)",
        "Size (Human Readable)",
        "Access Tier",
        "Last Modified",
        "Days Since Last Modified",
        "Recommended Tier"
    ]
    
    # Add headers to sheet
    main_sheet.append(headers)
    for col_num, header in enumerate(headers, start=1):
        cell = main_sheet.cell(row=1, column=col_num)
        cell.font = PresentationStyles.HEADER_FONT
        cell.fill = PresentationStyles.HEADER_FILL
        cell.alignment = PresentationStyles.CENTER
        cell.border = thin_border
    
    # Set column widths
    column_widths = {
        1: 25,  # Storage Account
        2: 25,  # Container
        3: 50,  # Blob Name
        4: 15,  # Size (Bytes)
        5: 20,  # Size (Human Readable)
        6: 15,  # Access Tier
        7: 20,  # Last Modified
        8: 15,  # Days Since Last Modified
        9: 20   # Recommended Tier
    }
    
    for col_num, width in column_widths.items():
        main_sheet.column_dimensions[get_column_letter(col_num)].width = width
    
    # Freeze the header row
    main_sheet.freeze_panes = "A2"
    
    # Get current date for age calculation
    now = datetime.now(timezone.utc)
    
    # Track statistics
    total_blobs_exported = 0
    sheet_num = 1
    main_sheet_row_idx = 2
    
    # Process each container result
    for container_idx, container_result in enumerate(container_results):
        # Extract container info
        container_name = container_result['container_name']
        account_name = container_result['account_name']
        client = container_result.get('client')
        
        if not client:
            logger.warning(f"Missing blob client for container {container_name} in account {account_name}")
            continue
        
        logger.info(f"Processing blobs in container {container_name} ({container_idx+1}/{len(container_results)})")
        
        try:
            # List blobs in the container
            blobs = list(client.list_blobs())
            
            # Apply max blobs limit if specified
            if max_blobs_per_container and len(blobs) > max_blobs_per_container:
                logger.info(f"Limiting export to {max_blobs_per_container} blobs for container {container_name}")
                blobs = blobs[:max_blobs_per_container]
            
            # Add each blob to the sheet
            for blob in blobs:
                # Extract blob data
                blob_name = blob.name
                size_bytes = blob.size
                size_human_readable = human_readable_size(size_bytes)
                size_human = f"{size_human_readable[0]} {size_human_readable[1]}"
                access_tier = blob.blob_tier if hasattr(blob, 'blob_tier') else "Unknown"
                last_modified = blob.last_modified
                days_since_modified = (now - last_modified).days
                
                # Recommend tier based on size and age
                recommended_tier = recommend_azure_tier(size_bytes, days_since_modified)
                
                # Create the data row
                row = [
                    account_name,
                    container_name,
                    blob_name,
                    size_bytes,
                    size_human,
                    access_tier,
                    last_modified.strftime('%Y-%m-%d %H:%M:%S'),
                    days_since_modified,
                    recommended_tier
                ]
                
                # Check if we're approaching Excel's row limit
                if main_sheet_row_idx >= MAX_EXCEL_ROWS:
                    logger.info(f"Creating additional sheet (sheet {sheet_num + 1})")
                    sheet_num += 1
                    main_sheet = workbook.create_sheet(title=f"Detailed Blob List {sheet_num}")
                    
                    # Add headers to new sheet
                    main_sheet.append(headers)
                    for col_num, header in enumerate(headers, start=1):
                        cell = main_sheet.cell(row=1, column=col_num)
                        cell.font = PresentationStyles.HEADER_FONT
                        cell.fill = PresentationStyles.HEADER_FILL
                        cell.alignment = PresentationStyles.CENTER
                        cell.border = thin_border
                    
                    # Set column widths in new sheet
                    for col_num, width in column_widths.items():
                        main_sheet.column_dimensions[get_column_letter(col_num)].width = width
                    
                    # Reset row index for new sheet
                    main_sheet_row_idx = 2
                
                # Add the row to the current sheet
                main_sheet.append(row)
                
                # Format cells
                for col_num in range(1, len(headers) + 1):
                    cell = main_sheet.cell(row=main_sheet_row_idx, column=col_num)
                    cell.border = thin_border
                    
                    # Apply different alignments based on column type
                    if col_num in [1, 2, 3, 6, 9]:  # Text columns
                        cell.alignment = PresentationStyles.LEFT
                    else:  # Numeric columns
                        cell.alignment = PresentationStyles.RIGHT
                    
                    # Format the date column
                    if col_num == 7:
                        cell.number_format = 'yyyy-mm-dd hh:mm:ss'
                
                # Color coding based on recommended tier changes
                if access_tier != recommended_tier and access_tier != "Unknown":
                    highlight_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    for col_num in range(1, len(headers) + 1):
                        main_sheet.cell(row=main_sheet_row_idx, column=col_num).fill = highlight_fill
                
                main_sheet_row_idx += 1
                total_blobs_exported += 1
                
                # Log progress periodically
                if total_blobs_exported % 10000 == 0:
                    logger.info(f"Processed {total_blobs_exported} blobs so far")
        
        except Exception as e:
            logger.error(f"Error processing container {container_name}: {e}")
            continue
    
    # Apply autofilter to all sheets
    for sheet_idx in range(1, sheet_num + 1):
        sheet_name = "Detailed Blob List" if sheet_idx == 1 else f"Detailed Blob List {sheet_idx}"
        sheet = workbook[sheet_name]
        
        # Apply autofilter if the sheet has data
        if sheet.max_row > 1:
            last_column = get_column_letter(len(headers))
            last_row = sheet.max_row
            sheet.auto_filter.ref = f"A1:{last_column}{last_row}"
    
    logger.info(f"Detailed blob list export complete:")
    logger.info(f"  - Total blobs exported: {total_blobs_exported}")
    logger.info(f"  - Sheets created: {sheet_num}")
    
    # Generate a summary sheet showing the distribution of access tiers
    if total_blobs_exported > 0:
        try:
            # Use the first sheet for the access tier summary
            first_sheet = workbook["Detailed Blob List"]
            create_access_tier_summary(workbook, first_sheet, total_blobs_exported)
        except Exception as e:
            logger.error(f"Error creating access tier summary: {e}")
    
    return total_blobs_exported

def create_access_tier_summary(workbook, detailed_blobs_sheet, total_blobs):
    """
    Generate a summary sheet showing the distribution of access tiers
    based on the detailed blob data.
    
    Args:
        workbook (openpyxl.Workbook): Excel workbook
        detailed_blobs_sheet (openpyxl.Worksheet): Worksheet containing detailed blob data
        total_blobs (int): Total number of blobs exported
    """
    # Create a new sheet for the access tier summary
    sheet = workbook.create_sheet(title="Access Tier Summary")
    
    # Add headers
    headers = ["Access Tier", "Blob Count", "Percentage", "Total Size (Bytes)", "Size (Human Readable)"]
    sheet.append(headers)
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.font = PresentationStyles.HEADER_FONT
        cell.fill = PresentationStyles.HEADER_FILL
        cell.alignment = PresentationStyles.CENTER
        cell.border = PresentationStyles.BORDER
    
    # Gather access tier statistics
    tiers = {}
    
    # Column indices in the detailed blobs sheet (adjust if needed)
    tier_col = 6  # Column F
    size_bytes_col = 4     # Column D
    
    # Initialize the dictionary properly
    for row in range(2, detailed_blobs_sheet.max_row + 1):
        tier = detailed_blobs_sheet.cell(row=row, column=tier_col).value
        if tier not in tiers and tier is not None:
            tiers[tier] = {"count": 0, "size": 0}
    
    # Skip the header row in the detailed sheet
    for row in range(2, detailed_blobs_sheet.max_row + 1):
        tier = detailed_blobs_sheet.cell(row=row, column=tier_col).value
        size_bytes = detailed_blobs_sheet.cell(row=row, column=size_bytes_col).value
        
        if tier and size_bytes and tier in tiers:
            try:
                size_value = int(size_bytes) if isinstance(size_bytes, (int, float)) else 0
                tiers[tier]["count"] += 1
                tiers[tier]["size"] += size_value
            except (ValueError, TypeError):
                logger.warning(f"Invalid size value at row {row}: {size_bytes}")
                # Continue processing other rows
    
    # Sort access tiers by blob count
    sorted_tiers = sorted(tiers.items(), key=lambda x: x[1]["count"], reverse=True)
    
    # Add data rows
    row_idx = 2
    for tier, data in sorted_tiers:
        count = data["count"]
        percentage = (count / total_blobs * 100) if total_blobs > 0 else 0
        size_bytes = data["size"]
        
        try:
            size_hr = human_readable_size(size_bytes)
            size_hr_str = f"{size_hr[0]} {size_hr[1]}"
        except Exception as e:
            logger.warning(f"Error formatting size for {tier}: {e}")
            size_hr_str = "Error"
        
        row = [
            tier,
            count,
            percentage,
            size_bytes,
            size_hr_str
        ]
        sheet.append(row)
        
        # Format cells
        for col_num in range(1, len(headers) + 1):
            cell = sheet.cell(row=row_idx, column=col_num)
            
            # Left align the tier column
            if col_num == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            # Right align all numeric columns
            else:
                cell.alignment = PresentationStyles.RIGHT
            
            # Format percentage column
            if col_num == 3:
                cell.number_format = '0.0"%"'
        
        row_idx += 1
    
    # Adjust column widths
    for column in sheet.columns:
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        sheet.column_dimensions[column[0].column_letter].width = max_length + 2
    
    # Only add pie chart if we have enough data
    if len(sorted_tiers) > 1:
        try:
            # Add pie chart for tier distribution
            pie = PieChart()
            pie.title = "Blob Distribution by Access Tier"
            
            # Create data references - be careful with ranges
            max_row = min(row_idx-1, 20)  # Limit to 20 slices max to avoid Excel errors
            
            if max_row > 1:  # Only add chart if we have data
                data_refs = Reference(sheet, min_col=2, min_row=1, max_row=max_row)
                labels_refs = Reference(sheet, min_col=1, min_row=2, max_row=max_row)
                
                pie.add_data(data_refs, titles_from_data=True)
                pie.set_categories(labels_refs)
                
                # Add chart to sheet
                sheet.add_chart(pie, "H2")
                
                logger.info(f"Access tier summary generated with {len(sorted_tiers)} tiers and chart")
        except Exception as e:
            logger.warning(f"Error creating chart in access tier summary: {e}")
            logger.info(f"Access tier summary generated with {len(sorted_tiers)} tiers (no chart)")
    else:
        logger.info(f"Access tier summary generated with {len(sorted_tiers)} tiers (insufficient data for chart)")
        
def analyze_file_share(share_service_client, share_name, account_name, subscription_id=None):
    """
    Analyze a single Azure File Share and collect statistics.
    
    Args:
        share_service_client (ShareServiceClient): Azure File Share Service client
        share_name (str): Name of the file share to analyze
        account_name (str): Storage account name
        subscription_id (str, optional): Azure subscription ID. Defaults to None.
    
    Returns:
        dict: File share analysis results or None
    """
    try:
        share_client = share_service_client.get_share_client(share_name)
        
        # Get share properties
        share_props = share_client.get_share_properties()
        
        # Initialize variables
        total_size = 0
        total_file_count = 0
        total_directory_count = 0
        
        # Size categories
        small_file_count = 0  # Files <= 1MB
        small_file_size = 0
        medium_file_count = 0  # Files 1MB - 100MB
        medium_file_size = 0
        large_file_count = 0  # Files > 100MB
        large_file_size = 0
        
        # Age categories
        files_30_days = 0
        files_30_90_days = 0
        files_90_180_days = 0
        files_over_180_days = 0
        
        # Define time thresholds
        now = datetime.now(timezone.utc)
        threshold_30_days = now - timedelta(days=30)
        threshold_90_days = now - timedelta(days=90)
        threshold_180_days = now - timedelta(days=180)
        
        # Size thresholds
        SMALL_FILE_SIZE = 1024 * 1024  # 1 MB
        MEDIUM_FILE_SIZE = 100 * 1024 * 1024  # 100 MB
        
        # Recursively analyze directories and files
        def analyze_directory(directory_client, path=""):
            nonlocal total_size, total_file_count, total_directory_count
            nonlocal small_file_count, small_file_size, medium_file_count, medium_file_size
            nonlocal large_file_count, large_file_size
            nonlocal files_30_days, files_30_90_days, files_90_180_days, files_over_180_days
            
            try:
                for item in directory_client.list_directories_and_files():
                    if item['is_directory']:
                        total_directory_count += 1
                        # Recursively analyze subdirectory
                        subdir_client = share_client.get_directory_client(
                            directory_name=f"{path}/{item['name']}" if path else item['name']
                        )
                        analyze_directory(subdir_client, f"{path}/{item['name']}" if path else item['name'])
                    else:
                        # It's a file
                        total_file_count += 1
                        file_size = item.get('size', 0)
                        total_size += file_size
                        
                        # Categorize by size
                        if file_size <= SMALL_FILE_SIZE:
                            small_file_count += 1
                            small_file_size += file_size
                        elif file_size <= MEDIUM_FILE_SIZE:
                            medium_file_count += 1
                            medium_file_size += file_size
                        else:
                            large_file_count += 1
                            large_file_size += file_size
                        
                        # Get file properties for age analysis if last_modified is available
                        if 'last_modified' in item and item['last_modified']:
                            last_modified = item['last_modified']
                            if last_modified >= threshold_30_days:
                                files_30_days += 1
                            elif last_modified >= threshold_90_days:
                                files_30_90_days += 1
                            elif last_modified >= threshold_180_days:
                                files_90_180_days += 1
                            else:
                                files_over_180_days += 1
                                
            except Exception as e:
                logger.warning(f"Error analyzing directory {path}: {e}")
        
        # Start analysis from root
        root_dir = share_client.get_directory_client(directory_name="")
        analyze_directory(root_dir)
        
        # Get share quota and usage
        quota_gb = share_props.quota
        
        # Calculate percentages
        pct_small_files = (small_file_count / total_file_count * 100) if total_file_count > 0 else 0
        pct_medium_files = (medium_file_count / total_file_count * 100) if total_file_count > 0 else 0
        pct_large_files = (large_file_count / total_file_count * 100) if total_file_count > 0 else 0
        
        pct_files_30_days = (files_30_days / total_file_count * 100) if total_file_count > 0 else 0
        pct_files_30_90_days = (files_30_90_days / total_file_count * 100) if total_file_count > 0 else 0
        pct_files_90_180_days = (files_90_180_days / total_file_count * 100) if total_file_count > 0 else 0
        pct_files_over_180_days = (files_over_180_days / total_file_count * 100) if total_file_count > 0 else 0
        
        # Check for snapshots
        snapshots = list(share_client.list_snapshots())
        snapshot_count = len(snapshots)
        
        # Log details
        hr_total_size = human_readable_size(total_size)
        logger.info(f"File Share: {share_name} in account {account_name}")
        logger.info(f"Quota: {quota_gb} GB")
        logger.info(f"Total Size: {hr_total_size[0]} {hr_total_size[1]}")
        logger.info(f"Total Files: {total_file_count}")
        logger.info(f"Total Directories: {total_directory_count}")
        logger.info(f"Snapshots: {snapshot_count}")
        logger.info(f"Small files (≤1MB): {small_file_count} ({pct_small_files:.1f}%)")
        logger.info(f"Medium files (1MB-100MB): {medium_file_count} ({pct_medium_files:.1f}%)")
        logger.info(f"Large files (>100MB): {large_file_count} ({pct_large_files:.1f}%)")
        logger.info("-" * 40)
        
        return {
            'share_name': share_name,
            'account_name': account_name,
            'subscription_id': subscription_id,
            'quota_gb': quota_gb,
            'total_size': total_size,
            'total_file_count': total_file_count,
            'total_directory_count': total_directory_count,
            'small_file_count': small_file_count,
            'small_file_size': small_file_size,
            'medium_file_count': medium_file_count,
            'medium_file_size': medium_file_size,
            'large_file_count': large_file_count,
            'large_file_size': large_file_size,
            'files_30_days': files_30_days,
            'files_30_90_days': files_30_90_days,
            'files_90_180_days': files_90_180_days,
            'files_over_180_days': files_over_180_days,
            'pct_small_files': pct_small_files,
            'pct_medium_files': pct_medium_files,
            'pct_large_files': pct_large_files,
            'pct_files_30_days': pct_files_30_days,
            'pct_files_30_90_days': pct_files_30_90_days,
            'pct_files_90_180_days': pct_files_90_180_days,
            'pct_files_over_180_days': pct_files_over_180_days,
            'snapshot_count': snapshot_count,
            'tier': share_props.tier if hasattr(share_props, 'tier') else 'Standard'
        }
        
    except Exception as e:
        logger.error(f"Error analyzing file share {share_name} in account {account_name}: {e}")
        return None

def select_file_shares_to_process(storage_client, account, auto_mode=False, share_names=None, share_pattern=None, max_shares_per_account=None):
    """
    Allow user to select which file shares to process in a storage account.
    
    Args:
        storage_client (StorageManagementClient): Azure Storage Management client
        account (StorageAccount): Storage account object
        auto_mode (bool, optional): If True, process all shares automatically. Defaults to False.
        share_names (list, optional): Specific share names to process. Defaults to None.
        share_pattern (str, optional): Pattern to match share names. Defaults to None.
        max_shares_per_account (int, optional): Maximum shares to process per account. Defaults to None.
    
    Returns:
        list: List of (share_name, share_service_client) tuples
    """
    try:
        # Get resource group from account ID
        resource_group = account.id.split('/')[4]
        
        # Get connection string
        conn_string = get_storage_account_connection_string(storage_client, resource_group, account.name)
        if not conn_string:
            logger.error(f"Could not get connection string for account {account.name}")
            return []
        
        # Create share service client
        share_service_client = ShareServiceClient.from_connection_string(conn_string)
        
        # List all file shares
        shares = list(share_service_client.list_shares())
        
        if not shares:
            logger.info(f"No file shares found in account {account.name}")
            return []
        
        logger.info(f"Found {len(shares)} file shares in account {account.name}")
        
        # Auto mode or specific filters provided
        if auto_mode or share_names or share_pattern:
            selected_shares = []
            
            if share_names:
                # Filter by specific share names
                for share_name in share_names:
                    matching_shares = [s for s in shares if s.name.lower() == share_name.lower()]
                    if matching_shares:
                        selected_shares.extend([(s.name, share_service_client) for s in matching_shares])
                    else:
                        logger.warning(f"File share '{share_name}' not found in account {account.name}")
                
                if not selected_shares:
                    logger.warning(f"No valid file shares found from specified names in account {account.name}. Processing all shares.")
                    selected_shares = [(share.name, share_service_client) for share in shares]
            
            elif share_pattern:
                # Filter by pattern
                for share in shares:
                    if fnmatch.fnmatch(share.name.lower(), share_pattern.lower()):
                        selected_shares.append((share.name, share_service_client))
                
                if not selected_shares:
                    logger.warning(f"No file shares matched pattern '{share_pattern}' in account {account.name}. Processing all shares.")
                    selected_shares = [(share.name, share_service_client) for share in shares]
            
            else:
                # Process all shares
                selected_shares = [(share.name, share_service_client) for share in shares]
            
            # Apply max_shares_per_account limit
            if max_shares_per_account and len(selected_shares) > max_shares_per_account:
                logger.info(f"Limiting to first {max_shares_per_account} file shares in account {account.name}")
                selected_shares = selected_shares[:max_shares_per_account]
            
            share_names_list = [s[0] for s in selected_shares]
            logger.info(f"Auto mode: Processing {len(selected_shares)} file shares in account {account.name}: {', '.join(share_names_list[:3])}" + 
                       (f" and {len(share_names_list) - 3} more" if len(share_names_list) > 3 else ""))
            return selected_shares
        
        # Interactive mode
        print(f"\nHow would you like to select file shares in account {account.name}?")
        print("1. Process all file shares")
        print("2. Select specific file shares by number")
        print("3. Enter file share name patterns (using wildcards)")
        
        selection_method = 0
        while selection_method not in [1, 2, 3]:
            try:
                selection_method = int(input("Enter your choice (1-3): ").strip())
                if selection_method not in [1, 2, 3]:
                    logger.warning("Invalid choice. Please enter a number between 1 and 3.")
            except ValueError:
                logger.warning("Invalid input. Please enter a number.")
        
        # Process all shares
        if selection_method == 1:
            logger.info(f"Processing all {len(shares)} file shares in account {account.name}")
            return [(share.name, share_service_client) for share in shares]
        
        # Select specific shares by number
        elif selection_method == 2:
            print("\nAvailable file shares:")
            for idx, share in enumerate(shares, start=1):
                # Get share properties
                try:
                    share_client = share_service_client.get_share_client(share.name)
                    props = share_client.get_share_properties()
                    quota = props.quota
                    print(f"{idx}. {share.name} (Quota: {quota} GB)")
                except Exception:
                    print(f"{idx}. {share.name}")
            
            selected_indices = []
            valid_selection = False
            
            while not valid_selection:
                try:
                    selection_input = input("Enter file share numbers to process (comma-separated, e.g., 1,3,5): ").strip()
                    indices = [int(idx.strip()) for idx in selection_input.split(',')]
                    
                    # Validate indices
                    invalid_indices = [idx for idx in indices if idx < 1 or idx > len(shares)]
                    if invalid_indices:
                        logger.warning(f"Invalid share numbers: {invalid_indices}. Please enter values between 1 and {len(shares)}.")
                    else:
                        selected_indices = indices
                        valid_selection = True
                except ValueError:
                    logger.warning("Invalid input. Please enter numbers separated by commas.")
            
            selected_shares = [(shares[idx-1].name, share_service_client) for idx in selected_indices]
            logger.info(f"Processing {len(selected_shares)} selected file shares")
            return selected_shares
        
        # Enter share name patterns
        elif selection_method == 3:
            print("\nEnter file share name patterns to match (e.g., 'prod-*' or '*-backup').")
            print("Use * as a wildcard to match any characters.")
            print("Enter an empty line when finished.")
            
            patterns = []
            while True:
                pattern = input("File share pattern (or empty to finish): ").strip()
                if not pattern:
                    break
                patterns.append(pattern)
            
            if not patterns:
                logger.warning("No patterns provided. Processing all file shares.")
                return [(share.name, share_service_client) for share in shares]
            
            # Match shares against patterns
            matched_shares = []
            for pattern in patterns:
                for share in shares:
                    if fnmatch.fnmatch(share.name.lower(), pattern.lower()):
                        matched_shares.append((share.name, share_service_client))
            
            # Remove duplicates
            unique_shares = list({s[0]: s for s in matched_shares}.values())
            
            if not unique_shares:
                logger.warning("No file shares matched the provided patterns. Processing all shares.")
                return [(share.name, share_service_client) for share in shares]
            
            share_names = [s[0] for s in unique_shares]
            logger.info(f"Found {len(unique_shares)} file shares matching the patterns: {', '.join(share_names[:5])}" + 
                       (f" and {len(share_names) - 5} more" if len(share_names) > 5 else ""))
            return unique_shares
    
    except Exception as e:
        logger.error(f"Error selecting file shares for account {account.name}: {e}")
        return []
      
# ============================================================================
# AZURE FILES EXCEL REPORTING FUNCTIONS
# ============================================================================

def export_file_shares_to_excel(workbook, file_share_results):
    """
    Export file share analysis results to Excel worksheet.
    
    Args:
        workbook (openpyxl.Workbook): Excel workbook
        file_share_results (list): List of file share analysis results
    """
    # Create a new sheet for file shares
    sheet = workbook.create_sheet(title="Azure Files Analysis")
    
    # Define styles
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    right_alignment = Alignment(horizontal="right", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    
    warning_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    
    # Add headers
    headers = [
        "Storage Account",
        "File Share",
        "Tier",
        "Quota (GB)",
        "Total Size (HR)",
        "Usage %",
        "Total Files",
        "Total Directories",
        "Small Files (≤1MB)",
        "Medium Files (1-100MB)",
        "Large Files (>100MB)",
        "Files >180 Days Old",
        "Old Files %",
        "Snapshots"
    ]
    
    sheet.append(headers)
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Aggregate totals
    total_size = 0
    total_quota = 0
    total_files = 0
    total_directories = 0
    total_small_files = 0
    total_medium_files = 0
    total_large_files = 0
    total_old_files = 0
    total_snapshots = 0
    
    # Add data rows
    for idx, result in enumerate(file_share_results, start=2):
        total_size_hr = human_readable_size(result['total_size'])
        usage_pct = (result['total_size'] / (result['quota_gb'] * 1024**3) * 100) if result['quota_gb'] > 0 else 0
        
        total_size += result['total_size']
        total_quota += result['quota_gb']
        total_files += result['total_file_count']
        total_directories += result['total_directory_count']
        total_small_files += result['small_file_count']
        total_medium_files += result['medium_file_count']
        total_large_files += result['large_file_count']
        total_old_files += result['files_over_180_days']
        total_snapshots += result['snapshot_count']
        
        row = [
            result['account_name'],
            result['share_name'],
            result['tier'],
            result['quota_gb'],
            f"{total_size_hr[0]} {total_size_hr[1]}",
            usage_pct,
            result['total_file_count'],
            result['total_directory_count'],
            result['small_file_count'],
            result['medium_file_count'],
            result['large_file_count'],
            result['files_over_180_days'],
            result['pct_files_over_180_days'],
            result['snapshot_count']
        ]
        
        sheet.append(row)
        
        # Apply warning highlighting for high usage or many old files
        if usage_pct > 80:
            cell = sheet.cell(row=idx, column=6)
            cell.fill = warning_fill
        
        if result['pct_files_over_180_days'] > 50:
            cell = sheet.cell(row=idx, column=13)
            cell.fill = warning_fill
        
        # Format cells
        for col_num in range(1, len(headers) + 1):
            cell = sheet.cell(row=idx, column=col_num)
            
            # Left align text columns
            if col_num in [1, 2, 3]:
                cell.alignment = left_alignment
            # Right align numeric columns
            else:
                cell.alignment = right_alignment
            
            # Format percentage columns
            if col_num == 6 or col_num == 13:
                cell.number_format = '0.0"%"'
    
    # Add summary row
    total_usage_pct = (total_size / (total_quota * 1024**3) * 100) if total_quota > 0 else 0
    total_old_files_pct = (total_old_files / total_files * 100) if total_files > 0 else 0
    total_size_hr = human_readable_size(total_size)
    
    summary_row = [
        "TOTAL",
        "",
        "",
        total_quota,
        f"{total_size_hr[0]} {total_size_hr[1]}",
        total_usage_pct,
        total_files,
        total_directories,
        total_small_files,
        total_medium_files,
        total_large_files,
        total_old_files,
        total_old_files_pct,
        total_snapshots
    ]
    
    # Add empty row before summary
    sheet.append([])
    
    # Add summary row with formatting
    summary_row_idx = len(file_share_results) + 3
    sheet.append(summary_row)
    
    summary_font = Font(bold=True)
    summary_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=summary_row_idx, column=col)
        cell.font = summary_font
        cell.fill = summary_fill
        
        # Left align the "TOTAL" label
        if col == 1:
            cell.alignment = left_alignment
        # Right align all numeric values
        else:
            cell.alignment = right_alignment
        
        # Format percentage cells in summary row
        if col == 6 or col == 13:
            cell.number_format = '0.0"%"'
    
    # Adjust column widths
    for column in sheet.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = max_length + 2
    
    # Add a chart showing file share usage
    add_file_share_usage_chart(sheet, file_share_results)
    
    logger.info(f"File share analysis exported with {len(file_share_results)} shares")

def add_file_share_usage_chart(sheet, file_share_results):
    """
    Add a bar chart showing file share usage percentages.
    
    Args:
        sheet (openpyxl.Worksheet): Excel worksheet
        file_share_results (list): File share analysis results
    """
    try:
        # Only create chart if we have data
        if not file_share_results:
            return
        
        # Add data for chart
        chart_row = sheet.max_row + 3
        sheet.cell(row=chart_row, column=1).value = "File Share"
        sheet.cell(row=chart_row, column=2).value = "Usage %"
        
        # Add top 10 shares by usage
        sorted_shares = sorted(file_share_results, 
                             key=lambda x: x['total_size'] / (x['quota_gb'] * 1024**3) if x['quota_gb'] > 0 else 0,
                             reverse=True)[:10]
        
        for i, share in enumerate(sorted_shares):
            usage_pct = (share['total_size'] / (share['quota_gb'] * 1024**3) * 100) if share['quota_gb'] > 0 else 0
            sheet.cell(row=chart_row+i+1, column=1).value = f"{share['share_name'][:20]}"
            sheet.cell(row=chart_row+i+1, column=2).value = usage_pct
        
        # Create bar chart
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Top 10 File Shares by Usage %"
        chart.y_axis.title = "File Share"
        chart.x_axis.title = "Usage %"
        
        # Add data to chart
        data = Reference(sheet, min_col=2, min_row=chart_row, max_row=chart_row+len(sorted_shares), max_col=2)
        cats = Reference(sheet, min_col=1, min_row=chart_row+1, max_row=chart_row+len(sorted_shares))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # Add chart to worksheet
        sheet.add_chart(chart, "P5")
        
    except Exception as e:
        logger.warning(f"Error creating file share usage chart: {e}")

def generate_file_share_optimization_report(file_share_results):
    """
    Generate optimization recommendations for Azure Files.
    
    Args:
        file_share_results (list): List of file share analysis results
    
    Returns:
        dict: File share optimization recommendations
    """
    recommendations = {
        'tier_upgrades': [],
        'tier_downgrades': [],
        'quota_adjustments': [],
        'old_file_cleanup': [],
        'snapshot_cleanup': [],
        'estimated_savings': 0.0
    }
    
    # Pricing per GB per month (approximate)
    pricing = {
        'Standard': 0.06,    # Standard file shares
        'Premium': 0.15,     # Premium SSD file shares
        'TransactionOptimized': 0.0225  # Transaction optimized tier
    }
    
    for result in file_share_results:
        share_name = result['share_name']
        account_name = result['account_name']
        tier = result.get('tier', 'Standard')
        quota_gb = result['quota_gb']
        total_size_gb = result['total_size'] / (1024**3)
        usage_pct = (total_size_gb / quota_gb * 100) if quota_gb > 0 else 0
        
        # 1. Check for tier optimization opportunities
        if tier == 'Premium' and result['large_file_count'] < 10 and total_size_gb < 100:
            # Downgrade to Standard if not many large files and small total size
            current_cost = total_size_gb * pricing['Premium']
            new_cost = total_size_gb * pricing['Standard']
            savings = current_cost - new_cost
            
            if savings > 5:  # Only recommend if savings > $5/month
                recommendations['tier_downgrades'].append({
                    'account': account_name,
                    'share': share_name,
                    'recommendation': 'Consider downgrading from Premium to Standard tier',
                    'reason': 'Low number of large files and small total size',
                    'current_tier': 'Premium',
                    'recommended_tier': 'Standard',
                    'monthly_savings': round(savings, 2)
                })
                recommendations['estimated_savings'] += savings
        
        elif tier == 'Standard' and result['large_file_count'] > 100 and result['pct_large_files'] > 50:
            # Upgrade to Premium for better performance with large files
            recommendations['tier_upgrades'].append({
                'account': account_name,
                'share': share_name,
                'recommendation': 'Consider upgrading to Premium tier for better performance',
                'reason': f"{result['large_file_count']} large files ({result['pct_large_files']:.1f}% of total)",
                'current_tier': 'Standard',
                'recommended_tier': 'Premium',
                'monthly_cost_increase': round(total_size_gb * (pricing['Premium'] - pricing['Standard']), 2)
            })
        
        # 2. Check for quota adjustment opportunities
        if usage_pct < 20 and quota_gb > 100:
            # Recommend reducing quota if very low usage
            recommended_quota = max(int(total_size_gb * 2), 100)  # 2x current usage or 100GB minimum
            if recommended_quota < quota_gb:
                recommendations['quota_adjustments'].append({
                    'account': account_name,
                    'share': share_name,
                    'recommendation': f'Reduce quota from {quota_gb}GB to {recommended_quota}GB',
                    'reason': f'Currently using only {usage_pct:.1f}% of allocated quota',
                    'current_quota_gb': quota_gb,
                    'recommended_quota_gb': recommended_quota,
                    'usage_pct': usage_pct
                })
        
        elif usage_pct > 90:
            # Recommend increasing quota if near limit
            recommended_quota = int(quota_gb * 1.5)
            recommendations['quota_adjustments'].append({
                'account': account_name,
                'share': share_name,
                'recommendation': f'Increase quota from {quota_gb}GB to {recommended_quota}GB',
                'reason': f'Currently using {usage_pct:.1f}% of allocated quota',
                'current_quota_gb': quota_gb,
                'recommended_quota_gb': recommended_quota,
                'usage_pct': usage_pct
            })
        
        # 3. Check for old file cleanup opportunities
        if result['pct_files_over_180_days'] > 30:
            old_files_size_gb = (result['files_over_180_days'] / result['total_file_count']) * total_size_gb if result['total_file_count'] > 0 else 0
            potential_savings = old_files_size_gb * pricing.get(tier, pricing['Standard'])
            
            if potential_savings > 1:  # Only recommend if savings > $1/month
                recommendations['old_file_cleanup'].append({
                    'account': account_name,
                    'share': share_name,
                    'recommendation': 'Archive or delete files not accessed in 180+ days',
                    'affected_files': result['files_over_180_days'],
                    'pct_old_files': result['pct_files_over_180_days'],
                    'estimated_size_gb': round(old_files_size_gb, 2),
                    'monthly_savings': round(potential_savings, 2)
                })
                recommendations['estimated_savings'] += potential_savings
        
        # 4. Check for snapshot cleanup opportunities
        if result['snapshot_count'] > 5:
            # Assume each snapshot uses 10% of share size (conservative estimate)
            snapshot_size_gb = result['snapshot_count'] * total_size_gb * 0.1
            snapshot_cost = snapshot_size_gb * pricing.get(tier, pricing['Standard']) * 0.5  # Snapshots typically cost less
            
            if snapshot_cost > 5:  # Only recommend if cost > $5/month
                recommendations['snapshot_cleanup'].append({
                    'account': account_name,
                    'share': share_name,
                    'recommendation': f'Review and clean up old snapshots ({result["snapshot_count"]} snapshots)',
                    'snapshot_count': result['snapshot_count'],
                    'estimated_cost': round(snapshot_cost, 2)
                })
    
    return recommendations

def export_file_share_optimization_to_excel(workbook, recommendations):
    """
    Export file share optimization recommendations to Excel.
    
    Args:
        workbook (openpyxl.Workbook): Excel workbook
        recommendations (dict): File share optimization recommendations
    """
    # Create a new sheet for file share optimization
    sheet = workbook.create_sheet(title="Files Optimization")
    
    # Modern presentation styles
    # Title
    sheet.append(["Azure Files Optimization Recommendations"])
    sheet.merge_cells('A1:F1')
    cell = sheet.cell(row=1, column=1)
    cell.font = PresentationStyles.TITLE_FONT
    cell.alignment = PresentationStyles.CENTER
    cell.fill = PresentationStyles.HEADER_FILL
    cell.border = PresentationStyles.BORDER

    # Add summary with total savings
    sheet.append([])
    sheet.append(["Total Estimated Monthly Savings:", f"${recommendations['estimated_savings']:.2f}"])
    cell = sheet.cell(row=3, column=1)
    cell.font = PresentationStyles.SUMMARY_FONT
    cell.fill = PresentationStyles.SUMMARY_FILL
    cell.border = PresentationStyles.BORDER

    savings_cell = sheet.cell(row=3, column=2)
    savings_cell.font = Font(bold=True, color="006100")
    savings_cell.fill = PresentationStyles.SUMMARY_FILL
    savings_cell.border = PresentationStyles.BORDER

    # Add spacing
    sheet.append([])

    current_row = 5
    
    # 1. Tier Downgrade Recommendations
    if recommendations['tier_downgrades']:
        sheet.append(["Tier Downgrade Recommendations"])
        section_cell = sheet.cell(row=current_row, column=1)
        section_cell.font = PresentationStyles.SECTION_FONT
        section_cell.fill = PresentationStyles.SECTION_FILL
        section_cell.alignment = PresentationStyles.LEFT
        section_cell.border = PresentationStyles.BORDER
        current_row += 1

        headers = ["Storage Account", "File Share", "Current Tier", "Recommended Tier", "Reason", "Monthly Savings"]
        sheet.append(headers)
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=current_row, column=col_num)
            cell.font = PresentationStyles.HEADER_FONT
            cell.fill = PresentationStyles.HEADER_FILL
            cell.alignment = PresentationStyles.CENTER
            cell.border = PresentationStyles.BORDER
        current_row += 1
        
        for rec in recommendations['tier_downgrades']:
            row = [
                rec['account'],
                rec['share'],
                rec['current_tier'],
                rec['recommended_tier'],
                rec['reason'],
                f"${rec['monthly_savings']:.2f}"
            ]
            sheet.append(row)
            current_row += 1
        
        sheet.append([])
        current_row += 1
    
    # 2. Quota Adjustment Recommendations
    if recommendations['quota_adjustments']:
        sheet.append(["Quota Adjustment Recommendations"])
        section_cell = sheet.cell(row=current_row, column=1)
        section_cell.font = PresentationStyles.CATEGORY_FONT
        section_cell.fill = PresentationStyles.CATEGORY_FILL
        current_row += 1

        headers = ["Storage Account", "File Share", "Current Quota (GB)", "Recommended Quota (GB)", "Usage %", "Reason"]
        sheet.append(headers)
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=current_row, column=col_num)
            cell.font = PresentationStyles.HEADER_FONT
            cell.fill = PresentationStyles.HEADER_FILL
            cell.alignment = PresentationStyles.CENTER
        current_row += 1

        for rec in recommendations['quota_adjustments']:
            row = [
                rec['account'],
                rec['share'],
                rec['current_quota_gb'],
                rec['recommended_quota_gb'],
                f"{rec['usage_pct']:.1f}%",
                rec['reason']
            ]
            sheet.append(row)
            current_row += 1

        sheet.append([])
        current_row += 1
    
    # 3. Old File Cleanup Recommendations
    if recommendations['old_file_cleanup']:
        sheet.append(["Old File Cleanup Recommendations"])
        section_cell = sheet.cell(row=current_row, column=1)
        section_cell.font = PresentationStyles.CATEGORY_FONT
        section_cell.fill = PresentationStyles.CATEGORY_FILL
        current_row += 1

        headers = ["Storage Account", "File Share", "Old Files", "% of Total", "Est. Size (GB)", "Monthly Savings"]
        sheet.append(headers)
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=current_row, column=col_num)
            cell.font = PresentationStyles.HEADER_FONT
            cell.fill = PresentationStyles.HEADER_FILL
            cell.alignment = PresentationStyles.CENTER
        current_row += 1

        for rec in recommendations['old_file_cleanup']:
            row = [
                rec['account'],
                rec['share'],
                rec['affected_files'],
                f"{rec['pct_old_files']:.1f}%",
                rec['estimated_size_gb'],
                f"${rec['monthly_savings']:.2f}"
            ]
            sheet.append(row)
            current_row += 1
    
    # Adjust column widths
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 35
    sheet.column_dimensions['F'].width = 18
    
# ============================================================================
# ENHANCED MAIN ORCHESTRATION FUNCTION WITH AZURE FILES
# ============================================================================

def get_azure_storage_analysis_enhanced(max_workers=10, export_detailed_blobs=False, max_blobs_per_container=None,
                                       analyze_file_shares=True, export_detailed_files=False, max_files_per_share=None,
                                       auto_mode=False, subscription_id=None, account_names=None, account_pattern=None,
                                       container_names=None, container_pattern=None, share_names=None, share_pattern=None,
                                       max_accounts=None, max_containers_per_account=None, max_shares_per_account=None,
                                       analyze_containers=True):
    """
    Analyze Azure Storage accounts including both Blob Storage and Azure Files.
    
    Args:
        max_workers (int, optional): Maximum number of concurrent workers. Defaults to 10.
        export_detailed_blobs (bool, optional): Whether to export detailed blob lists. Defaults to False.
        max_blobs_per_container (int, optional): Maximum number of blobs to export per container. Defaults to None.
        analyze_file_shares (bool, optional): Whether to analyze Azure Files. Defaults to True.
        export_detailed_files (bool, optional): Whether to export detailed file lists. Defaults to False.
        max_files_per_share (int, optional): Maximum number of files to export per share. Defaults to None.
        auto_mode (bool, optional): If True, run automatically without user prompts. Defaults to False.
        subscription_id (str, optional): Specific subscription ID to use. Defaults to None.
        account_names (list, optional): Specific storage account names to process. Defaults to None.
        account_pattern (str, optional): Pattern to match storage account names. Defaults to None.
        container_names (list, optional): Specific container names to process. Defaults to None.
        container_pattern (str, optional): Pattern to match container names. Defaults to None.
        share_names (list, optional): Specific file share names to process. Defaults to None.
        share_pattern (str, optional): Pattern to match file share names. Defaults to None.
        max_accounts (int, optional): Maximum number of accounts to process. Defaults to None.
        max_containers_per_account (int, optional): Maximum containers to process per account. Defaults to None.
        max_shares_per_account (int, optional): Maximum file shares to process per account. Defaults to None.
        analyze_containers (bool, optional): Whether to analyze containers. Defaults to True.
    
    Returns:
        bool: True if analysis completed successfully
    """
    try:
        # Log what we're planning to analyze
        logger.info(f"Analysis configuration:")
        logger.info(f"  - Analyze containers: {analyze_containers}")
        logger.info(f"  - Analyze file shares: {analyze_file_shares}")
        
        # Validate that we have at least one analysis type enabled
        if not analyze_containers and not analyze_file_shares:
            logger.error("Both container and file share analysis are disabled. Nothing to analyze!")
            return False
        
        # Initialize Azure clients
        logger.info("Initializing Azure clients...")
        credential, subscription_id, resource_client, storage_client = initialize_azure_clients(
            subscription_id=subscription_id, auto_mode=auto_mode
        )
        
        # Get all storage accounts
        logger.info("Discovering storage accounts...")
        storage_accounts = get_all_storage_accounts(storage_client)
        
        if not storage_accounts:
            logger.error("No storage accounts found in the subscription.")
            return False
        
        # Select storage accounts to process
        selected_accounts = select_storage_accounts_to_process(
            storage_accounts, 
            auto_mode=auto_mode,
            account_names=account_names,
            account_pattern=account_pattern,
            max_accounts=max_accounts
        )
        
        if not selected_accounts:
            logger.error("No storage accounts selected for processing.")
            return False
        
        # Collect containers and file shares to process
        containers_to_process = []
        file_shares_to_process = []
        
        for account in selected_accounts:
            logger.info(f"Processing storage account: {account.name}")
            
            # Get containers for this account ONLY if container analysis is enabled
            if analyze_containers:
                logger.info(f"Container analysis ENABLED - Selecting containers for storage account: {account.name}")
                try:
                    account_containers = select_containers_to_process(
                        storage_client, 
                        account, 
                        auto_mode=auto_mode,
                        container_names=container_names,
                        container_pattern=container_pattern,
                        max_containers_per_account=max_containers_per_account
                    )
                    
                    # Add subscription_id and account_name to each container tuple
                    for container_name, blob_service_client in account_containers:
                        containers_to_process.append((
                            blob_service_client,
                            container_name,
                            account.name,
                            subscription_id
                        ))
                    
                    logger.info(f"Found {len(account_containers)} containers in account {account.name}")
                except Exception as e:
                    logger.warning(f"Error processing containers for account {account.name}: {e}")
            else:
                logger.info(f"Container analysis DISABLED - Skipping container processing for storage account: {account.name}")
            
            # Get file shares for this account if requested
            if analyze_file_shares:
                logger.info(f"File share analysis ENABLED - Selecting file shares for storage account: {account.name}")
                try:
                    account_shares = select_file_shares_to_process(
                        storage_client,
                        account,
                        auto_mode=auto_mode,
                        share_names=share_names,
                        share_pattern=share_pattern,
                        max_shares_per_account=max_shares_per_account
                    )
                    
                    # Add subscription_id and account_name to each share tuple
                    for share_name, share_service_client in account_shares:
                        file_shares_to_process.append((
                            share_service_client,
                            share_name,
                            account.name,
                            subscription_id
                        ))
                    
                    logger.info(f"Found {len(account_shares)} file shares in account {account.name}")
                except Exception as e:
                    logger.warning(f"Error processing file shares for account {account.name}: {e}")
            else:
                logger.info(f"File share analysis DISABLED - Skipping file share processing for storage account: {account.name}")
        
        # Validate that we have something to analyze
        if not containers_to_process and not file_shares_to_process:
            logger.error("No containers or file shares selected for processing.")
            if not analyze_containers:
                logger.info("Container analysis was disabled")
            if not analyze_file_shares:
                logger.info("File share analysis was disabled")
            return False
        
        logger.info(f"Total containers to analyze: {len(containers_to_process)}")
        logger.info(f"Total file shares to analyze: {len(file_shares_to_process)}")
        
        # Process containers concurrently ONLY if we have containers to process AND container analysis is enabled
        container_results = []
        if containers_to_process and analyze_containers:
            logger.info("Starting container analysis...")
            container_results = process_containers_concurrently(
                containers_to_process, 
                max_workers=max_workers
            )
            logger.info(f"Successfully analyzed {len(container_results)} containers")
        elif not analyze_containers:
            logger.info("Container analysis disabled - skipping blob storage analysis")
        else:
            logger.info("No containers found to analyze")
        
        # Process file shares concurrently ONLY if we have file shares to process AND file share analysis is enabled
        file_share_results = []
        if file_shares_to_process and analyze_file_shares:
            logger.info("Starting file share analysis...")
            file_share_results = process_file_shares_concurrently(
                file_shares_to_process,
                max_workers=max_workers
            )
            logger.info(f"Successfully analyzed {len(file_share_results)} file shares")
        elif not analyze_file_shares:
            logger.info("File share analysis disabled - skipping Azure Files analysis")
        else:
            logger.info("No file shares found to analyze")
        
        # Generate Excel report
        logger.info("Generating Excel report...")
        _generate_enhanced_excel_report(
            container_results,
            file_share_results,
            export_detailed_blobs=export_detailed_blobs,
            max_blobs_per_container=max_blobs_per_container,
            export_detailed_files=export_detailed_files,
            max_files_per_share=max_files_per_share
        )
        
        logger.info("Azure Storage analysis completed successfully!")
        return True
        
    except KeyboardInterrupt:
        logger.info("Analysis interrupted by user")
        return False
    except Exception as e:
        logger.error(f"Error during analysis: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False

def process_file_shares_concurrently(shares_to_process, max_workers=10):
    """
    Process file shares concurrently using ThreadPoolExecutor.
    
    Args:
        shares_to_process (list): List of (share_service_client, share_name, account_name, subscription_id) tuples
        max_workers (int, optional): Maximum number of concurrent workers. Defaults to 10.
    
    Returns:
        list: Processed file share results
    """
    share_results = []
    effective_workers = min(max_workers, len(shares_to_process))
    
    logger.info(f"Starting file share analysis with {effective_workers} concurrent workers")
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=effective_workers) as executor:
        futures = {
            executor.submit(
                analyze_file_share,
                share_service_client,
                share_name,
                account_name,
                subscription_id
            ): (share_name, account_name)
            for share_service_client, share_name, account_name, subscription_id in shares_to_process
        }
        
        completed = 0
        total = len(shares_to_process)
        
        for future in concurrent.futures.as_completed(futures):
            share_name, account_name = futures[future]
            completed += 1
            logger.info(f"Progress: {completed}/{total} file shares ({(completed/total)*100:.1f}%)")
            
            result = future.result()
            if result:
                share_results.append(result)
    
    return share_results

def _generate_enhanced_excel_report(
    container_results, file_share_results,
    export_detailed_blobs=False, max_blobs_per_container=None,
    export_detailed_files=False, max_files_per_share=None):
    # Add watermark/info to all sheets after creation
    watermark_text = "Automation developed by Prashant Kumar, Cloud Engineer @ AHEAD India"
    def add_watermark_to_sheet(sheet):
        from openpyxl.styles import Font, Alignment
        row = sheet.max_row + 2
        cell = sheet.cell(row=row, column=1)
        cell.value = watermark_text
        cell.font = Font(italic=True, color="888888", size=10)
        cell.alignment = Alignment(horizontal="left", vertical="center")
    """
    Generate an enhanced Excel report including both Blob Storage and Azure Files analysis.
    
    Args:
        container_results (list): List of container analysis results
        file_share_results (list): List of file share analysis results
        export_detailed_blobs (bool, optional): Whether to export detailed blob lists. Defaults to False.
        max_blobs_per_container (int, optional): Maximum number of blobs to export per container. Defaults to None.
        export_detailed_files (bool, optional): Whether to export detailed file lists. Defaults to False.
        max_files_per_share (int, optional): Maximum number of files to export per share. Defaults to None.
    """
    # Initialize workbook
    workbook = Workbook()
    
    # Remove default sheet
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    
    # Create Overview sheet
    overview_sheet = workbook.create_sheet(title="Overview", index=0)
    create_overview_sheet(overview_sheet, container_results, file_share_results)
    
    # Generate Blob Storage report if we have container results
    if container_results:
        logger.info("Generating Blob Storage analysis sheets...")
        
        # Create container analysis sheet
        _generate_excel_report_containers_only(workbook, container_results)
        
        # Generate cost optimization recommendations for blobs
        blob_cost_recommendations = generate_cost_optimization_report(container_results)
        export_cost_optimization_report_to_excel(workbook, blob_cost_recommendations)
        
        # Export detailed blob list if requested
        if export_detailed_blobs:
            logger.info("Exporting detailed blob list...")
            container_results_with_clients = [r for r in container_results if 'client' in r]
            if container_results_with_clients:
                total_blobs = export_detailed_blob_list_to_excel(
                    container_results_with_clients, workbook, max_blobs_per_container)
                logger.info(f"Exported details for {total_blobs} blobs")
    
    # Generate Azure Files report if we have file share results
    if file_share_results:
        logger.info("Generating Azure Files analysis sheets...")
        
        # Create file share analysis sheet
        export_file_shares_to_excel(workbook, file_share_results)
        
        # Generate optimization recommendations for file shares
        file_share_recommendations = generate_file_share_optimization_report(file_share_results)
        export_file_share_optimization_to_excel(workbook, file_share_recommendations)
    
    # Save the Excel file

    import os
    output_file = os.path.abspath(f"azure_storage_analysis_enhanced_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    # Add watermark to all sheets before saving
    for sheet in workbook.worksheets:
        add_watermark_to_sheet(sheet)
    workbook.save(output_file)
    logger.info(f"Analysis complete. Results saved to {output_file}")
    print(f"\nExcel file saved at: {output_file}\n")

    # Export to CSV as well
    export_analysis_to_csv(container_results, file_share_results)

    # Log summary statistics
    logger.info("=" * 60)
    logger.info("SUMMARY STATISTICS")
    logger.info("=" * 60)

    if container_results:
        total_blob_size = sum(r['total_size'] for r in container_results)
        total_blobs = sum(r['total_blob_count'] for r in container_results)
        size_hr = human_readable_size(total_blob_size)
        logger.info(f"Blob Storage:")
        logger.info(f"  - Containers analyzed: {len(container_results)}")
        logger.info(f"  - Total size: {size_hr[0]} {size_hr[1]}")
        logger.info(f"  - Total blobs: {total_blobs:,}")

    if file_share_results:
        total_file_size = sum(r['total_size'] for r in file_share_results)
        total_files = sum(r['total_file_count'] for r in file_share_results)
        size_hr = human_readable_size(total_file_size)
        logger.info(f"Azure Files:")
        logger.info(f"  - File shares analyzed: {len(file_share_results)}")
        logger.info(f"  - Total size: {size_hr[0]} {size_hr[1]}")
        logger.info(f"  - Total files: {total_files:,}")

def export_analysis_to_csv(container_results, file_share_results):
    """
    Export both Blob Storage and Azure Files results to a clean, well-formatted CSV file.
    """
    import csv
    from datetime import datetime
    
    import os
    output_file = os.path.abspath(f"azure_storage_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")
    with open(output_file, mode='w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)

        # Add watermark at the top of the CSV
        watermark_text = "Automation developed by Prashant Kumar, Cloud Engineer @ AHEAD India"
        writer.writerow([watermark_text])
        writer.writerow([])

        # Write Blob Storage section
        if container_results:
            writer.writerow(["Blob Storage Analysis"])
            headers = [
                "Storage Account", "Container", "Total Size (HR)", "Blobs 0KB-1MB", "Blobs 0KB-1MB Size (HR)",
                "Large Blobs (>1MB)", "Large Blobs Size (HR)", "Total Blobs", "30-90 Days Old", "30-90 Days %",
                "≥90 Days Old", "≥90 Days %"
            ]
            writer.writerow(headers)
            for r in container_results:
                total_size_hr = human_readable_size(r['total_size'])
                small_blob_size_hr = human_readable_size(r['small_blob_size'])
                large_blob_size_hr = human_readable_size(r['large_blob_size'])
                row = [
                    r['account_name'],
                    r['container_name'],
                    f"{total_size_hr[0]} {total_size_hr[1]}",
                    r['small_blob_count'],
                    f"{small_blob_size_hr[0]} {small_blob_size_hr[1]}",
                    r['large_blob_count'],
                    f"{large_blob_size_hr[0]} {large_blob_size_hr[1]}",
                    r['total_blob_count'],
                    r['between_30_and_90_days'],
                    f"{r['pct_30_90_days']:.1f}%",
                    r['not_accessed_90_days'],
                    f"{r['pct_over_90_days']:.1f}%"
                ]
                writer.writerow(row)
            writer.writerow([])

        # Write Azure Files section
        if file_share_results:
            writer.writerow(["Azure Files Analysis"])
            headers = [
                "Storage Account", "File Share", "Tier", "Quota (GB)", "Total Size (HR)", "Usage %",
                "Total Files", "Total Directories", "Small Files (≤1MB)", "Medium Files (1-100MB)",
                "Large Files (>100MB)", "Files >180 Days Old", "Old Files %", "Snapshots"
            ]
            writer.writerow(headers)
            for r in file_share_results:
                total_size_hr = human_readable_size(r['total_size'])
                usage_pct = (r['total_size'] / (r['quota_gb'] * 1024**3) * 100) if r['quota_gb'] > 0 else 0
                row = [
                    r['account_name'],
                    r['share_name'],
                    r.get('tier', ''),
                    r['quota_gb'],
                    f"{total_size_hr[0]} {total_size_hr[1]}",
                    f"{usage_pct:.1f}%",
                    r['total_file_count'],
                    r['total_directory_count'],
                    r['small_file_count'],
                    r['medium_file_count'],
                    r['large_file_count'],
                    r['files_over_180_days'],
                    f"{r['pct_files_over_180_days']:.1f}%",
                    r['snapshot_count']
                ]
                writer.writerow(row)
            writer.writerow([])

    logger.info(f"CSV analysis results saved to {output_file}")
    print(f"CSV file saved at: {output_file}\n")

def create_overview_sheet(sheet, container_results, file_share_results):
    """
    Create an overview sheet summarizing both Blob Storage and Azure Files.
    
    Args:
        sheet: Excel worksheet
        container_results: List of container analysis results
        file_share_results: List of file share analysis results
    """
    # Title
    sheet.append(["Azure Storage Analysis Overview"])
    sheet.merge_cells('A1:D1')
    cell = sheet.cell(row=1, column=1)
    cell.font = PresentationStyles.TITLE_FONT
    cell.alignment = PresentationStyles.CENTER
    cell.fill = PresentationStyles.HEADER_FILL
    cell.border = PresentationStyles.BORDER

    # Generated timestamp
    sheet.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    sheet.cell(row=2, column=1).font = Font(name="Segoe UI", italic=True, color="888888", size=10)
    sheet.append([])

    current_row = 4

    # Blob Storage Summary
    if container_results:
        sheet.append(["Blob Storage Summary"])
        cell = sheet.cell(row=current_row, column=1)
        cell.font = PresentationStyles.SECTION_FONT
        cell.fill = PresentationStyles.SECTION_FILL
        cell.alignment = PresentationStyles.LEFT
        cell.border = PresentationStyles.BORDER
        current_row += 1

        total_blob_size = sum(r['total_size'] for r in container_results)
        total_blobs = sum(r['total_blob_count'] for r in container_results)
        size_hr = human_readable_size(total_blob_size)

        # Metrics table header
        sheet.append(["Metric", "Value"])
        for col in range(1, 3):
            hcell = sheet.cell(row=current_row, column=col)
            hcell.font = PresentationStyles.HEADER_FONT
            hcell.fill = PresentationStyles.HEADER_FILL
            hcell.alignment = PresentationStyles.CENTER
            hcell.border = PresentationStyles.BORDER
        current_row += 1

        # Metrics rows
        metrics = [
            ("Total Containers", len(container_results)),
            ("Total Size", f"{size_hr[0]} {size_hr[1]}"),
            ("Total Blobs", f"{total_blobs:,}"),
            ("Average Blobs per Container", f"{total_blobs // len(container_results):,}" if container_results else "0")
        ]
        for metric, value in metrics:
            sheet.append([metric, value])
            for col in range(1, 3):
                dcell = sheet.cell(row=current_row, column=col)
                dcell.font = Font(name="Segoe UI", size=11)
                dcell.alignment = PresentationStyles.LEFT if col == 1 else PresentationStyles.RIGHT
                dcell.border = PresentationStyles.BORDER
            current_row += 1

        sheet.append([])
        current_row += 1

    # Azure Files Summary
    if file_share_results:
        sheet.append(["Azure Files Summary"])
        cell = sheet.cell(row=current_row, column=1)
        cell.font = PresentationStyles.SECTION_FONT
        cell.fill = PresentationStyles.SECTION_FILL
        cell.alignment = PresentationStyles.LEFT
        cell.border = PresentationStyles.BORDER
        current_row += 1

        total_file_size = sum(r['total_size'] for r in file_share_results)
        total_files = sum(r['total_file_count'] for r in file_share_results)
        total_quota = sum(r['quota_gb'] for r in file_share_results)
        size_hr = human_readable_size(total_file_size)

        # Metrics table header
        sheet.append(["Metric", "Value"])
        for col in range(1, 3):
            hcell = sheet.cell(row=current_row, column=col)
            hcell.font = PresentationStyles.HEADER_FONT
            hcell.fill = PresentationStyles.HEADER_FILL
            hcell.alignment = PresentationStyles.CENTER
            hcell.border = PresentationStyles.BORDER
        current_row += 1

        # Metrics rows
        metrics = [
            ("Total File Shares", len(file_share_results)),
            ("Total Quota", f"{total_quota:,} GB"),
            ("Total Size", f"{size_hr[0]} {size_hr[1]}"),
            ("Total Files", f"{total_files:,}"),
            ("Average Files per Share", f"{total_files // len(file_share_results):,}" if file_share_results else "0"),
            ("Overall Usage %", f"{(total_file_size / (total_quota * 1024**3) * 100):.1f}%" if total_quota > 0 else "0%")
        ]
        for metric, value in metrics:
            sheet.append([metric, value])
            for col in range(1, 3):
                dcell = sheet.cell(row=current_row, column=col)
                dcell.font = Font(name="Segoe UI", size=11)
                dcell.alignment = PresentationStyles.LEFT if col == 1 else PresentationStyles.RIGHT
                dcell.border = PresentationStyles.BORDER
            current_row += 1

    # Adjust column widths
    sheet.column_dimensions['A'].width = 32
    sheet.column_dimensions['B'].width = 28
    
# ============================================================================
# ENHANCED MAIN ENTRY POINT WITH AZURE FILES SUPPORT
# ============================================================================

def main():
    """
    Main entry point for the enhanced Azure Storage Analysis Tool.
    """
    parser = argparse.ArgumentParser(
        description="Analyze Azure Storage accounts (Blob Storage and Azure Files)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Analyze both Blob Storage and Azure Files
  python azure_storage_analysis.py --auto
  
  # Analyze only Blob Storage
  python azure_storage_analysis.py --auto --no-file-shares
  
  # Analyze only Azure Files
  python azure_storage_analysis.py --auto --no-containers
  
  # Analyze specific accounts with pattern matching
  python azure_storage_analysis.py --auto --account-pattern "prod-*"
  
  # Export detailed information with limits
  python azure_storage_analysis.py --auto --export-detailed-blobs --max-blobs-per-container 1000
  
  # Analyze specific file shares
  python azure_storage_analysis.py --auto --share-names myshare1 myshare2
        """
    )
    
    # General options
    parser.add_argument("--auto", action="store_true", help="Run in automatic mode without prompts")
    parser.add_argument("--max-workers", type=int, default=10, help="Maximum number of concurrent workers")
    
    # Blob Storage options
    parser.add_argument("--no-containers", action="store_true", help="Skip Blob Storage analysis")
    parser.add_argument("--export-detailed-blobs", action="store_true", help="Export detailed blob lists")
    parser.add_argument("--max-blobs-per-container", type=int, help="Maximum blobs to export per container")
    parser.add_argument("--container-names", nargs="+", help="Specific container names to process")
    parser.add_argument("--container-pattern", help="Pattern to match container names")
    parser.add_argument("--max-containers-per-account", type=int, help="Maximum containers per account")
    
    # Azure Files options
    parser.add_argument("--no-file-shares", action="store_true", help="Skip Azure Files analysis")
    parser.add_argument("--export-detailed-files", action="store_true", help="Export detailed file lists")
    parser.add_argument("--max-files-per-share", type=int, help="Maximum files to export per share")
    parser.add_argument("--share-names", nargs="+", help="Specific file share names to process")
    parser.add_argument("--share-pattern", help="Pattern to match file share names")
    parser.add_argument("--max-shares-per-account", type=int, help="Maximum file shares per account")
    
    # Account selection options
    parser.add_argument("--subscription-id", help="Specific subscription ID to use")
    parser.add_argument("--account-names", nargs="+", help="Specific storage account names to process")
    parser.add_argument("--account-pattern", help="Pattern to match storage account names")
    parser.add_argument("--max-accounts", type=int, help="Maximum number of accounts to process")
    
    args = parser.parse_args()
    
    # Display banner
    print("\n" + "="*80)
    print(" AZURE STORAGE ANALYSIS TOOL v3.0 ".center(80, "="))
    print(" Enhanced with Azure Files Support ".center(80, "="))
    print("="*80)
    print("This tool analyzes Azure Storage accounts including:")
    if not args.no_containers:
        print("  - Blob Storage containers and blobs")
    if not args.no_file_shares:
        print("  - Azure Files shares and files")
    print("  - Cost optimization recommendations")
    print("  - Detailed usage reports")
    print("="*80)
    
    # Determine what to analyze
    analyze_containers = not args.no_containers
    analyze_file_shares = not args.no_file_shares
    
    if not analyze_containers and not analyze_file_shares:
        print("\nError: Both --no-containers and --no-file-shares specified.")
        print("Nothing to analyze!")
        sys.exit(1)
    
    # Log what will be analyzed
    if args.auto:
        if analyze_containers and analyze_file_shares:
            print("\nAuto mode: Will analyze both Blob Storage and Azure Files")
        elif analyze_containers:
            print("\nAuto mode: Will analyze only Blob Storage (Azure Files disabled)")
        elif analyze_file_shares:
            print("\nAuto mode: Will analyze only Azure Files (Blob Storage disabled)")
    
    # Run the analysis
    success = get_azure_storage_analysis_enhanced(
        max_workers=args.max_workers,
        # Blob Storage options
        export_detailed_blobs=args.export_detailed_blobs,
        max_blobs_per_container=args.max_blobs_per_container,
        # Azure Files options
        analyze_file_shares=analyze_file_shares,
        export_detailed_files=args.export_detailed_files,
        max_files_per_share=args.max_files_per_share,
        # General options
        auto_mode=args.auto,
        subscription_id=args.subscription_id,
        account_names=args.account_names,
        account_pattern=args.account_pattern,
        # Container options - CRITICAL: Pass analyze_containers parameter
        analyze_containers=analyze_containers,
        container_names=args.container_names if analyze_containers else None,
        container_pattern=args.container_pattern if analyze_containers else None,
        max_containers_per_account=args.max_containers_per_account if analyze_containers else None,
        # File share options
        share_names=args.share_names if analyze_file_shares else None,
        share_pattern=args.share_pattern if analyze_file_shares else None,
        max_shares_per_account=args.max_shares_per_account if analyze_file_shares else None,
        # Account limit
        max_accounts=args.max_accounts
    )
    
    if success:
        print("\n" + "="*80)
        print(" ANALYSIS COMPLETED SUCCESSFULLY ".center(80, "="))
        print("="*80)
        sys.exit(0)
    else:
        print("\n" + "="*80)
        print(" ANALYSIS FAILED ".center(80, "="))
        print("="*80)
        sys.exit(1)

# Additional helper function for _generate_excel_report_containers_only
def _generate_excel_report_containers_only(workbook, container_results):
    sheet = workbook.active
    headers = [
        "Storage Account",
        "Container",
        "Total Size (HR)",
        "Blobs 0KB-1MB",
        "Blobs 0KB-1MB Size (HR)",
        "Large Blobs (>1MB)",
        "Large Blobs Size (HR)",
        "Total Blobs",
        "30-90 Days Old",
        "30-90 Days %",
        "≥90 Days Old",
        "≥90 Days %"
    ]
    total_size = 0
    total_small_blob_count = 0
    total_small_blob_size = 0
    total_large_blob_count = 0
    total_large_blob_size = 0
    total_blob_count = 0
    total_30_90_days = 0
    total_over_90_days = 0

    # Write headers
    sheet.append(headers)
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.font = PresentationStyles.HEADER_FONT
        cell.fill = PresentationStyles.HEADER_FILL
        cell.alignment = PresentationStyles.CENTER

    # Add data rows
    for idx, result in enumerate(container_results, start=2):
        total_size_hr = human_readable_size(result['total_size'])
        small_blob_size_hr = human_readable_size(result['small_blob_size'])
        large_blob_size_hr = human_readable_size(result['large_blob_size'])

        total_size += result['total_size']
        total_small_blob_count += result['small_blob_count']
        total_small_blob_size += result['small_blob_size']
        total_large_blob_count += result['large_blob_count']
        total_large_blob_size += result['large_blob_size']
        total_blob_count += result['total_blob_count']
        total_30_90_days += result['between_30_and_90_days']
        total_over_90_days += result['not_accessed_90_days']

        row = [
            result['account_name'],
            result['container_name'],
            f"{total_size_hr[0]} {total_size_hr[1]}",
            result['small_blob_count'],
            f"{small_blob_size_hr[0]} {small_blob_size_hr[1]}",
            result['large_blob_count'],
            f"{large_blob_size_hr[0]} {large_blob_size_hr[1]}",
            result['total_blob_count'],
            result['between_30_and_90_days'],
            result['pct_30_90_days'],
            result['not_accessed_90_days'],
            result['pct_over_90_days']
        ]
        sheet.append(row)

        # Apply warning highlighting to rows with high percentages of old objects
        if result['pct_over_90_days'] > 50:
            for col in range(10, 13):
                cell = sheet.cell(row=idx, column=col)
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

        # Format percentage cells properly and align all data cells
        for col_num in range(1, len(headers) + 1):
            cell = sheet.cell(row=idx, column=col_num)
            if col_num in [1, 2]:
                cell.alignment = PresentationStyles.LEFT
            else:
                cell.alignment = PresentationStyles.RIGHT
            if col_num == 10 or col_num == 12:
                cell.number_format = '0.0"%"'

    # Add summary row
    total_pct_30_90 = (total_30_90_days / total_blob_count * 100) if total_blob_count > 0 else 0
    total_pct_over_90 = (total_over_90_days / total_blob_count * 100) if total_blob_count > 0 else 0
    total_size_hr = human_readable_size(total_size)
    total_small_size_hr = human_readable_size(total_small_blob_size)
    total_large_size_hr = human_readable_size(total_large_blob_size)

    summary_row = [
        "TOTAL",
        "",
        f"{total_size_hr[0]} {total_size_hr[1]}",
        total_small_blob_count,
        f"{total_small_size_hr[0]} {total_small_size_hr[1]}",
        total_large_blob_count,
        f"{total_large_size_hr[0]} {total_large_size_hr[1]}",
        total_blob_count,
        total_30_90_days,
        total_pct_30_90,
        total_over_90_days,
        total_pct_over_90
    ]

    # Add empty row before summary
    sheet.append([])

    # Add summary row with formatting
    summary_row_idx = len(container_results) + 3
    sheet.append(summary_row)

    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=summary_row_idx, column=col)
        cell.font = PresentationStyles.SUMMARY_FONT
        cell.fill = PresentationStyles.SUMMARY_FILL
        if col == 1:
            cell.alignment = PresentationStyles.LEFT
        else:
            cell.alignment = PresentationStyles.RIGHT
        if col == 10 or col == 12:
            cell.number_format = '0.0"%"'
    """
    Generate container analysis sheet only (without creating a new workbook).
    This is used when we have both containers and file shares.
    
    Args:
        workbook (openpyxl.Workbook): Existing Excel workbook
        container_results (list): List of container analysis results
    """
    sheet = workbook.create_sheet(title="Blob Storage Analysis")

    warning_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

    headers = [
        "Storage Account",
        "Container", 
        "Total Size (HR)", 
        "Blobs 0KB-1MB", 
        "Blobs 0KB-1MB Size (HR)",
        "Large Blobs (>1MB)", 
        "Large Blobs Size (HR)", 
        "Total Blobs", 
        "30-90 Days Old", 
        "30-90 Days %", 
        "≥90 Days Old",
        "≥90 Days %"
    ]
    sheet.append(headers)
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.font = PresentationStyles.HEADER_FONT
        cell.fill = PresentationStyles.HEADER_FILL
        cell.alignment = PresentationStyles.CENTER
        cell.border = PresentationStyles.BORDER

    warning_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

    headers = [
        "Storage Account",
        "File Share",
        "Tier",
        "Quota (GB)",
        "Total Size (HR)",
        "Usage %",
        "Total Files",
        "Total Directories",
        "Small Files (≤1MB)",
        "Medium Files (1-100MB)",
        "Large Files (>100MB)",
        "Files >180 Days Old",
        "Old Files %",
        "Snapshots"
    ]
    sheet.append(headers)
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.font = PresentationStyles.HEADER_FONT
        cell.fill = PresentationStyles.HEADER_FILL
        cell.alignment = PresentationStyles.CENTER
        cell.border = PresentationStyles.BORDER

    # Add summary row
    total_pct_30_90 = (total_30_90_days / total_blob_count * 100) if total_blob_count > 0 else 0
    total_pct_over_90 = (total_over_90_days / total_blob_count * 100) if total_blob_count > 0 else 0
    total_size_hr = human_readable_size(total_size)
    total_small_size_hr = human_readable_size(total_small_blob_size)
    total_large_size_hr = human_readable_size(total_large_blob_size)
    
    summary_row = [
        "TOTAL",
        "",
        f"{total_size_hr[0]} {total_size_hr[1]}",
        total_small_blob_count,
        f"{total_small_size_hr[0]} {total_small_size_hr[1]}",
        total_large_blob_count,
        f"{total_large_size_hr[0]} {total_large_size_hr[1]}",
        total_blob_count,
        total_30_90_days,
        total_pct_30_90,
        total_over_90_days,
        total_pct_over_90
    ]
    
    # Add empty row before summary
    sheet.append([])
    
    # Add summary row with formatting
    summary_row_idx = len(container_results) + 3
    sheet.append(summary_row)
    
    summary_font = Font(bold=True)
    summary_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=summary_row_idx, column=col)
        cell.font = summary_font
        cell.fill = summary_fill
        
        if col == 1:
            cell.alignment = PresentationStyles.LEFT
        else:
            cell.alignment = PresentationStyles.RIGHT
        
        if col == 10 or col == 12:
            cell.number_format = '0.0"%"'

    # Adjust column widths
    for column in sheet.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = max_length + 2

if __name__ == "__main__":
    main()